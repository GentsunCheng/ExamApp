import importlib
import os
import tempfile
import unittest

from openpyxl import Workbook, load_workbook


class ProgressExcelTest(unittest.TestCase):
    def test_import_progress_from_excel(self):
        with tempfile.TemporaryDirectory() as td:
            os.environ['HOME'] = td
            import database
            import models
            from views.admin_modules import study_progress_module as progress_excel

            importlib.reload(database)
            importlib.reload(models)
            importlib.reload(progress_excel)

            wb = Workbook()
            ws = wb.active
            ws.title = '模块A'
            ws.append(['任务名', '描述', '顺序'])
            ws.append(['任务1', '描述1', 1])
            ws.append(['任务2', '描述2', 2])
            ws2 = wb.create_sheet('模块B')
            ws2.append(['任务名', '描述', '顺序'])
            ws2.append(['任务3', '描述3', 1])
            ws3 = wb.create_sheet('说明')
            ws3.append(['x'])
            in_path = os.path.join(td, 'import.xlsx')
            wb.save(in_path)

            summary = progress_excel.import_progress_from_excel(in_path, replace=False)
            self.assertEqual(summary['modules'], 2)
            self.assertEqual(summary['tasks'], 3)

            modules = models.list_progress_modules()
            self.assertEqual([m[1] for m in modules], ['模块A', '模块B'])

    def test_export_progress_template(self):
        with tempfile.TemporaryDirectory() as td:
            out = os.path.join(td, 'tpl.xlsx')
            from views.admin_modules import study_progress_module as progress_excel
            out2 = progress_excel.export_progress_template(out)
            wb = load_workbook(out2)
            self.assertIn('说明', wb.sheetnames)
            self.assertIn('示例模块', wb.sheetnames)

    def test_export_user_progress_to_excel(self):
        with tempfile.TemporaryDirectory() as td:
            os.environ['HOME'] = td
            import database
            import models
            from views.admin_modules import study_progress_module as progress_excel

            importlib.reload(database)
            importlib.reload(models)
            importlib.reload(progress_excel)

            models.create_user('u_excel', 'pw', role='user', active=1, full_name='U')
            user_id = next(u for u in models.list_users() if u[1] == 'u_excel')[0]

            module_id = models.upsert_progress_module('模块A')
            t1 = models.upsert_progress_task(module_id, '任务1', '描述1', 1)
            models.set_user_task_progress(user_id, t1, models.PROGRESS_STATUS_IN_PROGRESS, updated_by='admin:1')

            out = os.path.join(td, 'user_progress.xlsx')
            out2 = progress_excel.export_user_progress_to_excel(user_id, out)
            wb = load_workbook(out2)
            self.assertIn('模块A', wb.sheetnames)


if __name__ == '__main__':
    unittest.main()
