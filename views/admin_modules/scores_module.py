from PySide6.QtCore import Qt
from PySide6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QPushButton, QTableWidget, QTableWidgetItem, QFileDialog
from icon_manager import IconManager
from theme_manager import theme_manager
from language import tr
from utils import show_info, show_warn
from models import list_attempts_with_user
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from models import get_exam_title
from windows.score_detail_window import ScoreDetailWindow
import os
import pathlib


class AdminScoresModule(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.icon_manager = IconManager()
        lay = QVBoxLayout()
        gb = QGroupBox(tr('scores.group'))
        vb = QVBoxLayout()
        self.scores_table = QTableWidget(0, 8)
        self.scores_table.setHorizontalHeaderLabels([tr('scores.headers.uuid'), tr('scores.headers.username'), tr('scores.headers.full_name'), tr('scores.headers.user_id'), tr('scores.headers.exam_title'), tr('scores.headers.started'), tr('scores.headers.submitted'), tr('scores.headers.score_total_status')])
        self.scores_table.horizontalHeader().setStretchLastSection(True)
        self.scores_table.setColumnWidth(0, 280)
        self.scores_table.setColumnWidth(1, 75)
        self.scores_table.setColumnWidth(2, 75)
        self.scores_table.setColumnWidth(3, 50)
        self.scores_table.setColumnWidth(4, 250)
        self.scores_table.setColumnWidth(5, 200)
        self.scores_table.setColumnWidth(6, 200)
        self.scores_table.setAlternatingRowColors(True)
        self.scores_table.setShowGrid(False)
        self.scores_table.cellDoubleClicked.connect(self.on_item_double_clicked)
        self.refresh_scores()
        vb.addWidget(self.scores_table)
        hb = QHBoxLayout()
        btn_export_scores = QPushButton(tr('scores.export_excel'))
        btn_export_scores.setIcon(self.icon_manager.get_icon('exam_export'))
        btn_export_scores.clicked.connect(self.export_scores_to_excel)
        hb.addWidget(btn_export_scores)
        vb.addLayout(hb)
        gb.setLayout(vb)
        lay.addWidget(gb)
        self.setLayout(lay)
    @staticmethod
    def make_tag(text, bg, fg):
        from PySide6.QtWidgets import QLabel
        lab = QLabel(text)
        lab.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lab.setStyleSheet(f"QLabel {{ background-color:{bg}; color:{fg}; border-radius:10px; padding:2px 8px; font-size:12px; }}")
        return lab
    def refresh_scores(self):
        self.scores_table = getattr(self, 'scores_table', QTableWidget(0, 8))
        self.scores_table.setRowCount(0)
        for a in list_attempts_with_user():
            r = self.scores_table.rowCount()
            self.scores_table.insertRow(r)
            self.scores_table.setItem(r, 0, QTableWidgetItem(a[0]))
            self.scores_table.setItem(r, 1, QTableWidgetItem(a[1] or ''))
            self.scores_table.setItem(r, 2, QTableWidgetItem(a[2] or ''))
            self.scores_table.setItem(r, 3, QTableWidgetItem(str(a[3])))
            exam_title = get_exam_title(int(a[4])) if a[4] is not None else ''
            self.scores_table.setItem(r, 4, QTableWidgetItem(exam_title or ''))
            self.scores_table.setItem(r, 5, QTableWidgetItem(a[5] or ''))
            self.scores_table.setItem(r, 6, QTableWidgetItem(a[6] or tr('scores.not_submitted')))
            is_valid = (len(a) > 10 and a[10] == 1)
            passed_text = '数据异常' if not is_valid else ('通过' if a[8] == 1 else '未通过')
            badge_bg = '#fff3cd' if not is_valid else ('#e1f3d8' if a[8] == 1 else '#fde2e2')
            badge_fg = '#8a6d3b' if not is_valid else ('#67c23a' if a[8] == 1 else '#f56c6c')
            total = int(a[9] or 0)
            self.scores_table.setCellWidget(r, 7, self.make_tag(f'{a[7]} / {total} / {passed_text}', badge_bg, badge_fg))
            for c in (3,4):
                it = self.scores_table.item(r, c)
                if it:
                    it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        try:
            for r in range(self.scores_table.rowCount()):
                for c in range(self.scores_table.columnCount()):
                    it = self.scores_table.item(r, c)
                    if it:
                        it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        except Exception:
            pass

    def on_item_double_clicked(self, row, column):
        uuid_item = self.scores_table.item(row, 0)
        if uuid_item:
            attempt_uuid = uuid_item.text()
            dlg = ScoreDetailWindow(attempt_uuid, self)
            dlg.exec()

    def export_scores_to_excel(self):
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents/scores')
        fn, sel = QFileDialog.getSaveFileName(self, tr('scores.export_excel'), suggested, 'Excel (*.xlsx)')
        if not fn:
            return
        try:
            ext = os.path.splitext(fn)[1].lower()
            out = fn if ext == '.xlsx' else fn + '.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = 'Scores'
            headers = ['尝试UUID', '用户名', '姓名', '用户ID', '试题标题', '开始', '提交', '分数', '状态']
            ws.append(headers)
            green_fill = PatternFill(start_color='FF67C23A', end_color='FF67C23A', fill_type='solid')
            red_fill = PatternFill(start_color='FFF56C6C', end_color='FFF56C6C', fill_type='solid')
            header_fill = PatternFill(start_color='FF409EFF', end_color='FF409EFF', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFFFF', size=13)
            data_font = Font(size=12)
            center = Alignment(horizontal='center', vertical='center')
            left = Alignment(horizontal='left', vertical='center')
            thin = Side(style='thin', color='FFDDDDDD')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for a in list_attempts_with_user():
                exam_title = get_exam_title(int(a[4])) if a[4] is not None else ''
                is_valid = (len(a) > 10 and a[10] == 1)
                text_pass = '数据异常' if not is_valid else ('通过' if a[8] == 1 else '未通过')
                ws.append([a[0], a[1] or '', a[2] or '', int(a[3]), exam_title or '', a[5] or '', a[6] or '', a[7], text_pass])
                cell = ws.cell(row=ws.max_row, column=9)
                if not is_valid:
                    cell.fill = PatternFill(start_color='FFFFF3CD', end_color='FFFFF3CD', fill_type='solid')
                else:
                    cell.fill = green_fill if a[8] == 1 else red_fill
            for c in range(1, len(headers)+1):
                cell = ws.cell(row=1, column=c)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
            ws.row_dimensions[1].height = 26
            for r in range(2, ws.max_row+1):
                for c in range(1, len(headers)+1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = border
                    cell.font = data_font
                ws.cell(row=r, column=4).alignment = center
                ws.cell(row=r, column=6).alignment = center
                ws.cell(row=r, column=7).alignment = center
                ws.cell(row=r, column=8).alignment = center
                ws.cell(row=r, column=9).alignment = center
                ws.cell(row=r, column=1).alignment = left
                ws.cell(row=r, column=2).alignment = left
                ws.cell(row=r, column=3).alignment = left
                ws.row_dimensions[r].height = 22
                ws.cell(row=r, column=5).alignment = left
            widths = [0] * len(headers)
            for r in ws.iter_rows(values_only=True):
                for idx, val in enumerate(r):
                    l = len(str(val)) if val is not None else 0
                    widths[idx] = max(widths[idx], l)
            for i, w in enumerate(widths, start=1):
                letter = get_column_letter(i)
                ws.column_dimensions[letter].width = max(16, min(48, w + 6))
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
            wb.save(out)
            show_info(self, tr('common.success'), tr('export.scores.done'))
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))
