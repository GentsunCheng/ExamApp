import os
import pathlib

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

from PySide6.QtCore import Qt
from PySide6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QComboBox, QPushButton, QFileDialog, QScrollArea, QTableWidget, QTableWidgetItem, QCheckBox, QMessageBox, QAbstractItemView

from icon_manager import IconManager
from theme_manager import theme_manager
from utils import show_info, show_warn, ask_yes_no
from language import tr
from windows.progress_overview_window import ProgressOverviewWindow

from models import (
    list_users,
    PROGRESS_STATUS_NOT_STARTED,
    PROGRESS_STATUS_IN_PROGRESS,
    PROGRESS_STATUS_COMPLETED,
    upsert_progress_module,
    upsert_progress_task,
    list_progress_modules,
    list_progress_tasks,
    delete_progress_module,
    get_user_progress_tree,
    set_user_task_progress,
)


_RESERVED_SHEET_NAMES = {'说明', '_meta', 'meta'}


def export_progress_template(file_path):
    out = _ensure_xlsx(file_path)
    wb = Workbook()
    ws = wb.active
    ws.title = '说明'
    ws.append(['字段', '说明', '示例'])
    ws.append(['任务名', '必填；模块内唯一', '观看第一章视频'])
    ws.append(['描述', '可选', '时长约 20 分钟'])
    ws.append(['顺序', '可选；整数；用于排序', '1'])
    _apply_table_style(ws, header_fill_color='FF409EFF')
    ws.freeze_panes = 'A2'

    sample = wb.create_sheet('示例模块')
    sample.append(['任务名', '描述', '顺序'])
    sample.append(['任务1', '示例描述1', 1])
    sample.append(['任务2', '示例描述2', 2])
    _apply_table_style(sample, header_fill_color='FF409EFF')
    sample.freeze_panes = 'A2'

    wb.save(out)
    return out


def export_progress_modules_to_excel(file_path):
    out = _ensure_xlsx(file_path)
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    modules = list_progress_modules()
    tasks = list_progress_tasks(None)
    tasks_by_module = {}
    for t in tasks:
        tasks_by_module.setdefault(int(t[1]), []).append(t)
    for m in modules:
        mid = int(m[0])
        name = m[1] or ''
        ws = wb.create_sheet(_safe_sheet_name(name) or f'Module_{mid}')
        ws.append(['任务名', '描述', '顺序'])
        rows = tasks_by_module.get(mid, [])
        for t in rows:
            ws.append([t[2] or '', t[3] or '', int(t[4] or 0)])
        _apply_table_style(ws, header_fill_color='FF409EFF')
        ws.freeze_panes = 'A2'
    if not wb.sheetnames:
        ws = wb.create_sheet('Tasks')
        ws.append(['任务名', '描述', '顺序'])
        _apply_table_style(ws, header_fill_color='FF409EFF')
        ws.freeze_panes = 'A2'
    wb.save(out)
    return out


def import_progress_from_excel(file_path, replace=False):
    wb = load_workbook(file_path)
    summary = {'modules': 0, 'tasks': 0, 'skipped_sheets': 0}
    for ws in wb.worksheets:
        sheet_name = (ws.title or '').strip()
        if not sheet_name or sheet_name in _RESERVED_SHEET_NAMES:
            summary['skipped_sheets'] += 1
            continue
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        header = [str(x).strip() if x is not None else '' for x in (header_row or [])]
        col_title = _find_header_index(header, {'任务名', '任务名称', 'title', '任务'})
        if col_title is None:
            raise Exception(f'{tr("progress.worksheet")} {sheet_name} {tr("error.missing_column")}: {tr("progress.headers.task_title")}')
        col_desc = _find_header_index(header, {'描述', '任务描述', 'desc', 'description'})
        col_order = _find_header_index(header, {'顺序', '排序', 'order', 'sort'})

        if replace:
            try:
                module_id = upsert_progress_module(sheet_name)
                delete_progress_module(module_id)
            except Exception:
                pass
        module_id = upsert_progress_module(sheet_name)
        summary['modules'] += 1
        for r in ws.iter_rows(min_row=2, values_only=True):
            title = _cell_str(r, col_title)
            if not title:
                continue
            desc = _cell_str(r, col_desc) if col_desc is not None else None
            order_val = _cell_int(r, col_order) if col_order is not None else 0
            upsert_progress_task(module_id, title, desc, order_val)
            summary['tasks'] += 1
    return summary


def export_user_progress_to_excel(user_id, file_path):
    out = _ensure_xlsx(file_path)
    tree = get_user_progress_tree(user_id)
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for md in tree:
        module_name = md.get('module_name') or ''
        ws = wb.create_sheet(_safe_sheet_name(module_name) or 'Progress')
        ws.append([
            tr('progress.headers.task_title'),
            tr('progress.headers.description'),
            tr('progress.headers.order'),
            tr('progress.headers.status'),
            tr('progress.headers.updated_at'),
            tr('progress.headers.updated_by'),
        ])
        for t in md.get('tasks') or []:
            status = int(t.get('status') or 0)
            ws.append([
                t.get('title') or '',
                t.get('description') or '',
                int(t.get('sort_order') or 0),
                _status_text(status),
                t.get('updated_at'),
                t.get('updated_by'),
            ])
        _apply_table_style(ws, header_fill_color='FF409EFF')
        _apply_status_color(ws, status_col=4)
        ws.freeze_panes = 'A2'
    if not wb.sheetnames:
        ws = wb.create_sheet('Progress')
        ws.append([
            tr('progress.headers.task_title'),
            tr('progress.headers.description'),
            tr('progress.headers.order'),
            tr('progress.headers.status'),
            tr('progress.headers.updated_at'),
            tr('progress.headers.updated_by'),
        ])
        _apply_table_style(ws, header_fill_color='FF409EFF')
        ws.freeze_panes = 'A2'
    wb.save(out)
    return out


class AdminProgressModule(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.icon_manager = IconManager()
        lay = QVBoxLayout()

        header = QGroupBox(tr('progress.group'))
        hb = QHBoxLayout()
        hb.setContentsMargins(12, 8, 12, 8)
        hb.setSpacing(8)

        self.user_select = QComboBox()
        self.user_select.setMinimumWidth(260)
        self.user_select.currentIndexChanged.connect(self.refresh_progress_view)
        hb.addWidget(self.user_select)

        self.replace_import = QCheckBox(tr('progress.replace_import'))
        colors_checkbox = theme_manager.get_theme_colors()
        self.replace_import.setStyleSheet(
            f"QCheckBox {{ color:{colors_checkbox['text_primary']}; font-size:14px; }}\n"
            f"QCheckBox::indicator {{ width:40px; height:22px; border-radius:11px; }}\n"
            f"QCheckBox::indicator:unchecked {{ background-color:{colors_checkbox['border_light']}; border:1px solid {colors_checkbox['border']}; }}\n"
            f"QCheckBox::indicator:checked {{ background-color:{colors_checkbox['primary']}; border:1px solid {colors_checkbox['primary']}; }}"
        )
        hb.addWidget(self.replace_import)

        btn_export_tpl = QPushButton(tr('progress.export_tpl'))
        btn_export_tpl.setIcon(self.icon_manager.get_icon('exam_export'))
        btn_export_tpl.clicked.connect(self.export_template)
        hb.addWidget(btn_export_tpl)

        btn_import = QPushButton(tr('progress.import_tpl'))
        btn_import.setIcon(self.icon_manager.get_icon('exam_import'))
        btn_import.clicked.connect(self.import_template)
        hb.addWidget(btn_import)

        btn_export_user = QPushButton(tr('progress.export_user_btn'))
        btn_export_user.setIcon(self.icon_manager.get_icon('exam_export'))
        btn_export_user.clicked.connect(self.export_user_progress)
        hb.addWidget(btn_export_user)

        btn_overview = QPushButton(tr('progress.overview'))
        btn_overview.setIcon(self.icon_manager.get_icon('score'))
        btn_overview.clicked.connect(self.open_overview)
        hb.addWidget(btn_overview)

        hb.addStretch()
        header.setLayout(hb)
        lay.addWidget(header)

        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        colors_scroll = theme_manager.get_theme_colors()
        self.scroll.setStyleSheet(
            f"QScrollArea {{ border:1px solid {colors_scroll['border']}; border-radius:8px; background-color:{colors_scroll['card_background']}; }}"
        )
        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout()
        self.content_widget.setLayout(self.content_layout)
        self.scroll.setWidget(self.content_widget)
        lay.addWidget(self.scroll)

        lay.addStretch()
        self.setLayout(lay)

        self.refresh_users_and_view()

        self.overview_window = None

    def refresh_users_and_view(self):
        current_id = self.get_selected_user_id()
        self.user_select.blockSignals(True)
        self.user_select.clear()
        users = list_users()
        for u in users:
            label = u[1] if not u[2] else f'{u[1]} ({u[2]})'
            self.user_select.addItem(label, int(u[0]))
        self.user_select.blockSignals(False)
        if current_id is not None:
            idx = self.user_select.findData(int(current_id))
            if idx >= 0:
                self.user_select.setCurrentIndex(idx)
        self.refresh_progress_view()

    def get_selected_user_id(self):
        try:
            v = self.user_select.currentData()
            return int(v) if v is not None else None
        except Exception:
            return None

    def export_template(self):
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents/progress_template')
        fn, sel = QFileDialog.getSaveFileName(self, tr('progress.export_tpl.title'), suggested, 'Excel (*.xlsx)')
        if not fn:
            return
        try:
            out = export_progress_template(fn)
            show_info(self, tr('common.success'), tr('progress.export_tpl.done', path=out))
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))

    def import_template(self):
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents')
        fn, sel = QFileDialog.getOpenFileName(self, tr('progress.import_tpl.title'), suggested, 'Excel (*.xlsx)')
        if not fn:
            return
        replace = bool(self.replace_import.isChecked())
        if replace:
            reply = ask_yes_no(self, tr('common.hint'), tr('progress.import_tpl.replace_confirm'), default_yes=False)
            if reply != QMessageBox.StandardButton.Yes:
                return
        try:
            summary = import_progress_from_excel(fn, replace=replace)
            self.refresh_progress_view()
            show_info(
                self,
                tr('common.success'),
                tr(
                    'progress.import_tpl.result',
                    modules=summary['modules'],
                    tasks=summary['tasks'],
                    skipped_sheets=summary['skipped_sheets'],
                ),
            )
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))

    def export_user_progress(self):
        user_id = self.get_selected_user_id()
        if user_id is None:
            show_warn(self, tr('common.error'), tr('error.select_user'))
            return
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents/user_progress')
        fn, sel = QFileDialog.getSaveFileName(self, tr('progress.export_user.title'), suggested, 'Excel (*.xlsx)')
        if not fn:
            return
        try:
            out = export_user_progress_to_excel(user_id, fn)
            show_info(self, tr('common.success'), tr('progress.export_user.done', path=out))
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))

    def open_overview(self):
        user_id = self.get_selected_user_id()
        if user_id is None:
            show_warn(self, tr('common.error'), tr('error.select_user'))
            return
        tree = get_user_progress_tree(user_id)
        label = self.user_select.currentText() or ''
        title = tr('progress.overview.title', user=label)
        self.overview_window = ProgressOverviewWindow(title, tree, self)
        self.overview_window.show()
        try:
            self.overview_window.raise_()
            self.overview_window.activateWindow()
        except Exception:
            pass

    def refresh_progress_view(self):
        self._clear_layout(self.content_layout)
        user_id = self.get_selected_user_id()
        if user_id is None:
            return
        tree = get_user_progress_tree(user_id)
        for md in tree:
            gb = QGroupBox(md.get('module_name') or '')
            vb = QVBoxLayout()
            tbl = QTableWidget(0, 4)
            tbl.setHorizontalHeaderLabels([
                tr('progress.headers.task_title'),
                tr('progress.headers.description'),
                tr('progress.headers.order'),
                tr('progress.headers.status'),
            ])
            tbl.setColumnWidth(0, 240)
            tbl.setColumnWidth(1, 520)
            tbl.setColumnWidth(2, 80)
            tbl.setColumnWidth(3, 120)
            tbl.horizontalHeader().setStretchLastSection(True)
            tbl.setAlternatingRowColors(True)
            tbl.setShowGrid(False)
            tbl.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
            tasks = md.get('tasks') or []
            tbl.setRowCount(len(tasks))
            for r, t in enumerate(tasks):
                it_title = QTableWidgetItem(t.get('title') or '')
                it_title.setData(Qt.ItemDataRole.UserRole, int(t.get('task_id') or 0))
                tbl.setItem(r, 0, it_title)
                tbl.setItem(r, 1, QTableWidgetItem(t.get('description') or ''))
                it_order = QTableWidgetItem(str(int(t.get('sort_order') or 0)))
                it_order.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                it_order.setFlags(it_order.flags() & ~Qt.ItemFlag.ItemIsEditable)
                tbl.setItem(r, 2, it_order)

                cb = QComboBox()
                cb.addItem(tr('progress.status.not_started'), PROGRESS_STATUS_NOT_STARTED)
                cb.addItem(tr('progress.status.in_progress'), PROGRESS_STATUS_IN_PROGRESS)
                cb.addItem(tr('progress.status.completed'), PROGRESS_STATUS_COMPLETED)
                cb.setProperty('task_id', int(t.get('task_id') or 0))
                cb.setProperty('user_id', int(user_id))
                status = int(t.get('status') or 0)
                idx = cb.findData(status)
                cb.blockSignals(True)
                cb.setCurrentIndex(idx if idx >= 0 else 0)
                cb.blockSignals(False)
                self._apply_status_style(cb, status)
                cb.currentIndexChanged.connect(self.on_status_changed)
                tbl.setCellWidget(r, 3, cb)
            vb.addWidget(tbl)
            gb.setLayout(vb)
            self.content_layout.addWidget(gb)
        self.content_layout.addStretch()

    def on_status_changed(self, _index):
        cb = self.sender()
        if cb is None:
            return
        if not isinstance(cb, QComboBox):
            return
        try:
            user_id = int(cb.property('user_id'))
            task_id = int(cb.property('task_id'))
            status = int(cb.currentData())
        except Exception:
            return
        try:
            set_user_task_progress(user_id, task_id, status, updated_by='admin')
            self._apply_status_style(cb, status)
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))

    @staticmethod
    def _apply_status_style(widget, status):
        colors = theme_manager.get_theme_colors()
        if int(status) == PROGRESS_STATUS_COMPLETED:
            bg = '#e1f3d8'
            fg = '#67c23a'
        elif int(status) == PROGRESS_STATUS_IN_PROGRESS:
            bg = '#d9ecff'
            fg = colors.get('primary') or '#409eff'
        else:
            bg = '#f4f4f5'
            fg = colors.get('text_secondary') or '#909399'
        widget.setStyleSheet(f"QComboBox {{ background-color:{bg}; color:{fg}; padding:4px 10px; border-radius:10px; }}")

    def _clear_layout(self, layout):
        while layout.count():
            item = layout.takeAt(0)
            w = item.widget()
            if w is not None:
                w.deleteLater()
                continue
            child_layout = item.layout()
            if child_layout is not None:
                self._clear_layout(child_layout)


def _status_text(status):
    if int(status) == PROGRESS_STATUS_COMPLETED:
        return tr('progress.status.completed')
    if int(status) == PROGRESS_STATUS_IN_PROGRESS:
        return tr('progress.status.in_progress')
    return tr('progress.status.not_started')


def _status_fill(status):
    if int(status) == PROGRESS_STATUS_COMPLETED:
        return PatternFill(start_color='FF67C23A', end_color='FF67C23A', fill_type='solid')
    if int(status) == PROGRESS_STATUS_IN_PROGRESS:
        return PatternFill(start_color='FF409EFF', end_color='FF409EFF', fill_type='solid')
    return PatternFill(start_color='FF909399', end_color='FF909399', fill_type='solid')


def _apply_status_color(ws, status_col):
    if ws.max_row < 2:
        return
    completed = _status_text(PROGRESS_STATUS_COMPLETED)
    in_progress = _status_text(PROGRESS_STATUS_IN_PROGRESS)
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=status_col)
        raw = str(cell.value).strip() if cell.value is not None else ''
        if raw == completed:
            fill = _status_fill(PROGRESS_STATUS_COMPLETED)
        elif raw == in_progress:
            fill = _status_fill(PROGRESS_STATUS_IN_PROGRESS)
        else:
            fill = _status_fill(PROGRESS_STATUS_NOT_STARTED)
        cell.fill = fill
        cell.font = Font(color='FFFFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _apply_table_style(ws, header_fill_color='FF409EFF'):
    header_fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', size=13)
    data_font = Font(size=12)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='FFDDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 26
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 22
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.font = data_font
            if c == 1:
                cell.alignment = left
            else:
                cell.alignment = center
    widths = [0] * ws.max_column
    for row in ws.iter_rows(values_only=True):
        for idx, val in enumerate(row):
            l = len(str(val)) if val is not None else 0
            widths[idx] = max(widths[idx], l)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(16, min(64, w + 6))


def _ensure_xlsx(path):
    ext = os.path.splitext(path)[1].lower().strip()
    return path if ext == '.xlsx' else path + '.xlsx'


def _safe_sheet_name(name):
    n = str(name).strip()
    if not n:
        return ''
    bad = ['\\', '/', '*', '[', ']', ':', '?']
    for ch in bad:
        n = n.replace(ch, ' ')
    n = n.strip()
    if len(n) > 31:
        n = n[:31]
    return n


def _find_header_index(header, candidates):
    for idx, h in enumerate(header):
        if not h:
            continue
        if str(h).strip() in candidates:
            return idx
    return None


def _cell_str(row, idx):
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _cell_int(row, idx):
    if idx is None or idx >= len(row):
        return 0
    v = row[idx]
    if v is None or str(v).strip() == '':
        return 0
    try:
        return int(v)
    except Exception:
        try:
            return int(float(v))
        except Exception:
            return 0
