import os
import pathlib
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QFormLayout, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem, QTextBrowser, QProgressDialog, QFileDialog, QMessageBox
from theme_manager import theme_manager
from language import tr
from icon_manager import get_icon
from utils import show_info, show_warn, ask_yes_no
from database import DB_DIR
from models import list_sync_targets, upsert_sync_target, delete_sync_target, update_sync_target, get_exam_title
from sync import rsync_push, rsync_pull_scores
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


class SyncWorker(QThread):
    progress = Signal(str)
    finished = Signal(str)
    error = Signal(str)
    def __init__(self, targets, operation='push'):
        super().__init__()
        self.targets = targets
        self.operation = operation
    def run(self):
        results = []
        for t in self.targets:
            try:
                ssh_password = t[5] if len(t) > 5 else None
                if self.operation == 'sync':
                    base_dir = os.path.join(DB_DIR, 'pulled')
                    os.makedirs(base_dir, exist_ok=True)
                    ip = t[2]
                    dest_dir = os.path.join(base_dir, ip)
                    os.makedirs(dest_dir, exist_ok=True)
                    pull_msg = ''
                    code, out, err = rsync_pull_scores(ip, t[3], t[4], dest_dir, ssh_password)
                    if code == 0:
                        pull_msg = f'{t[1]} ({ip}) 拉取成功'
                        rp = os.path.join(dest_dir, 'scores.db')
                        try:
                            from models import merge_remote_scores_db
                            merge_remote_scores_db(rp)
                            pull_msg += ' (成绩已合并)'
                        except Exception as me:
                            pull_msg += f' (合并失败: {str(me)})'
                    else:
                        pull_msg = f'{t[1]} ({ip}) 未找到远端成绩，跳过合并'
                    self.progress.emit(pull_msg)
                    code2, out2, err2 = rsync_push(t[2], t[3], t[4], ssh_password)
                    if code2 == 0:
                        push_msg = f'{t[1]} ({t[2]}) 上传完成'
                    else:
                        push_msg = f'{t[1]} ({t[2]}) 上传失败: {err2 or "未知错误"}'
                    self.progress.emit(push_msg)
                    result = pull_msg + '；' + push_msg
                elif self.operation == 'push':
                    code, out, err = rsync_push(t[2], t[3], t[4], ssh_password)
                    result = f'{t[1]} ({t[2]}) ' + ('推送成功' if code == 0 else f'推送失败: {err or "未知错误"}')
                    self.progress.emit(result)
                else:
                    base_dir = os.path.join(DB_DIR, 'pulled')
                    os.makedirs(base_dir, exist_ok=True)
                    ip = t[2]
                    dest_dir = os.path.join(base_dir, ip)
                    os.makedirs(dest_dir, exist_ok=True)
                    code, out, err = rsync_pull_scores(ip, t[3], t[4], dest_dir, ssh_password)
                    if code == 0:
                        result = f'{t[1]} ({ip}) 拉取成功'
                        rp = os.path.join(dest_dir, 'scores.db')
                        try:
                            from models import merge_remote_scores_db
                            merge_remote_scores_db(rp)
                            result += ' (成绩已合并)'
                        except Exception as me:
                            result += f' (合并失败: {str(me)})'
                    else:
                        result = f'{t[1]} ({ip}) 拉取失败: {err or "未知错误"}'
                    self.progress.emit(result)
                results.append(result)
            except Exception as e:
                error_msg = f'{t[1]} 错误: {str(e)}'
                results.append(error_msg)
                self.error.emit(error_msg)
        self.finished.emit('\n'.join(results))


class AdminSyncModule(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        lay = QVBoxLayout()
        gb1 = QGroupBox(tr('admin.targets.group'))
        vb1 = QVBoxLayout()
        self.targets_table = QTableWidget(0, 6)
        self.targets_table.setHorizontalHeaderLabels([tr('admin.targets.headers.name'), tr('admin.targets.headers.ip'), tr('admin.targets.headers.username'), tr('admin.targets.headers.remote_path'), tr('admin.targets.headers.ssh_password'), tr('admin.users.headers.actions')])
        self.targets_table.setColumnWidth(0, 150)
        self.targets_table.setColumnWidth(1, 150)
        self.targets_table.setColumnWidth(2, 150)
        self.targets_table.setColumnWidth(3, 300)
        self.targets_table.setColumnWidth(4, 150)
        self.targets_table.horizontalHeader().setStretchLastSection(True)
        self.targets_table.setShowGrid(False)
        self.refresh_targets()
        vb1.addWidget(self.targets_table)
        gb1.setLayout(vb1)
        lay.addWidget(gb1)
        gb2 = QGroupBox(tr('admin.targets.add_group'))
        form = QFormLayout()
        self.t_name = QLineEdit()
        self.t_name.setPlaceholderText(tr('admin.targets.name_ph'))
        try:
            self.t_name.setValidator(None)
        except Exception:
            pass
        self.t_ip = QLineEdit()
        self.t_ip.setPlaceholderText(tr('admin.targets.ip_ph'))
        self.t_user = QLineEdit()
        self.t_user.setPlaceholderText(tr('admin.targets.username_ph'))
        self.t_path = QLineEdit()
        self.t_path.setText(tr('admin.targets.remote_path_ph'))
        self.t_password = QLineEdit()
        self.t_password.setPlaceholderText(tr('admin.targets.ssh_password_ph'))
        self.t_password.setEchoMode(QLineEdit.EchoMode.Password)
        add_btn = QPushButton(tr('admin.targets.add_btn'))
        add_btn.clicked.connect(self.add_target)
        form.addRow('名称', self.t_name)
        form.addRow('IP', self.t_ip)
        form.addRow('用户名', self.t_user)
        form.addRow('远程路径', self.t_path)
        form.addRow('SSH密码', self.t_password)
        form.addRow(add_btn)
        gb2.setLayout(form)
        lay.addWidget(gb2)
        hb_tpl = QHBoxLayout()
        btn_export_targets_tpl = QPushButton(tr('admin.export.targets_tpl.title'))
        btn_export_targets_tpl.setIcon(get_icon('exam_export'))
        btn_export_targets_tpl.clicked.connect(self.export_targets_template)
        btn_import_targets_excel = QPushButton(tr('admin.import.targets.title'))
        btn_import_targets_excel.setIcon(get_icon('exam_import'))
        btn_import_targets_excel.clicked.connect(self.import_targets_from_excel)
        hb_tpl.addWidget(btn_export_targets_tpl)
        hb_tpl.addWidget(btn_import_targets_excel)
        lay.addLayout(hb_tpl)
        hb = QHBoxLayout()
        self.sync_btn = QPushButton(tr('sync.sync_btn'))
        self.sync_btn.setIcon(get_icon('push'))
        self.sync_btn.clicked.connect(self.sync_all)
        hb.addWidget(self.sync_btn)
        lay.addLayout(hb)
        self.sync_spinner = None
        colors_log = theme_manager.get_theme_colors()
        self.sync_log = QTextBrowser()
        self.sync_log.setReadOnly(True)
        self.sync_log.setMinimumHeight(140)
        self.sync_log.setStyleSheet(
            f"QTextBrowser {{ background-color:{colors_log['card_background']}; border:1px solid {colors_log['border']}; border-radius:8px; padding:8px; color:{colors_log['text_primary']}; }}"
        )
        lay.addWidget(self.sync_log)
        lay.addStretch()
        self.setLayout(lay)
    def show_sync_progress(self, message):
        if hasattr(self, 'sync_progress_dialog') and getattr(self, 'sync_progress_dialog'):
            try:
                self.sync_progress_dialog.close()
            except Exception:
                pass
        colors = theme_manager.get_theme_colors()
        dlg = QProgressDialog(message, '', 0, 0, self)
        dlg.setWindowTitle(tr('sync.progress.title'))
        dlg.setCancelButton(None)
        dlg.setMinimumDuration(0)
        dlg.setAutoClose(False)
        dlg.setAutoReset(False)
        dlg.setModal(True)
        dlg.setRange(0, 0)
        dlg.setLabelText(message)
        dlg.setFixedSize(420, 140)
        try:
            dlg.setWindowFlags(dlg.windowFlags() & ~Qt.WindowType.WindowContextHelpButtonHint)
        except Exception:
            pass
        dlg.setStyleSheet(
            f"QProgressDialog {{ background-color:{colors['card_background']}; border:1px solid {colors['border']}; border-radius:12px; }}\n"
            f"QLabel {{ color:{colors['text_primary']}; font-size:14px; padding:12px; }}\n"
            f"QProgressBar {{ background-color:{colors['border_light']}; border:1px solid {colors['border']}; border-radius:10px; height:16px; margin:8px 12px; }}\n"
            f"QProgressBar::chunk {{ background-color:{colors['primary']}; border-radius:10px; }}"
        )
        dlg.show()
        self.sync_progress_dialog = dlg
    def make_tag(self, text, bg, fg):
        from PySide6.QtWidgets import QLabel
        lab = QLabel(text)
        lab.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lab.setStyleSheet(f"QLabel {{ background-color:{bg}; color:{fg}; border-radius:10px; padding:2px 8px; font-size:12px; }}")
        return lab
    def append_sync_log(self, message):
        colors = theme_manager.get_theme_colors()
        fg_ok = '#67c23a'
        bg_ok = '#e1f3d8'
        fg_err = '#f56c6c'
        bg_err = '#fde2e2'
        fg_info = colors['text_secondary']
        bg_info = colors['border_light']
        status = '成功' if ('成功' in message and '失败' not in message) else ('失败' if '失败' in message else '信息')
        if status == '成功':
            badge_fg, badge_bg = fg_ok, bg_ok
        elif status == '失败':
            badge_fg, badge_bg = fg_err, bg_err
        else:
            badge_fg, badge_bg = fg_info, bg_info
        html = f"<div style=\"margin:4px 0;\">"
        html += f"<span style=\"background:{badge_bg}; color:{badge_fg}; border-radius:10px; padding:2px 8px; font-size:12px;\">{status}</span>"
        html += f"<span style=\"margin-left:8px;\">{message}</span>"
        html += "</div>"
        self.sync_log.append(html)
    def set_sync_buttons_enabled(self, enabled):
        for i in range(self.targets_table.rowCount()):
            widget = self.targets_table.cellWidget(i, 5)
            if widget:
                for child in widget.findChildren(QPushButton):
                    child.setEnabled(enabled)
        if hasattr(self, 'sync_btn'):
            self.sync_btn.setEnabled(enabled)
    def update_progress_message(self, msg):
        try:
            if hasattr(self, 'sync_progress_dialog') and getattr(self, 'sync_progress_dialog'):
                self.sync_progress_dialog.setLabelText(msg)
        except Exception:
            pass
    def sync_all(self):
        targets = list_sync_targets()
        if not targets:
            show_info(self, tr('sync.status.info'), tr('info.no_targets'))
            return
        self.set_sync_buttons_enabled(False)
        if hasattr(self, 'sync_btn'):
            self.sync_btn.setEnabled(False)
        if hasattr(self, 'sync_log'):
            self.sync_log.clear()
        self.sync_worker = SyncWorker(targets, 'sync')
        self.sync_worker.finished.connect(self.on_sync_finished)
        self.sync_worker.error.connect(self.on_sync_error)
        self.sync_worker.progress.connect(self.append_sync_log)
        self.sync_worker.progress.connect(self.update_progress_message)
        self.sync_worker.start()
        self.show_sync_progress(tr('sync.syncing_message'))
    def push_all(self):
        targets = list_sync_targets()
        if not targets:
            show_info(self, tr('sync.status.info'), tr('info.no_targets'))
            return
        self.set_sync_buttons_enabled(False)
        if hasattr(self, 'sync_log'):
            self.sync_log.clear()
        self.sync_worker = SyncWorker(targets, 'push')
        self.sync_worker.finished.connect(self.on_sync_finished)
        self.sync_worker.error.connect(self.on_sync_error)
        self.sync_worker.progress.connect(self.append_sync_log)
        self.sync_worker.progress.connect(self.update_progress_message)
        self.sync_worker.start()
        self.show_sync_progress(tr('sync.pushing_message'))
    def pull_all(self):
        targets = list_sync_targets()
        if not targets:
            show_info(self, tr('sync.status.info'), tr('info.no_targets'))
            return
        self.set_sync_buttons_enabled(False)
        if hasattr(self, 'sync_log'):
            self.sync_log.clear()
        self.sync_worker = SyncWorker(targets, 'pull')
        self.sync_worker.finished.connect(self.on_sync_finished)
        self.sync_worker.error.connect(self.on_sync_error)
        self.sync_worker.progress.connect(self.append_sync_log)
        self.sync_worker.progress.connect(self.update_progress_message)
        self.sync_worker.start()
        self.show_sync_progress(tr('sync.pulling_message'))
    def on_sync_finished(self, results):
        self.set_sync_buttons_enabled(True)
        if hasattr(self, 'sync_worker'):
            self.sync_worker.deleteLater()
        if hasattr(self, 'sync_progress_dialog') and getattr(self, 'sync_progress_dialog'):
            try:
                self.sync_progress_dialog.close()
            except Exception:
                pass
            self.sync_progress_dialog = None
        try:
            parent = self.parent()
            if parent and hasattr(parent, 'refresh_scores'):
                parent.refresh_scores()
        except Exception:
            pass
        show_info(self, tr('sync.finished.title'), tr('sync.operation_done', results=results))
    def on_sync_error(self, error):
        self.set_sync_buttons_enabled(True)
        if hasattr(self, 'sync_worker'):
            self.sync_worker.deleteLater()
        if hasattr(self, 'sync_progress_dialog') and getattr(self, 'sync_progress_dialog'):
            try:
                self.sync_progress_dialog.close()
            except Exception:
                pass
            self.sync_progress_dialog = None
        show_warn(self, tr('sync.error.title'), tr('sync.error.message', error=error))
    def export_targets_template(self):
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents/targets_template')
        fn, sel = QFileDialog.getSaveFileName(self, '导出设备Excel模板', suggested, 'Excel (*.xlsx)')
        if not fn:
            return
        headers = ['名称', 'IP', '用户名', '远程路径', 'SSH密码']
        try:
            ext = os.path.splitext(fn)[1].lower()
            out = fn if ext == '.xlsx' else fn + '.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = 'Targets'
            ws.append(headers)
            ws.append(['设备A', '192.168.1.10', 'user', '~/.exam_system/', ''])
            header_fill = PatternFill(start_color='FF409EFF', end_color='FF409EFF', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFFFF', size=13)
            data_font = Font(size=12)
            center = Alignment(horizontal='center', vertical='center')
            left = Alignment(horizontal='left', vertical='center')
            thin = Side(style='thin', color='FFDDDDDD')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for c in range(1, len(headers)+1):
                cell = ws.cell(row=1, column=c)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
            ws.row_dimensions[1].height = 26
            for r in range(2, ws.max_row+1):
                for c in range(1, len(headers)+1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = border
                    cell.font = data_font
                ws.cell(row=r, column=2).alignment = center
                ws.cell(row=r, column=1).alignment = left
                ws.cell(row=r, column=3).alignment = left
                ws.cell(row=r, column=4).alignment = left
                ws.cell(row=r, column=5).alignment = left
                ws.row_dimensions[r].height = 22
            widths = [0] * len(headers)
            for r in ws.iter_rows(values_only=True):
                for idx, val in enumerate(r):
                    l = len(str(val)) if val is not None else 0
                    widths[idx] = max(widths[idx], l)
            for i, w in enumerate(widths, start=1):
                letter = get_column_letter(i)
                ws.column_dimensions[letter].width = max(16, min(48, w + 6))
            ws.freeze_panes = 'A2'
            wb.save(out)
            show_info(self, tr('common.success'), tr('admin.export.targets_tpl.done'))
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))
    def import_targets_from_excel(self):
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents')
        fn, sel = QFileDialog.getOpenFileName(self, '选择设备Excel', suggested, 'Excel (*.xlsx)')
        if not fn:
            return
        try:
            wb = load_workbook(fn)
            ws = wb['Targets'] if 'Targets' in wb.sheetnames else wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            header = [str(x).strip() if x else '' for x in header_row]
            def idx(name):
                try:
                    return header.index(name)
                except Exception:
                    return -1
            iname = idx('名称'); iip = idx('IP'); iuser = idx('用户名'); ipath = idx('远程路径'); ipwd = idx('SSH密码')
            if min(iname, iip, iuser, ipath) < 0:
                show_warn(self, tr('common.error'), tr('admin.import.targets.error.missing'))
                return
            ok = 0; fail = 0
            for r in ws.iter_rows(min_row=2, values_only=True):
                try:
                    name = (str(r[iname]).strip() if iname >= 0 and iname < len(r) and r[iname] is not None else '')
                    ip = (str(r[iip]).strip() if iip >= 0 and iip < len(r) and r[iip] is not None else '')
                    user = (str(r[iuser]).strip() if iuser >= 0 and iuser < len(r) and r[iuser] is not None else '')
                    path = (str(r[ipath]).strip() if ipath >= 0 and ipath < len(r) and r[ipath] is not None else '')
                    password = (str(r[ipwd]).strip() if ipwd >= 0 and ipwd < len(r) and r[ipwd] is not None else '') or None
                    if not name or not ip or not user or not path:
                        fail += 1
                        continue
                    upsert_sync_target(name, ip, user, path, password)
                    ok += 1
                except Exception:
                    fail += 1
            self.refresh_targets()
            show_info(self, tr('common.success'), f'导入成功:{ok} 失败:{fail}')
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))
    def refresh_targets(self):
        targets = list_sync_targets()
        self.targets_table.blockSignals(True)
        self.targets_table.setRowCount(0)
        for t in targets:
            r = self.targets_table.rowCount()
            self.targets_table.insertRow(r)
            it_name = QTableWidgetItem(t[1])
            it_name.setData(Qt.ItemDataRole.UserRole, t[0])
            self.targets_table.setItem(r, 0, it_name)
            self.targets_table.setItem(r, 1, QTableWidgetItem(t[2]))
            self.targets_table.setItem(r, 2, QTableWidgetItem(t[3]))
            self.targets_table.setItem(r, 3, QTableWidgetItem(t[4]))
            password_text = '******' if t[5] else ''
            self.targets_table.setItem(r, 4, QTableWidgetItem(password_text))
            action_widget = QWidget()
            action_layout = QHBoxLayout()
            action_layout.setContentsMargins(4, 4, 4, 4)
            edit_btn = QPushButton('编辑')
            edit_btn.setStyleSheet("QPushButton { background-color:#409eff; color:#fff; padding:4px 8px; font-size:12px; border-radius:6px; }")
            edit_btn.clicked.connect(lambda checked, tid=t[0]: self.edit_target(tid))
            action_layout.addWidget(edit_btn)
            delete_btn = QPushButton('删除')
            delete_btn.setStyleSheet("QPushButton { background-color:#f56c6c; color:#fff; padding:4px 8px; font-size:12px; }")
            delete_btn.clicked.connect(lambda checked, tid=t[0]: self.delete_target(tid))
            action_layout.addWidget(delete_btn)
            action_widget.setLayout(action_layout)
            self.targets_table.setCellWidget(r, 5, action_widget)
        try:
            for r in range(self.targets_table.rowCount()):
                for c in range(self.targets_table.columnCount()):
                    it = self.targets_table.item(r, c)
                    if it:
                        it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        except Exception:
            pass
        self.targets_table.blockSignals(False)
        try:
            self.targets_table.itemChanged.disconnect()
        except Exception:
            pass
        self.targets_table.itemChanged.connect(self.on_target_item_changed)
    def on_target_item_changed(self, item):
        row = item.row()
        col = item.column()
        if col not in (0,1,2,3):
            return
        id_item = self.targets_table.item(row, 0)
        target_id = id_item.data(Qt.ItemDataRole.UserRole) if id_item else None
        if target_id is None:
            return
        name = self.targets_table.item(row, 0).text() if self.targets_table.item(row, 0) else ''
        ip = self.targets_table.item(row, 1).text() if self.targets_table.item(row, 1) else ''
        username = self.targets_table.item(row, 2).text() if self.targets_table.item(row, 2) else ''
        remote_path = self.targets_table.item(row, 3).text() if self.targets_table.item(row, 3) else ''
        if not name or not ip or not username or not remote_path:
            show_warn(self, tr('common.error'), '设备信息不能为空')
            self.refresh_targets()
            return
        try:
            update_sync_target(int(target_id), name, ip, username, remote_path, None)
            self.refresh_targets()
            show_info(self, tr('common.success'), '设备信息已更新')
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))
    def add_target(self):
        name = self.t_name.text().strip()
        ip = self.t_ip.text().strip()
        user = self.t_user.text().strip()
        path = self.t_path.text().strip()
        password = self.t_password.text().strip()
        if not name or not ip or not user or not path:
            show_warn(self, tr('common.error'), '请完整填写设备信息')
            return
        upsert_sync_target(name, ip, user, path, password if password else None)
        self.refresh_targets()
        self.t_name.clear()
        self.t_ip.clear()
        self.t_user.clear()
        self.t_path.clear()
        self.t_password.clear()
        show_info(self, tr('common.success'), '设备已添加')
    def edit_target(self, target_id):
        targets = list_sync_targets()
        target = None
        for t in targets:
            if t[0] == target_id:
                target = t
                break
        if not target:
            return
        from PySide6.QtWidgets import QDialog, QDialogButtonBox
        dialog = QDialog(self)
        dialog.setWindowTitle('编辑设备')
        dialog.setFixedSize(400, 300)
        layout = QFormLayout()
        name_edit = QLineEdit(target[1])
        ip_edit = QLineEdit(target[2])
        user_edit = QLineEdit(target[3])
        path_edit = QLineEdit(target[4])
        password_edit = QLineEdit()
        password_edit.setPlaceholderText('留空保持原密码')
        password_edit.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addRow('名称', name_edit)
        layout.addRow('IP', ip_edit)
        layout.addRow('用户名', user_edit)
        layout.addRow('远程路径', path_edit)
        layout.addRow('SSH密码', password_edit)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        layout.addWidget(buttons)
        dialog.setLayout(layout)
        def save_changes():
            new_name = name_edit.text().strip()
            new_ip = ip_edit.text().strip()
            new_user = user_edit.text().strip()
            new_path = path_edit.text().strip()
            new_password = password_edit.text().strip()
            if not new_name or not new_ip or not new_user or not new_path:
                show_warn(dialog, tr('common.error'), '请完整填写设备信息')
                return
            try:
                update_sync_target(target_id, new_name, new_ip, new_user, new_path, new_password if new_password else None)
                self.refresh_targets()
                dialog.accept()
                show_info(self, tr('common.success'), '设备已更新')
            except Exception as e:
                show_warn(dialog, tr('common.error'), str(e))
        buttons.accepted.connect(save_changes)
        buttons.rejected.connect(dialog.reject)
        dialog.exec()
    def delete_target(self, target_id):
        reply = ask_yes_no(self, '确认', '确定要删除该设备吗？', default_yes=False)
        if reply == QMessageBox.StandardButton.Yes:
            try:
                delete_sync_target(target_id)
                self.refresh_targets()
                show_info(self, tr('common.success'), '设备已删除')
            except Exception as e:
                show_warn(self, tr('common.error'), str(e))
