import os
import pathlib

from PySide6.QtCore import Qt
from PySide6.QtWidgets import QWidget, QVBoxLayout, QGroupBox, QScrollArea, QTableWidget, QTableWidgetItem, QLabel, QPushButton, QHBoxLayout, QFileDialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

from theme_manager import theme_manager
from language import tr
from models import list_exams, list_exam_user_overview
from utils import show_info, show_warn
from icon_manager import get_icon


class AdminScoresOverviewModule(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        lay = QVBoxLayout()
        header = QHBoxLayout()
        header.addStretch()
        btn_export = QPushButton(tr('admin.scores_overview.export_excel'))
        btn_export.setIcon(get_icon('exam_export'))
        btn_export.clicked.connect(self.export_overview)
        header.addWidget(btn_export)
        lay.addLayout(header)
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        colors_scroll = theme_manager.get_theme_colors()
        self.scroll.setStyleSheet(
            f"QScrollArea {{ border:1px solid {colors_scroll['border']}; border-radius:8px; background-color:{colors_scroll['card_background']}; }}"
        )
        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout()
        self.content_widget.setLayout(self.content_layout)
        self.scroll.setWidget(self.content_widget)
        lay.addWidget(self.scroll)
        lay.addStretch()
        self.setLayout(lay)
        self.refresh_overview()

    @staticmethod
    def make_tag(text, bg, fg):
        lab = QLabel(text)
        lab.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lab.setStyleSheet(f"QLabel {{ background-color:{bg}; color:{fg}; border-radius:10px; padding:2px 8px; font-size:12px; }}")
        return lab

    def _clear_layout(self, layout):
        while layout.count():
            item = layout.takeAt(0)
            w = item.widget()
            if w is not None:
                w.deleteLater()
                continue
            child_layout = item.layout()
            if child_layout is not None:
                self._clear_layout(child_layout)

    def refresh_overview(self):
        self._clear_layout(self.content_layout)
        exams = list_exams(include_expired=True)
        for e in exams:
            exam_id = int(e[0])
            exam_title = e[1] or ''
            gb = QGroupBox(exam_title or f'Exam {exam_id}')
            vb = QVBoxLayout()
            tbl = QTableWidget(0, 6)
            tbl.setHorizontalHeaderLabels([
                tr('admin.users.headers.id'),
                tr('admin.users.headers.username'),
                tr('admin.users.headers.full_name'),
                tr('scores.headers.submitted'),
                tr('exams.best'),
                tr('progress.headers.status'),
            ])
            tbl.setColumnWidth(0, 60)
            tbl.setColumnWidth(1, 120)
            tbl.setColumnWidth(2, 140)
            tbl.setColumnWidth(3, 200)
            tbl.setColumnWidth(4, 100)
            tbl.setColumnWidth(5, 140)
            tbl.horizontalHeader().setStretchLastSection(True)
            tbl.setAlternatingRowColors(True)
            tbl.setShowGrid(False)
            rows = list_exam_user_overview(exam_id)
            for row in rows:
                r = tbl.rowCount()
                tbl.insertRow(r)
                uid = row[0]
                uname = row[1] or ''
                full_name = row[2] or ''
                last_ts = row[3] or ''
                best_score = row[4] or 0.0
                passed = int(row[5] or 0)
                tbl.setItem(r, 0, QTableWidgetItem(str(uid)))
                tbl.setItem(r, 1, QTableWidgetItem(uname))
                tbl.setItem(r, 2, QTableWidgetItem(full_name))
                tbl.setItem(r, 3, QTableWidgetItem(last_ts))
                tbl.setItem(r, 4, QTableWidgetItem(str(best_score)))
                status_text = tr('attempts.pass') if passed == 1 else tr('attempts.fail')
                badge_bg = '#e1f3d8' if passed == 1 else '#fde2e2'
                badge_fg = '#67c23a' if passed == 1 else '#f56c6c'
                tbl.setCellWidget(r, 5, self.make_tag(status_text, badge_bg, badge_fg))
            try:
                for r in range(tbl.rowCount()):
                    for c in range(tbl.columnCount()):
                        it = tbl.item(r, c)
                        if it:
                            it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            except Exception:
                pass
            vb.addWidget(tbl)
            gb.setLayout(vb)
            self.content_layout.addWidget(gb)
        self.content_layout.addStretch()

    def export_overview(self):
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents/scores_overview')
        fn, sel = QFileDialog.getSaveFileName(self, tr('admin.scores_overview.export'), suggested, 'Excel (*.xlsx)')
        if not fn:
            return
        try:
            out = _ensure_xlsx(fn)
            wb = Workbook()
            default = wb.active
            wb.remove(default)
            exams = list_exams(include_expired=True)
            for e in exams:
                exam_id = int(e[0])
                exam_title = e[1] or ''
                sheet_name = _safe_sheet_name(exam_title) or f'Exam_{exam_id}'
                ws = wb.create_sheet(sheet_name)
                ws.append([
                    tr('admin.users.headers.id'),
                    tr('admin.users.headers.username'),
                    tr('admin.users.headers.full_name'),
                    tr('scores.headers.submitted'),
                    tr('exams.best'),
                    tr('progress.headers.status'),
                ])
                rows = list_exam_user_overview(exam_id)
                for row in rows:
                    uid = row[0]
                    uname = row[1] or ''
                    full_name = row[2] or ''
                    last_ts = row[3] or ''
                    best_score = row[4] or 0.0
                    passed = int(row[5] or 0)
                    status_text = tr('attempts.pass') if passed == 1 else tr('attempts.fail')
                    ws.append([
                        uid,
                        uname,
                        full_name,
                        last_ts,
                        best_score,
                        status_text,
                    ])
                _apply_table_style(ws, header_fill_color='FF409EFF')
            if not wb.sheetnames:
                ws = wb.create_sheet('Scores')
                ws.append([
                    tr('admin.users.headers.id'),
                    tr('admin.users.headers.username'),
                    tr('admin.users.headers.full_name'),
                    tr('scores.headers.submitted'),
                    tr('exams.best'),
                    tr('progress.headers.status'),
                ])
                _apply_table_style(ws, header_fill_color='FF409EFF')
            wb.save(out)
            show_info(self, tr('common.success'), f'{tr("admin.scores_overview.export_info")}: {out}')
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))


def _apply_table_style(ws, header_fill_color='FF409EFF'):
    header_fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', size=13)
    data_font = Font(size=12)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='FFDDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 26
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 22
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.font = data_font
            cell.alignment = left if c in (2, 3) else center
    widths = [0] * ws.max_column
    for row in ws.iter_rows(values_only=True):
        for idx, val in enumerate(row):
            l = len(str(val)) if val is not None else 0
            widths[idx] = max(widths[idx], l)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(10, min(40, w + 4))


def _ensure_xlsx(path):
    ext = os.path.splitext(path)[1].lower().strip()
    return path if ext == '.xlsx' else path + '.xlsx'


def _safe_sheet_name(name):
    n = str(name).strip()
    if not n:
        return ''
    bad = ['\\', '/', '*', '[', ']', ':', '?']
    for ch in bad:
        n = n.replace(ch, ' ')
    n = n.strip()
    if len(n) > 31:
        n = n[:31]
    return n
