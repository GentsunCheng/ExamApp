from PySide6.QtCore import Qt, QSize
from PySide6.QtGui import QRegularExpressionValidator
from PySide6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QFormLayout, QLineEdit, QComboBox, QListView, QPushButton, QTableWidget, QTableWidgetItem, QMessageBox, QFileDialog, QDialog, QDialogButtonBox
from icon_manager import IconManager
from theme_manager import theme_manager
from language import tr
from utils import show_info, show_warn, ask_yes_no
from models import list_users, list_admins, create_user, create_admin, delete_user, delete_admin, demote_admin_to_user, promote_user_to_admin, update_user_role, update_user_active, update_user_basic, update_admin_active, update_admin_basic
from PySide6.QtCore import QRegularExpression
import re
import os
import pathlib
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


class AdminUsersModule(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.icon_manager = IconManager()
        lay = QVBoxLayout()
        gb = QGroupBox(tr('admin.users_group'))
        vb = QVBoxLayout()
        self.users_table = QTableWidget(0, 7)
        self.users_table.setHorizontalHeaderLabels([tr('admin.users.headers.id'), tr('admin.users.headers.username'), tr('admin.users.headers.full_name'), tr('admin.users.headers.role'), tr('admin.users.headers.status'), tr('admin.users.headers.created_at'), tr('admin.users.headers.actions')])
        self.users_table.setColumnWidth(0, 50)
        self.users_table.setColumnWidth(1, 100)
        self.users_table.setColumnWidth(2, 100)
        self.users_table.setColumnWidth(3, 80)
        self.users_table.setColumnWidth(4, 80)
        self.users_table.setColumnWidth(5, 300)
        self.users_table.horizontalHeader().setStretchLastSection(True)
        self.users_table.setAlternatingRowColors(True)
        self.users_table.setShowGrid(False)
        self.refresh_users()
        vb.addWidget(self.users_table)
        gb.setLayout(vb)
        lay.addWidget(gb)
        self.users_table.itemChanged.connect(self.on_user_item_changed)
        gb2 = QGroupBox(tr('admin.new_user_group'))
        form = QFormLayout()
        self.new_user = QLineEdit()
        self.new_user.setPlaceholderText(tr('admin.users.username_ph'))
        self.new_user.setInputMethodHints(Qt.InputMethodHint.ImhNoPredictiveText | Qt.InputMethodHint.ImhNoAutoUppercase | Qt.InputMethodHint.ImhPreferLowercase)
        self.new_user.setValidator(QRegularExpressionValidator(QRegularExpression(r"^[A-Za-z0-9_@.\-]+$")))
        self.new_pwd = QLineEdit()
        self.new_pwd.setEchoMode(QLineEdit.EchoMode.Password)
        self.new_pwd.setPlaceholderText(tr('admin.users.password_ph'))
        self.new_pwd.setInputMethodHints(Qt.InputMethodHint.ImhHiddenText | Qt.InputMethodHint.ImhNoPredictiveText | Qt.InputMethodHint.ImhSensitiveData)
        self.new_pwd.setValidator(QRegularExpressionValidator(QRegularExpression(r"^[\x20-\x7E]+$")))
        self.new_fullname = QLineEdit()
        self.new_fullname.setPlaceholderText(tr('admin.users.full_name_ph'))
        try:
            self.new_fullname.setValidator(None)
        except Exception:
            pass
        self.new_role = QComboBox()
        colors = theme_manager.get_theme_colors()
        self.new_role.setStyleSheet(
            f"QComboBox {{ padding:6px 10px; border:1px solid {colors['input_border']}; border-radius:12px; background-color:{colors['input_background']}; color:{colors['text_primary']}; min-height:32px; }}\n"
            f"QComboBox:focus {{ border-color:{colors['primary']}; }}\n"
            f"QComboBox::drop-down {{ width:28px; border:none; }}\n"
            f"QComboBox QAbstractItemView {{ background-color:{colors['card_background']}; border:1px solid {colors['border']}; padding:6px; outline:0; }}\n"
            f"QComboBox QAbstractItemView::item {{ padding:6px 10px; height:28px; color:{colors['text_primary']}; }}\n"
            f"QComboBox QAbstractItemView::item:hover {{ background-color:{colors['border_light']}; }}\n"
            f"QComboBox QAbstractItemView::item:selected {{ background-color:{colors['primary']}; color:{colors['text_inverse']}; }}"
        )
        role_view = QListView()
        role_view.setStyleSheet(
            f"QListView {{ background-color:{colors['card_background']}; border:1px solid {colors['border']}; padding:6px; }}\n"
            f"QListView::item {{ padding:6px 10px; height:28px; color:{colors['text_primary']}; }}\n"
            f"QListView::item:hover {{ background-color:{colors['border_light']}; }}\n"
            f"QListView::item:selected {{ background-color:{colors['primary']}; color:{colors['text_inverse']}; }}"
        )
        self.new_role.setView(role_view)
        self.new_role.setIconSize(QSize(16, 16))
        self.new_role.addItem(self.icon_manager.get_icon('user'), 'user')
        self.new_role.addItem(self.icon_manager.get_icon('user_admin'), 'admin')
        add_btn = QPushButton(tr('admin.users.add_button'))
        add_btn.clicked.connect(self.add_user)
        form.addRow(tr('admin.users.headers.username'), self.new_user)
        form.addRow(tr('admin.users.password_ph'), self.new_pwd)
        form.addRow(tr('admin.users.headers.full_name'), self.new_fullname)
        form.addRow(tr('admin.users.headers.role'), self.new_role)
        form.addRow(add_btn)
        gb2.setLayout(form)
        lay.addWidget(gb2)
        hb_users_excel = QHBoxLayout()
        btn_export_users_tpl = QPushButton(tr('admin.users.export_tpl'))
        btn_export_users_tpl.setIcon(self.icon_manager.get_icon('exam_export'))
        btn_export_users_tpl.clicked.connect(self.export_users_template)
        btn_import_users_excel = QPushButton(tr('admin.users.import_excel'))
        btn_import_users_excel.setIcon(self.icon_manager.get_icon('exam_import'))
        btn_import_users_excel.clicked.connect(self.import_users_from_excel)
        hb_users_excel.addWidget(btn_export_users_tpl)
        hb_users_excel.addWidget(btn_import_users_excel)
        lay.addLayout(hb_users_excel)
        lay.addStretch()
        self.setLayout(lay)
        
    @staticmethod
    def make_tag(text, bg, fg):
        from PySide6.QtWidgets import QLabel
        lab = QLabel(text)
        lab.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lab.setStyleSheet(f"QLabel {{ background-color:{bg}; color:{fg}; border-radius:10px; padding:2px 8px; font-size:12px; }}")
        return lab
    def refresh_users(self):
        self.users_table = getattr(self, 'users_table', QTableWidget(0, 7))
        self.users_table.blockSignals(True)
        self.users_table.setRowCount(0)
        admins = []
        try:
            admins = list_admins()
        except Exception:
            admins = []
        for a in admins:
            row = self.users_table.rowCount()
            self.users_table.insertRow(row)
            it_id = QTableWidgetItem(str(a[0]))
            it_id.setData(Qt.ItemDataRole.UserRole, 'admin')
            it_id.setFlags(it_id.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.users_table.setItem(row, 0, it_id)
            it_un = QTableWidgetItem(a[1] or '')
            self.users_table.setItem(row, 1, it_un)
            it_fn = QTableWidgetItem(a[2] or '')
            self.users_table.setItem(row, 2, it_fn)
            role_text = tr('admin.role.admin')
            role_bg = '#e1f3d8'
            role_fg = '#67c23a'
            self.users_table.setCellWidget(row, 3, self.make_tag(role_text, role_bg, role_fg))
            status_text = tr('admin.status.active') if a[4] == 1 else tr('admin.status.inactive')
            status_bg = '#e1f3d8' if a[4] == 1 else '#fde2e2'
            status_fg = '#67c23a' if a[4] == 1 else '#f56c6c'
            self.users_table.setCellWidget(row, 4, self.make_tag(status_text, status_bg, status_fg))
            it_ct = QTableWidgetItem(a[5] or '')
            it_ct.setFlags(it_ct.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.users_table.setItem(row, 5, it_ct)
            action_widget = QWidget()
            action_layout = QHBoxLayout()
            action_layout.setContentsMargins(4, 4, 4, 4)
            demote_btn = QPushButton(tr('admin.user.set_user'))
            demote_btn.setIcon(self.icon_manager.get_icon('user_edit'))
            demote_btn.setStyleSheet("QPushButton { background-color:#67c23a; color:#fff; padding:4px 8px; font-size:12px; border-radius:6px; }")
            demote_btn.clicked.connect(lambda checked, aid=a[0]: self.demote_admin(aid))
            delete_btn = QPushButton(tr('admin.user.delete'))
            delete_btn.setIcon(self.icon_manager.get_icon('delete'))
            delete_btn.setStyleSheet("QPushButton { background-color:#f56c6c; color:#fff; padding:4px 8px; font-size:12px; border-radius:6px; }")
            delete_btn.clicked.connect(lambda checked, aid=a[0]: self.delete_admin(aid))
            active_btn = QPushButton(tr('admin.user.disable') if a[4] == 1 else tr('admin.user.enable'))
            active_btn.setIcon(self.icon_manager.get_icon('user_active' if a[4] == 1 else 'user_inactive'))
            active_btn.setStyleSheet("QPushButton { background-color:#e6a23c; color:#fff; padding:4px 8px; font-size:12px; border-radius:6px; }")
            active_btn.clicked.connect(lambda checked, aid=a[0], current_active=a[4]: self.toggle_admin_active(aid, current_active))
            action_layout.addWidget(delete_btn)
            action_layout.addWidget(demote_btn)
            action_layout.addWidget(active_btn)
            
            edit_btn = QPushButton(tr('common.edit'))
            edit_btn.setIcon(self.icon_manager.get_icon('user_edit'))
            edit_btn.setStyleSheet("QPushButton { background-color:#409eff; color:#fff; padding:4px 8px; font-size:12px; border-radius:6px; }")
            edit_btn.clicked.connect(lambda checked, aid=a[0]: self.edit_user(aid, 'admin'))
            action_layout.addWidget(edit_btn)
            action_widget.setLayout(action_layout)
            self.users_table.setCellWidget(row, 6, action_widget)
        for u in list_users():
            row = self.users_table.rowCount()
            self.users_table.insertRow(row)
            it_uid = QTableWidgetItem(str(u[0]))
            it_uid.setData(Qt.ItemDataRole.UserRole, 'user')
            self.users_table.setItem(row, 0, it_uid)
            self.users_table.setItem(row, 1, QTableWidgetItem(u[1] or ''))
            self.users_table.setItem(row, 2, QTableWidgetItem((u[2] or '') if u[2] else ''))
            role_text = tr('admin.role.admin') if u[3] == 'admin' else tr('admin.role.user')
            role_bg = '#e1f3d8' if u[3] == 'admin' else '#d9ecff'
            role_fg = '#67c23a' if u[3] == 'admin' else '#409eff'
            self.users_table.setCellWidget(row, 3, self.make_tag(role_text, role_bg, role_fg))
            status_text = tr('admin.status.active') if u[4] == 1 else tr('admin.status.inactive')
            status_bg = '#e1f3d8' if u[4] == 1 else '#fde2e2'
            status_fg = '#67c23a' if u[4] == 1 else '#f56c6c'
            self.users_table.setCellWidget(row, 4, self.make_tag(status_text, status_bg, status_fg))
            self.users_table.setItem(row, 5, QTableWidgetItem(u[5]))
            it_id = self.users_table.item(row, 0)
            it_ct = self.users_table.item(row, 5)
            if it_id:
                it_id.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if it_ct:
                it_ct.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            action_widget = QWidget()
            action_layout = QHBoxLayout()
            action_layout.setContentsMargins(4, 4, 4, 4)
            delete_btn = QPushButton(tr('admin.user.delete'))
            delete_btn.setIcon(self.icon_manager.get_icon('delete'))
            delete_btn.setStyleSheet("QPushButton { background-color:#f56c6c; color:#fff; padding:4px 8px; font-size:12px; border-radius:6px; }")
            delete_btn.clicked.connect(lambda checked, uid=u[0]: self.delete_user(uid))
            action_layout.addWidget(delete_btn)
            role_btn = QPushButton(tr('admin.user.set_admin') if u[3] == 'user' else tr('admin.user.set_user'))
            role_btn.setIcon(self.icon_manager.get_icon('user_edit'))
            role_btn.setStyleSheet("QPushButton { background-color:#67c23a; color:#fff; padding:4px 8px; font-size:12px; border-radius:6px; }")
            role_btn.clicked.connect(lambda checked, uid=u[0], current_role=u[3]: self.toggle_user_role(uid, current_role))
            action_layout.addWidget(role_btn)
            active_btn = QPushButton(tr('admin.user.disable') if u[4] == 1 else tr('admin.user.enable'))
            active_btn.setIcon(self.icon_manager.get_icon('user_active' if u[4] == 1 else 'user_inactive'))
            active_btn.setStyleSheet("QPushButton { background-color:#e6a23c; color:#fff; padding:4px 8px; font-size:12px; border-radius:6px; }")
            active_btn.clicked.connect(lambda checked, uid=u[0], current_active=u[4]: self.toggle_user_active(uid, current_active))
            action_layout.addWidget(active_btn)

            edit_btn = QPushButton(tr('common.edit'))
            edit_btn.setIcon(self.icon_manager.get_icon('user_edit'))
            edit_btn.setStyleSheet("QPushButton { background-color:#409eff; color:#fff; padding:4px 8px; font-size:12px; border-radius:6px; }")
            edit_btn.clicked.connect(lambda checked, uid=u[0]: self.edit_user(uid, 'user'))
            action_layout.addWidget(edit_btn)
            action_widget.setLayout(action_layout)
            self.users_table.setCellWidget(row, 6, action_widget)
        try:
            for r in range(self.users_table.rowCount()):
                for c in range(self.users_table.columnCount()):
                    it = self.users_table.item(r, c)
                    if it:
                        it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        except Exception:
            pass
        self.users_table.blockSignals(False)

    def edit_user(self, user_id, role):
        target = None
        if role == 'admin':
            admins = list_admins()
            for a in admins:
                if a[0] == user_id:
                    target = a
                    break
        else:
            users = list_users()
            for u in users:
                if u[0] == user_id:
                    target = u
                    break
        
        if not target:
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(tr('common.edit'))
        dialog.setFixedSize(400, 250)
        layout = QFormLayout()

        username_edit = QLineEdit(target[1])
        fullname_edit = QLineEdit(target[2] if target[2] else "")
        password_edit = QLineEdit()
        password_edit.setPlaceholderText(tr('info.leave_empty_to_keep_original'))
        password_edit.setEchoMode(QLineEdit.EchoMode.Password)

        layout.addRow(tr('admin.users.headers.username'), username_edit)
        layout.addRow(tr('admin.users.headers.full_name'), fullname_edit)
        layout.addRow(tr('admin.users.password_ph'), password_edit)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.button(QDialogButtonBox.StandardButton.Ok).setText(tr('common.ok'))
        buttons.button(QDialogButtonBox.StandardButton.Cancel).setText(tr('common.cancel'))
        layout.addWidget(buttons)
        dialog.setLayout(layout)

        def save_changes():
            new_username = username_edit.text().strip()
            new_fullname = fullname_edit.text().strip() or None
            new_password = password_edit.text().strip() or None

            if not new_username:
                show_warn(dialog, tr('common.error'), tr('error.username_empty'))
                return

            if not re.fullmatch(r"[A-Za-z0-9_@.\-]+", new_username):
                show_warn(dialog, tr('common.error'), tr('error.username_format'))
                return

            if new_password and not re.fullmatch(r"[\x20-\x7E]+", new_password):
                show_warn(dialog, tr('common.error'), tr('error.password_format'))
                return

            try:
                if role == 'admin':
                    update_admin_basic(user_id, username=new_username, full_name=new_fullname, password=new_password)
                else:
                    update_user_basic(user_id, username=new_username, full_name=new_fullname, password=new_password)
                
                self.refresh_users()
                dialog.accept()
                show_info(self, tr('common.success'), tr('info.user_updated'))
            except Exception as e:
                show_warn(dialog, tr('common.error'), str(e))

        buttons.accepted.connect(save_changes)
        buttons.rejected.connect(dialog.reject)
        dialog.exec()

    def add_user(self):
        name = self.new_user.text().strip()
        pwd = self.new_pwd.text()
        role = self.new_role.currentText()
        if not name or not pwd:
            show_warn(self, tr('common.error'), tr('error.input_username_password'))
            return
        try:
            full_name = self.new_fullname.text().strip() or None
            if not re.fullmatch(r"[A-Za-z0-9_@.\-]+", name):
                show_warn(self, tr('common.error'), tr('error.username_format'))
                return
            if not re.fullmatch(r"[\x20-\x7E]+", pwd):
                show_warn(self, tr('common.error'), tr('error.password_format'))
                return
            if role == 'admin':
                create_admin(name, pwd, 1, full_name)
            else:
                create_user(name, pwd, role, 1, full_name)
            self.refresh_users()
            self.new_user.clear()
            self.new_pwd.clear()
            self.new_fullname.clear()
            show_info(self, tr('common.success'), tr('info.user_created'))
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))
    def delete_user(self, user_id):
        reply = ask_yes_no(self, tr('common.hint'), tr('confirm.delete_user'), default_yes=False)
        if reply == QMessageBox.StandardButton.Yes:
            try:
                delete_user(user_id)
                self.refresh_users()
                show_info(self, tr('common.success'), tr('info.user_deleted'))
            except Exception as e:
                show_warn(self, tr('common.error'), str(e))
    def toggle_user_role(self, user_id, current_role):
        if current_role == 'user':
            try:
                promote_user_to_admin(user_id)
                self.refresh_users()
                show_info(self, tr('common.success'), tr('info.user_role_updated', role='admin'))
            except Exception as e:
                show_warn(self, tr('common.error'), str(e))
        else:
            try:
                update_user_role(user_id, 'user')
                self.refresh_users()
                show_info(self, tr('common.success'), tr('info.user_role_updated', role='user'))
            except Exception as e:
                show_warn(self, tr('common.error'), str(e))
    def toggle_user_active(self, user_id, current_active):
        new_active = 0 if current_active == 1 else 1
        try:
            update_user_active(user_id, new_active)
            self.refresh_users()
            status = tr('admin.user.enable') if new_active == 1 else tr('admin.user.disable')
            show_info(self, tr('common.success'), tr('info.user_status_updated', status=status))
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))
    def on_user_item_changed(self, item):
        row = item.row()
        col = item.column()
        try:
            user_id = int(self.users_table.item(row, 0).text())
        except Exception:
            return
        role_tag = self.users_table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        if col == 1:
            username = item.text().strip()
            if not username:
                show_warn(self, tr('common.error'), '用户名不能为空')
                self.refresh_users()
                return
            if role_tag == 'admin':
                update_admin_basic(user_id, username=username)
            else:
                update_user_basic(user_id, username=username)
            self.refresh_users()
        elif col == 2:
            full_name = item.text().strip() or None
            if role_tag == 'admin':
                update_admin_basic(user_id, full_name=full_name)
            else:
                update_user_basic(user_id, full_name=full_name)
            self.refresh_users()
    def toggle_admin_active(self, admin_id, current_active):
        new_active = 0 if current_active == 1 else 1
        try:
            update_admin_active(admin_id, new_active)
            self.refresh_users()
            status = tr('admin.user.enable') if new_active == 1 else tr('admin.user.disable')
            show_info(self, tr('common.success'), tr('info.user_status_updated', status=status))
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))
    def delete_admin(self, admin_id):
        reply = ask_yes_no(self, tr('common.hint'), tr('confirm.delete_user'), default_yes=False)
        if reply == QMessageBox.StandardButton.Yes:
            try:
                delete_admin(admin_id)
                self.refresh_users()
                show_info(self, tr('common.success'), tr('info.user_deleted'))
            except Exception as e:
                show_warn(self, tr('common.error'), str(e))
    def demote_admin(self, admin_id):
        reply = ask_yes_no(self, tr('common.hint'), '确定要将该管理员设为普通用户吗？', default_yes=False)
        if reply == QMessageBox.StandardButton.Yes:
            try:
                demote_admin_to_user(admin_id)
                self.refresh_users()
                show_info(self, tr('common.success'), tr('info.user_role_updated', role='user'))
            except Exception as e:
                show_warn(self, tr('common.error'), str(e))
    def export_users_template(self):
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents/users_template')
        fn, sel = QFileDialog.getSaveFileName(self, '导出用户Excel模板', suggested, 'Excel (*.xlsx)')
        if not fn:
            return
        headers = ['用户名', '密码', '姓名', '角色', '状态']
        try:
            ext = os.path.splitext(fn)[1].lower()
            out = fn if ext == '.xlsx' else fn + '.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = 'Users'
            ws.append(headers)
            ws.append(['', '', '', 'user/admin', '1或0'])
            header_fill = PatternFill(start_color='FF409EFF', end_color='FF409EFF', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFFFF', size=13)
            data_font = Font(size=12)
            center = Alignment(horizontal='center', vertical='center')
            left = Alignment(horizontal='left', vertical='center')
            thin = Side(style='thin', color='FFDDDDDD')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for c in range(1, len(headers)+1):
                cell = ws.cell(row=1, column=c)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
            ws.row_dimensions[1].height = 26
            for r in range(2, ws.max_row+1):
                for c in range(1, len(headers)+1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = border
                    cell.font = data_font
                ws.cell(row=r, column=2).alignment = center
                ws.cell(row=r, column=4).alignment = center
                ws.cell(row=r, column=5).alignment = center
                ws.cell(row=r, column=1).alignment = left
                ws.cell(row=r, column=3).alignment = left
                ws.row_dimensions[r].height = 22
            widths = [0] * len(headers)
            for r in ws.iter_rows(values_only=True):
                for idx, val in enumerate(r):
                    l = len(str(val)) if val is not None else 0
                    widths[idx] = max(widths[idx], l)
            for i, w in enumerate(widths, start=1):
                letter = get_column_letter(i)
                ws.column_dimensions[letter].width = max(16, min(48, w + 6))
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
            ws.freeze_panes = 'A2'
            wb.save(out)
            show_info(self, tr('common.success'), tr('admin.export.users_tpl.done'))
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))
    def import_users_from_excel(self):
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents')
        fn, sel = QFileDialog.getOpenFileName(self, '选择用户Excel', suggested, 'Excel (*.xlsx)')
        if not fn:
            return
        try:
            wb = load_workbook(fn)
            ws = wb['Users'] if 'Users' in wb.sheetnames else wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            header = [str(x).strip() if x else '' for x in header_row]
            def idx(name):
                try:
                    return header.index(name)
                except Exception:
                    return -1
            iu = idx('用户名'); ip = idx('密码'); iname = idx('姓名'); ir = idx('角色'); ia = idx('状态')
            if min(iu, ip, ir, ia) < 0:
                show_warn(self, tr('common.error'), tr('admin.import.users.error.missing'))
                return
            ok = 0; fail = 0
            format_errs = []
            for idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    username = (str(r[iu]).strip() if iu >= 0 and iu < len(r) and r[iu] is not None else '')
                    password = (str(r[ip]).strip() if ip >= 0 and ip < len(r) and r[ip] is not None else '')
                    full_name = (str(r[iname]).strip() if iname >= 0 and iname < len(r) and r[iname] is not None else '') or None
                    role = (str(r[ir]).strip().lower() if ir >= 0 and ir < len(r) and r[ir] is not None else 'user')
                    active_str = (str(r[ia]).strip() if ia >= 0 and ia < len(r) and r[ia] is not None else '1')
                    active = 1 if active_str in ('1', '是', '启用', 'true', 'True') else 0
                    if not username or not password:
                        fail += 1
                        format_errs.append(f'第{idx}行：用户名或密码为空')
                        continue
                    if not re.fullmatch(r"[A-Za-z0-9_@.\-]+", username):
                        fail += 1
                        format_errs.append(f'第{idx}行：用户名格式错误')
                        continue
                    if not re.fullmatch(r"[\x20-\x7E]+", password):
                        fail += 1
                        format_errs.append(f'第{idx}行：密码格式错误')
                        continue
                    if role not in ('user', 'admin'):
                        role = 'user'
                    if role == 'admin':
                        create_admin(username, password, active, full_name)
                    else:
                        create_user(username, password, role, active, full_name)
                    ok += 1
                except Exception:
                    fail += 1
            self.refresh_users()
            show_info(self, tr('common.success'), tr('admin.import.users.result', ok=ok, fail=fail))
            if format_errs:
                detail = '\n'.join(format_errs[:20])
                show_warn(self, tr('admin.import.users.format_error'), detail)
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))
