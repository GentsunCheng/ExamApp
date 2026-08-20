import os
import json
import pathlib
from io import BytesIO
from PySide6.QtCore import Qt, QDateTime
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QFormLayout, QLineEdit,
    QTextEdit, QSpinBox, QDoubleSpinBox, QDateTimeEdit, QPushButton,
    QFileDialog, QTableWidget, QTableWidgetItem, QAbstractItemView,
    QCheckBox, QComboBox, QStackedWidget, QListWidget, QListWidgetItem,
    QScrollArea, QLabel
)
from PySide6.QtWidgets import QMessageBox
from PySide6.QtGui import QPixmap
from icon_manager import IconManager
from theme_manager import theme_manager
from language import tr
from utils import show_info, show_warn, ask_yes_no
from models import (
    list_exams, add_exam, import_questions_from_json, get_exam_stats,
    update_exam_title_desc, save_pic, update_question, delete_question,
    list_questions_by_pool, get_exam_random_pick_count,
    update_exam_random_pick_count, get_pic,
)
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


class QuestionPage(QWidget):
    """题目管理页：左侧上下两栏（必考题/随机题库），右侧查看与编辑题目"""

    LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    TYPE_COLORS = {
        'single': '#409eff',
        'multiple': '#9254de',
        'truefalse': '#e6a23c',
        'fill': '#13c2c2',
        'essay': '#67c23a',
    }

    def __init__(self, owner=None, parent=None):
        super().__init__(parent)
        self.owner = owner
        self.icon_manager = IconManager()
        self.exam_id = None
        self.exam_uuid = None
        self.exam_title = ''
        self.current_qid = None
        self._editing = False
        self._loading = False
        self._option_edits = []
        self._multi_checks = []
        self._correct_widget = None
        self._mandatory_qs = []
        self._random_qs = []
        self._build_ui()

    # ---------- UI ----------
    def _build_ui(self):
        colors = theme_manager.get_theme_colors()
        self.setStyleSheet(self._page_qss(colors))
        main_lay = QVBoxLayout(self)
        main_lay.setContentsMargins(12, 10, 12, 10)
        main_lay.setSpacing(10)

        # 顶部栏：返回 + 标题 + 随机抽取数量
        topbar = QHBoxLayout()
        topbar.setSpacing(10)
        self.back_btn = QPushButton(tr('admin.questions.back'))
        self.back_btn.setObjectName('btn-secondary')
        self.back_btn.setIcon(self.icon_manager.get_icon('back'))
        self.back_btn.clicked.connect(self.on_back)
        topbar.addWidget(self.back_btn)
        self.title_label = QLabel('')
        self.title_label.setStyleSheet(f"font-size:17px; font-weight:bold; color:{colors['primary']};")
        topbar.addWidget(self.title_label)
        topbar.addStretch()
        pick_lab = QLabel(tr('admin.questions.random_pick') + ':')
        pick_lab.setStyleSheet(f"color:{colors['text_secondary']}; font-size:13px; font-weight:600;")
        topbar.addWidget(pick_lab)
        self.pick_spin = QSpinBox()
        self.pick_spin.setRange(0, 999)
        self.pick_spin.setMinimumWidth(90)
        self.pick_spin.valueChanged.connect(self.on_pick_changed)
        topbar.addWidget(self.pick_spin)
        main_lay.addLayout(topbar)

        body = QHBoxLayout()
        body.setSpacing(12)

        # 左侧：必考题 / 随机题库
        left_panel = QWidget()
        left_panel.setFixedWidth(310)
        left_lay = QVBoxLayout(left_panel)
        left_lay.setContentsMargins(0, 0, 0, 0)
        left_lay.setSpacing(10)
        self.mand_group = QGroupBox(tr('admin.questions.mandatory_group'))
        m_lay = QVBoxLayout(self.mand_group)
        m_lay.setContentsMargins(6, 8, 6, 6)
        m_lay.setSpacing(0)
        self.mand_list = QListWidget()
        self.mand_list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.mand_list.itemActivated.connect(self.on_question_activated)
        self.mand_list.currentRowChanged.connect(self._on_list_row_changed(self.mand_list))
        m_lay.addWidget(self.mand_list)
        self.rand_group = QGroupBox(tr('admin.questions.random_group'))
        r_lay = QVBoxLayout(self.rand_group)
        r_lay.setContentsMargins(6, 8, 6, 6)
        r_lay.setSpacing(0)
        self.rand_list = QListWidget()
        self.rand_list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.rand_list.itemActivated.connect(self.on_question_activated)
        self.rand_list.currentRowChanged.connect(self._on_list_row_changed(self.rand_list))
        r_lay.addWidget(self.rand_list)
        left_lay.addWidget(self.mand_group, 1)
        left_lay.addWidget(self.rand_group, 1)
        body.addWidget(left_panel)

        # 右侧：详情 / 编辑
        right_panel = QWidget()
        right_lay = QVBoxLayout(right_panel)
        right_lay.setContentsMargins(0, 0, 0, 0)
        right_lay.setSpacing(10)

        self.hint_label = QLabel(tr('admin.questions.view_hint'))
        self.hint_label.setObjectName('hint-chip')
        right_lay.addWidget(self.hint_label)

        # 元信息徽章：类型 / 分数 / 题库
        self.meta_bar = QHBoxLayout()
        self.meta_bar.setSpacing(8)
        self.meta_type_badge = QLabel('')
        self.meta_score_badge = QLabel('')
        self.meta_pool_badge = QLabel('')
        for lab in (self.meta_type_badge, self.meta_score_badge, self.meta_pool_badge):
            lab.setVisible(False)
        self.meta_bar.addWidget(self.meta_type_badge)
        self.meta_bar.addWidget(self.meta_score_badge)
        self.meta_bar.addWidget(self.meta_pool_badge)
        self.meta_bar.addStretch()
        right_lay.addLayout(self.meta_bar)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        scroll.setStyleSheet(f"QScrollArea {{ border:1px solid {colors['border']}; border-radius:12px; background-color:transparent; }}")
        content = QWidget()
        content.setStyleSheet("background-color: transparent;")
        form_lay = QVBoxLayout(content)
        form_lay.setContentsMargins(16, 16, 16, 16)
        form_lay.setSpacing(8)

        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        form.setVerticalSpacing(14)
        form.setHorizontalSpacing(18)

        self.type_combo = QComboBox()
        for tkey in ('single', 'multiple', 'truefalse', 'fill', 'essay'):
            self.type_combo.addItem(tr('exam.type.' + tkey), tkey)
        self.type_combo.currentIndexChanged.connect(self.on_type_changed)
        self.type_combo.setMinimumWidth(170)
        form.addRow(tr('admin.questions.type'), self.type_combo)

        self.score_spin = QDoubleSpinBox()
        self.score_spin.setRange(0, 999)
        self.score_spin.setDecimals(1)
        self.score_spin.setSingleStep(1.0)
        self.score_spin.setValue(1.0)
        self.score_spin.setMinimumWidth(170)
        self.score_spin.valueChanged.connect(self.on_score_changed)
        form.addRow(tr('admin.questions.score'), self.score_spin)

        self.pool_combo = QComboBox()
        self.pool_combo.addItem(tr('admin.questions.pool.mandatory'), 'mandatory')
        self.pool_combo.addItem(tr('admin.questions.pool.random'), 'random')
        self.pool_combo.setMinimumWidth(170)
        form.addRow(tr('admin.questions.pool'), self.pool_combo)

        self.text_edit = QTextEdit()
        self.text_edit.setPlaceholderText(tr('admin.questions.content'))
        self.text_edit.setMinimumHeight(110)
        form.addRow(tr('admin.questions.content'), self.text_edit)

        # 选项编辑区（单选/多选可见）
        self.options_container = QWidget()
        opts_lay = QVBoxLayout(self.options_container)
        opts_lay.setContentsMargins(0, 0, 0, 0)
        opts_lay.setSpacing(8)
        for i in range(4):
            row = QHBoxLayout()
            row.setContentsMargins(0, 0, 0, 0)
            row.setSpacing(8)
            lab = QLabel(self.LETTERS[i] + '.')
            lab.setFixedWidth(26)
            lab.setStyleSheet(f"font-size:14px; font-weight:700; color:{colors['primary']};")
            le = QLineEdit()
            le.setPlaceholderText(tr('admin.questions.options') + f" {self.LETTERS[i]}")
            row.addWidget(lab)
            row.addWidget(le, 1)
            opts_lay.addLayout(row)
            self._option_edits.append(le)
        form.addRow(tr('admin.questions.options'), self.options_container)

        # 正确答案区（按题型动态切换）
        self.correct_container = QWidget()
        self.correct_container.setLayout(QVBoxLayout())
        self.correct_container.layout().setContentsMargins(0, 0, 0, 0)
        form.addRow(tr('admin.questions.correct'), self.correct_container)
        self.rebuild_correct_widget()

        # 表单字段标签样式
        for i in range(form.rowCount()):
            item = form.itemAt(i, QFormLayout.LabelRole)
            if item is not None and item.widget() is not None:
                item.widget().setStyleSheet(
                    f"font-size:14px; color:{colors['text_secondary']}; font-weight:600;"
                )

        form_lay.addLayout(form)

        # 图片展示区（仅查看模式）
        self.picture_widget = QWidget()
        pic_lay = QVBoxLayout(self.picture_widget)
        pic_lay.setContentsMargins(0, 8, 0, 0)
        self.picture_title = QLabel('')
        self.picture_title.setStyleSheet(f"font-weight:bold; color:{colors['text_secondary']}; font-size:13px;")
        self.picture_rows = QHBoxLayout()
        self.picture_rows.setSpacing(10)
        pic_lay.addWidget(self.picture_title)
        pic_lay.addLayout(self.picture_rows)
        self.picture_widget.setVisible(False)
        form_lay.addWidget(self.picture_widget)
        form_lay.addStretch()

        scroll.setWidget(content)
        right_lay.addWidget(scroll, 1)

        # 操作按钮
        btn_lay = QHBoxLayout()
        btn_lay.setSpacing(10)
        self.add_btn = QPushButton(tr('admin.questions.add_btn'))
        self.add_btn.setObjectName('btn-primary')
        self.add_btn.setIcon(self.icon_manager.get_icon('exam_add'))
        self.add_btn.clicked.connect(self.new_question)
        self.edit_btn = QPushButton(tr('common.edit'))
        self.edit_btn.setObjectName('btn-secondary')
        self.edit_btn.clicked.connect(self.edit_current)
        self.save_btn = QPushButton(tr('admin.questions.save_btn'))
        self.save_btn.setObjectName('btn-success')
        self.save_btn.setIcon(self.icon_manager.get_icon('save'))
        self.save_btn.clicked.connect(self.save_question)
        self.delete_btn = QPushButton(tr('admin.questions.delete_btn'))
        self.delete_btn.setObjectName('btn-danger')
        self.delete_btn.setIcon(self.icon_manager.get_icon('exam_delete'))
        self.delete_btn.clicked.connect(self.delete_current)
        btn_lay.addWidget(self.add_btn)
        btn_lay.addWidget(self.edit_btn)
        btn_lay.addWidget(self.save_btn)
        btn_lay.addWidget(self.delete_btn)
        btn_lay.addStretch()
        right_lay.addLayout(btn_lay)

        body.addWidget(right_panel, 1)
        main_lay.addLayout(body, 1)
        self.set_form_enabled(False)
        self.edit_btn.setEnabled(False)
        self.save_btn.setEnabled(False)
        self.delete_btn.setEnabled(False)

    # ---------- 数据加载 ----------
    def open_exam(self, exam_id, exam_uuid, title):
        self.exam_id = exam_id
        self.exam_uuid = exam_uuid
        self.exam_title = title or ''
        self.title_label.setText(f"{tr('admin.questions.title')} - {self.exam_title}")
        self.reload_lists()
        self.clear_form()
        self.set_edit_mode(False)

    def reload_lists(self):
        if not self.exam_uuid:
            return
        self._mandatory_qs = list_questions_by_pool(self.exam_uuid, 'mandatory')
        self._random_qs = list_questions_by_pool(self.exam_uuid, 'random')
        self._fill_list(self.mand_list, self._mandatory_qs, self.mand_group,
                        tr('admin.questions.mandatory_group'))
        self._fill_list(self.rand_list, self._random_qs, self.rand_group,
                        tr('admin.questions.random_group'))
        try:
            self.pick_spin.blockSignals(True)
            self.pick_spin.setValue(get_exam_random_pick_count(self.exam_uuid))
            self.pick_spin.blockSignals(False)
        except Exception:
            pass

    def _fill_list(self, lst, questions, group, base_title):
        lst.blockSignals(True)
        lst.clear()
        for i, q in enumerate(questions, start=1):
            item = QListWidgetItem()
            item.setData(Qt.ItemDataRole.UserRole, q['id'])
            item.setToolTip(
                (q.get('text') or '')
                + f"\n[{tr('exam.type.' + str(q.get('type')))}] "
                + f"{tr('admin.questions.score')}: {q.get('score', '')}"
            )
            lst.addItem(item)
            row_w = self._make_question_row(q, i)
            item.setSizeHint(row_w.sizeHint())
            lst.setItemWidget(item, row_w)
        lst.blockSignals(False)
        group.setTitle(f"{base_title} ({len(questions)})")

    # ---------- 列表行 / 徽章样式 ----------
    def _page_qss(self, c):
        return f"""
        QLabel {{ color: {c['text_primary']}; }}
        QGroupBox {{
            border: 1px solid {c['border']}; border-radius: 12px;
            margin-top: 12px; padding-top: 8px;
            font-size: 13px; font-weight: 600; color: {c['text_secondary']};
        }}
        QGroupBox::title {{ subcontrol-origin: margin; left: 12px; padding: 0 6px; }}
        QListWidget {{
            background: {c['card_background']}; color: {c['text_primary']};
            border: 1px solid {c['border']}; border-radius: 10px; padding: 4px; font-size: 13px;
        }}
        QListWidget::item {{ border-radius: 8px; margin: 2px; }}
        QListWidget::item:hover {{ background: {c['border_light']}; }}
        QListWidget::item:selected {{ background: {c['primary']}; }}
        QLineEdit, QTextEdit, QComboBox, QDoubleSpinBox, QSpinBox {{
            background: {c['input_background']}; color: {c['text_primary']};
            border: 1px solid {c['input_border']}; border-radius: 8px; padding: 7px 10px;
            selection-background-color: {c['primary']}; selection-color: {c['text_inverse']};
            font-size: 14px;
        }}
        QLineEdit:focus, QTextEdit:focus, QComboBox:focus, QDoubleSpinBox:focus, QSpinBox:focus {{
            border: 1px solid {c['primary']};
        }}
        QLineEdit:disabled, QTextEdit:disabled, QComboBox:disabled, QDoubleSpinBox:disabled, QSpinBox:disabled {{
            background: transparent; color: {c['text_primary']}; border: 1px dashed {c['border']};
        }}
        QComboBox::drop-down {{ width: 26px; border: none; }}
        QComboBox QAbstractItemView {{
            background: {c['card_background']}; color: {c['text_primary']};
            border: 1px solid {c['border']}; border-radius: 8px; padding: 4px;
            selection-background-color: {c['primary']}; selection-color: {c['text_inverse']}; outline: 0;
        }}
        QDoubleSpinBox::up-button, QDoubleSpinBox::down-button, QSpinBox::up-button, QSpinBox::down-button {{
            width: 20px; border: none; background: {c['border_light']}; border-radius: 6px; margin: 1px;
        }}
        QDoubleSpinBox::up-button:hover, QDoubleSpinBox::down-button:hover, QSpinBox::up-button:hover, QSpinBox::down-button:hover {{
            background: {c['button_primary_hover']};
        }}
        QCheckBox {{ color: {c['text_primary']}; font-size: 14px; spacing: 6px; }}
        QCheckBox::indicator {{ width: 18px; height: 18px; border-radius: 5px; border: 1px solid {c['input_border']}; background: {c['input_background']}; }}
        QCheckBox::indicator:checked {{ background: {c['primary']}; border-color: {c['primary']}; }}
        QPushButton {{ border: none; border-radius: 8px; padding: 8px 18px; font-size: 14px; font-weight: 600; }}
        QPushButton#btn-primary {{ background: {c['primary']}; color: {c['text_inverse']}; }}
        QPushButton#btn-primary:hover {{ background: {c['button_primary_hover']}; }}
        QPushButton#btn-secondary {{ background: {c['card_background']}; color: {c['text_primary']}; border: 1px solid {c['border']}; }}
        QPushButton#btn-secondary:hover {{ background: {c['border_light']}; }}
        QPushButton#btn-success {{ background: {c['success']}; color: {c['text_inverse']}; }}
        QPushButton#btn-success:hover {{ background: #85ce61; }}
        QPushButton#btn-danger {{ background: {c['error']}; color: {c['text_inverse']}; }}
        QPushButton#btn-danger:hover {{ background: #f78989; }}
        QPushButton:disabled {{ background: {c['border_light']}; color: {c['text_tertiary']}; }}
        QLabel#hint-chip {{
            background: {c['info_light']}; color: {c['primary']}; font-size: 13px;
            border-radius: 8px; padding: 7px 12px;
        }}
        """

    def _badge_style(self, accent):
        colors = theme_manager.get_theme_colors()
        return (
            f"background:{colors['card_background']}; color:{accent}; "
            f"border:1px solid {accent}; font-size:12px; font-weight:600; "
            f"padding:2px 10px; border-radius:10px;"
        )

    def _type_badge(self, qtype):
        colors = theme_manager.get_theme_colors()
        accent = self.TYPE_COLORS.get(str(qtype), colors['primary'])
        lab = QLabel(tr('exam.type.' + str(qtype)))
        lab._qtype = str(qtype)
        lab.setStyleSheet(self._badge_style(accent))
        return lab

    def _make_question_row(self, q, idx):
        colors = theme_manager.get_theme_colors()
        container = QWidget()
        container.setObjectName('qrow')
        container.setMinimumHeight(40)
        lay = QHBoxLayout(container)
        lay.setContentsMargins(6, 7, 8, 7)
        lay.setSpacing(10)
        num_lab = QLabel(str(idx))
        num_lab.setFixedWidth(24)
        num_lab.setAlignment(Qt.AlignmentFlag.AlignCenter)
        num_lab.setStyleSheet(f"color:{colors['text_tertiary']}; font-size:13px; font-weight:600;")
        badge = self._type_badge(q.get('type'))
        text = (q.get('text') or '').replace('\n', ' ')
        if len(text) > 34:
            text = text[:34] + '...'
        preview = QLabel(text)
        preview.setStyleSheet(f"color:{colors['text_primary']}; font-size:13px;")
        lay.addWidget(num_lab)
        lay.addWidget(badge)
        lay.addWidget(preview, 1)
        container._q_labels = [num_lab, preview]
        container._q_badge = badge
        return container

    def _style_row_widget(self, w, selected):
        colors = theme_manager.get_theme_colors()
        accent = self.TYPE_COLORS.get(w._q_badge._qtype, colors['primary'])
        if selected:
            w.setStyleSheet(f"QWidget#qrow {{ background:{colors['primary']}; border-radius:8px; }}")
            w._q_labels[0].setStyleSheet(f"color:{colors['text_inverse']}; font-size:13px; font-weight:600;")
            w._q_labels[1].setStyleSheet(f"color:{colors['text_inverse']}; font-size:13px;")
            w._q_badge.setStyleSheet(
                f"color:{colors['text_inverse']}; background:transparent; border:none; "
                f"font-size:12px; font-weight:600; padding:2px 10px;"
            )
        else:
            w.setStyleSheet("QWidget#qrow { background:transparent; border-radius:8px; }")
            w._q_labels[0].setStyleSheet(f"color:{colors['text_tertiary']}; font-size:12px; font-weight:600;")
            w._q_labels[1].setStyleSheet(f"color:{colors['text_primary']}; font-size:13px;")
            w._q_badge.setStyleSheet(self._badge_style(accent))

    def _on_list_row_changed(self, lst):
        def handler(current, previous):
            for row in (current, previous):
                item = lst.item(row)
                if item is None:
                    continue
                w = lst.itemWidget(item)
                if w is not None:
                    self._style_row_widget(w, row == current)
        return handler

    def _update_meta_badges(self, q):
        colors = theme_manager.get_theme_colors()
        qtype = str(q.get('type'))
        self.meta_type_badge.setText(tr('exam.type.' + qtype))
        self.meta_type_badge.setStyleSheet(self._badge_style(self.TYPE_COLORS.get(qtype, colors['primary'])))
        self.meta_type_badge.setVisible(True)
        self.meta_score_badge.setText(f"{tr('admin.questions.score')}: {q.get('score', 0)}")
        self.meta_score_badge.setStyleSheet(self._badge_style(colors['info']))
        self.meta_score_badge.setVisible(True)
        pool = q.get('pool') or 'mandatory'
        self.meta_pool_badge.setText(tr('admin.questions.pool.' + pool))
        self.meta_pool_badge.setStyleSheet(
            self._badge_style(colors['success'] if pool == 'mandatory' else colors['warning'])
        )
        self.meta_pool_badge.setVisible(True)

    def _clear_meta_badges(self):
        for lab in (self.meta_type_badge, self.meta_score_badge, self.meta_pool_badge):
            lab.setText('')
            lab.setStyleSheet('')
            lab.setVisible(False)

    # ---------- 查看 / 编辑 ----------
    def on_question_activated(self, item):
        qid = item.data(Qt.ItemDataRole.UserRole)
        if qid is not None:
            self.load_question(qid)

    def _find_question(self, qid):
        for q in self._mandatory_qs:
            if q['id'] == qid:
                return q, self.mand_list
        for q in self._random_qs:
            if q['id'] == qid:
                return q, self.rand_list
        return None, None

    def load_question(self, qid):
        q, lst = self._find_question(qid)
        if q is None:
            return
        self._loading = True
        try:
            self.current_qid = qid
            idx = self.type_combo.findData(q.get('type'))
            self.type_combo.setCurrentIndex(max(0, idx))
            self.score_spin.setValue(float(q.get('score') or 0))
            pool = q.get('pool') or 'mandatory'
            pidx = self.pool_combo.findData(pool)
            self.pool_combo.setCurrentIndex(max(0, pidx))
            self.text_edit.setPlainText(q.get('text') or '')
            opts = q.get('options') or []
            for i, le in enumerate(self._option_edits):
                if i < len(opts):
                    le.setText(str(opts[i].get('text') or '') if isinstance(opts[i], dict) else str(opts[i]))
                else:
                    le.clear()
            self.rebuild_correct_widget()
            self.fill_correct(q.get('correct') or [])
            self._show_pictures(q.get('pictures') or '')
            self._update_meta_badges(q)
            # 高亮左侧对应题目
            if lst is not None:
                for i in range(lst.count()):
                    it = lst.item(i)
                    if it and it.data(Qt.ItemDataRole.UserRole) == qid:
                        lst.setCurrentRow(i)
                        break
            self.set_edit_mode(False)
        finally:
            self._loading = False

    def _show_pictures(self, pictures):
        while self.picture_rows.count():
            item = self.picture_rows.takeAt(0)
            if item.widget():
                item.widget().setParent(None)
        try:
            hashes = json.loads(pictures) if pictures else []
        except Exception:
            hashes = []
        if not hashes:
            self.picture_widget.setVisible(False)
            return
        colors = theme_manager.get_theme_colors()
        self.picture_title.setText(tr('admin.questions.pictures') + ':')
        for h in hashes:
            img = get_pic(h)
            if img:
                lab = QLabel()
                lab.setAlignment(Qt.AlignmentFlag.AlignCenter)
                pm = QPixmap.fromImage(img)
                if pm.width() > 400:
                    pm = pm.scaledToWidth(400, Qt.TransformationMode.SmoothTransformation)
                lab.setPixmap(pm)
                self.picture_rows.addWidget(lab)
        self.picture_widget.setVisible(True)

    def clear_form(self):
        self._loading = True
        try:
            self.current_qid = None
            self.type_combo.setCurrentIndex(0)
            self.score_spin.setValue(1.0)
            self.text_edit.clear()
            for le in self._option_edits:
                le.clear()
            self.pool_combo.setCurrentIndex(0)
            self.rebuild_correct_widget()
            self.picture_widget.setVisible(False)
            self._clear_meta_badges()
        finally:
            self._loading = False

    def new_question(self):
        self.clear_form()
        self.set_edit_mode(True)

    def edit_current(self):
        if self.current_qid is None:
            show_warn(self, tr('common.error'), tr('admin.questions.select_first'))
            return
        self.set_edit_mode(True)

    def set_edit_mode(self, editing):
        self._editing = editing
        self.set_form_enabled(editing)
        if editing:
            self.hint_label.setText(tr('admin.questions.new_question') if self.current_qid is None else tr('admin.questions.view_hint'))
            self.edit_btn.setEnabled(False)
            self.save_btn.setEnabled(True)
            self.delete_btn.setEnabled(self.current_qid is not None)
        else:
            self.hint_label.setText(tr('admin.questions.view_hint'))
            self.edit_btn.setEnabled(self.current_qid is not None)
            self.save_btn.setEnabled(False)
            self.delete_btn.setEnabled(self.current_qid is not None)

    def set_form_enabled(self, enabled):
        for w in (self.type_combo, self.pool_combo, self.text_edit):
            w.setEnabled(enabled)
        for le in self._option_edits:
            le.setEnabled(enabled)
        if self._correct_widget is not None:
            self._correct_widget.setEnabled(enabled)
        # 分数始终可编辑（查看模式下修改后自动保存）
        self.score_spin.setEnabled(True)

    def on_score_changed(self, value):
        if self._loading or self.current_qid is None or self._editing:
            return
        q, _lst = self._find_question(self.current_qid)
        if q is None:
            return
        try:
            update_question(
                self.current_qid,
                q.get('type'), q.get('text') or '',
                q.get('options') or [], q.get('correct') or [],
                float(value), q.get('pool') or 'mandatory',
            )
            q['score'] = float(value)
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))
            return
        self._update_meta_badges(q)

    # ---------- 正确答案控件 ----------
    def rebuild_correct_widget(self):
        if self._correct_widget is not None:
            self._correct_widget.setParent(None)
            self._correct_widget.deleteLater()
            self._correct_widget = None
        self._multi_checks = []
        lay = self.correct_container.layout()
        while lay.count():
            item = lay.takeAt(0)
            if item.widget():
                item.widget().setParent(None)
        qtype = self.type_combo.currentData()
        colors = theme_manager.get_theme_colors()
        if qtype == 'single':
            w = QComboBox()
            for i in range(4):
                text = self._option_edits[i].text().strip()
                if text:
                    w.addItem(self.LETTERS[i], self.LETTERS[i])
        elif qtype == 'multiple':
            w = QWidget()
            wl = QHBoxLayout(w)
            wl.setContentsMargins(0, 0, 0, 0)
            wl.setSpacing(14)
            for i in range(4):
                if not self._option_edits[i].text().strip():
                    continue
                cb = QCheckBox(self.LETTERS[i])
                self._multi_checks.append(cb)
                wl.addWidget(cb)
            wl.addStretch()
        elif qtype == 'truefalse':
            w = QComboBox()
            w.addItem(tr('exam.true'), True)
            w.addItem(tr('exam.false'), False)
        elif qtype == 'fill':
            w = QLineEdit()
            w.setPlaceholderText(tr('admin.questions.fill_hint'))
        else:  # essay
            w = QLineEdit()
            w.setPlaceholderText(tr('admin.questions.correct_none'))
            w.setReadOnly(True)
        self._correct_widget = w
        w.setEnabled(self._editing)
        lay.addWidget(w)

    def fill_correct(self, correct):
        qtype = self.type_combo.currentData()
        w = self._correct_widget
        if w is None:
            return
        if qtype == 'single':
            idx = w.findData(correct[0]) if correct else -1
            w.setCurrentIndex(max(0, idx))
        elif qtype == 'multiple':
            sset = set(str(c) for c in correct)
            for cb in self._multi_checks:
                cb.setChecked(cb.text() in sset)
        elif qtype == 'truefalse':
            idx = w.findData(bool(correct[0])) if correct else 0
            w.setCurrentIndex(max(0, idx))
        elif qtype == 'fill':
            w.setText(' / '.join(str(c) for c in correct))

    def collect_correct(self):
        qtype = self.type_combo.currentData()
        w = self._correct_widget
        if qtype == 'single':
            return [str(w.currentData())] if w and w.currentData() is not None else []
        if qtype == 'multiple':
            return [cb.text() for cb in self._multi_checks if cb.isChecked()]
        if qtype == 'truefalse':
            return [bool(w.currentData())] if w and w.currentData() is not None else []
        if qtype == 'fill':
            parts = [p.strip() for p in w.text().replace('，', '/').replace(';', '/').split('/') if p.strip()] if w else []
            return parts
        return []

    def collect_options(self):
        qtype = self.type_combo.currentData()
        if qtype not in ('single', 'multiple'):
            return []
        opts = []
        for i, le in enumerate(self._option_edits):
            text = le.text().strip()
            if text:
                opts.append({'key': self.LETTERS[i], 'text': text})
        return opts

    def on_type_changed(self, _idx):
        # 非选项类题型隐藏选项编辑区
        qtype = self.type_combo.currentData()
        self.options_container.setVisible(qtype in ('single', 'multiple'))
        self.rebuild_correct_widget()

    def on_pick_changed(self, value):
        if self.exam_id is not None:
            update_exam_random_pick_count(self.exam_id, value)

    # ---------- 操作 ----------
    def save_question(self):
        text = self.text_edit.toPlainText().strip()
        if not text:
            show_warn(self, tr('common.error'), tr('admin.questions.text_required'))
            return
        qtype = self.type_combo.currentData()
        score = self.score_spin.value()
        pool = self.pool_combo.currentData()
        options = self.collect_options()
        correct = self.collect_correct()
        if qtype in ('single', 'multiple'):
            if not options:
                show_warn(self, tr('common.error'), tr('admin.questions.no_options'))
                return
            keys = {o['key'] for o in options}
            if not correct or not set(correct).issubset(keys):
                show_warn(self, tr('common.error'), tr('admin.questions.invalid_correct'))
                return
            if qtype == 'single' and len(correct) != 1:
                show_warn(self, tr('common.error'), tr('admin.questions.invalid_correct'))
                return
        elif qtype == 'fill':
            if not correct:
                show_warn(self, tr('common.error'), tr('admin.questions.invalid_correct'))
                return
        try:
            if self.current_qid is None:
                new_id = add_question(self.exam_uuid, qtype, text, options, correct, score, pool)
                qid = new_id
            else:
                update_question(self.current_qid, qtype, text, options, correct, score, pool)
                qid = self.current_qid
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))
            return
        self.reload_lists()
        self._select_by_id(qid, pool)
        self.load_question(qid)
        self.set_edit_mode(False)
        show_info(self, tr('common.success'), tr('admin.questions.saved'))

    def _select_by_id(self, qid, pool):
        lst = self.mand_list if pool == 'mandatory' else self.rand_list
        for i in range(lst.count()):
            it = lst.item(i)
            if it and it.data(Qt.ItemDataRole.UserRole) == qid:
                lst.setCurrentRow(i)
                break

    def delete_current(self):
        if self.current_qid is None:
            show_warn(self, tr('common.error'), tr('admin.questions.select_first'))
            return
        reply = ask_yes_no(self, tr('common.hint'), tr('admin.questions.delete_confirm'), default_yes=False)
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            delete_question(self.current_qid)
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))
            return
        self.reload_lists()
        self.clear_form()
        self.set_edit_mode(False)
        show_info(self, tr('common.success'), tr('admin.questions.deleted'))

    def on_back(self):
        if self.owner is not None:
            self.owner.back_to_exams()


class AdminExamsModule(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.icon_manager = IconManager()
        self.stack = QStackedWidget()
        self.exam_list_page = self._build_exam_list_page()
        self.question_page = QuestionPage(owner=self)
        self.stack.addWidget(self.exam_list_page)
        self.stack.addWidget(self.question_page)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.addWidget(self.stack)

    def _build_exam_list_page(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(0, 0, 0, 0)
        gb1 = QGroupBox(tr('admin.exams_group'))
        vb1 = QVBoxLayout()
        self.exams_table = QTableWidget(0, 9)
        self.exams_table.setHorizontalHeaderLabels([tr('admin.exams.headers.id'), tr('admin.exams.headers.title'), tr('admin.exams.headers.pass_ratio'), tr('admin.exams.headers.time_limit'), tr('admin.exams.headers.deadline'), tr('admin.exams.headers.description'), tr('admin.exams.headers.q_count'), tr('admin.exams.headers.total'), tr('admin.exams.headers.actions')])
        self.exams_table.setColumnWidth(0, 50)
        self.exams_table.setColumnWidth(1, 120)
        self.exams_table.setColumnWidth(2, 120)
        self.exams_table.setColumnWidth(3, 120)
        self.exams_table.setColumnWidth(4, 120)
        self.exams_table.setColumnWidth(5, 480)
        self.exams_table.setColumnWidth(6, 80)
        self.exams_table.setColumnWidth(7, 80)
        self.exams_table.setColumnWidth(8, 210)
        self.exams_table.horizontalHeader().setStretchLastSection(True)
        self.exams_table.setAlternatingRowColors(True)
        self.exams_table.setShowGrid(False)
        self.exams_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.exams_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.exams_table.itemChanged.connect(self.on_exam_item_changed)
        self.exams_table.cellDoubleClicked.connect(self.on_exam_double_clicked)
        self.refresh_exams()
        vb1.addWidget(self.exams_table)
        colors = theme_manager.get_theme_colors()
        hint = QLabel(tr('admin.questions.open_hint'))
        hint.setStyleSheet(f"color:{colors['text_secondary']}; font-size:12px; padding:2px 0;")
        vb1.addWidget(hint)
        gb1.setLayout(vb1)
        gb2 = QGroupBox(tr('admin.new_exam_group'))
        vb2 = QVBoxLayout()
        form = QFormLayout()
        self.ex_title = QLineEdit()
        self.ex_title.setPlaceholderText(tr('admin.exams.form.title'))
        self.ex_desc = QTextEdit()
        self.ex_desc.setPlaceholderText(tr('admin.exams.form.description'))
        self.ex_desc.setMaximumHeight(80)
        self.ex_pass = QSpinBox()
        self.ex_pass.setRange(0, 100)
        self.ex_pass.setValue(60)
        colors_inputs = theme_manager.get_theme_colors()
        spin_style = (
            f"QSpinBox {{ padding:6px 10px; border:1px solid {colors_inputs['input_border']}; border-radius:12px; background-color:{colors_inputs['input_background']}; color:{colors_inputs['text_primary']}; }}\n"
            f"QSpinBox:focus {{ border-color:{colors_inputs['primary']}; }}\n"
            f"QSpinBox::up-button, QSpinBox::down-button {{ width:20px; border:none; background-color:{colors_inputs['border_light']}; border-radius:8px; }}\n"
            f"QSpinBox::up-button:hover, QSpinBox::down-button:hover {{ background-color:{colors_inputs['button_primary_hover']}; }}"
        )
        self.ex_pass.setStyleSheet(spin_style)
        self.ex_time = QSpinBox()
        self.ex_time.setRange(1, 600)
        self.ex_time.setValue(60)
        self.ex_time.setStyleSheet(spin_style)
        self.ex_end = QDateTimeEdit()
        self.ex_end.setDateTime(QDateTime.currentDateTime())
        self.ex_end.setDisplayFormat('yyyy-MM-dd HH:mm')
        self.ex_end.setCalendarPopup(True)
        dt_style = (
            f"QDateTimeEdit {{ padding:6px 10px; border:1px solid {colors_inputs['input_border']}; border-radius:12px; background-color:{colors_inputs['input_background']}; color:{colors_inputs['text_primary']}; }}\n"
            f"QDateTimeEdit:focus {{ border-color:{colors_inputs['primary']}; }}\n"
            f"QDateTimeEdit::up-button, QDateTimeEdit::down-button {{ width:22px; border:none; background-color:{colors_inputs['border_light']}; border-radius:10px; margin:2px; }}\n"
            f"QDateTimeEdit::up-button:hover, QDateTimeEdit::down-button:hover {{ background-color:{colors_inputs['button_primary_hover']}; }}\n"
            f"QDateTimeEdit::up-button:pressed, QDateTimeEdit::down-button:pressed {{ background-color:{colors_inputs['primary']}; }}\n"
            f"QDateTimeEdit::up-arrow, QDateTimeEdit::down-arrow {{ width: 0; height: 0; }}"
        )
        self.ex_end.setStyleSheet(dt_style)
        cal = self.ex_end.calendarWidget()
        cal.setStyleSheet(
            f"QCalendarWidget {{ background-color:{colors_inputs['card_background']}; border:1px solid {colors_inputs['border']}; border-radius:8px; }}\n"
            f"QCalendarWidget QWidget#qt_calendar_navigationbar {{ background-color:{colors_inputs['card_background']}; border:none; padding:6px; }}\n"
            f"QCalendarWidget QToolButton#qt_calendar_prevmonth, QCalendarWidget QToolButton#qt_calendar_nextmonth {{ background-color:{colors_inputs['button_primary']}; color:{colors_inputs['text_inverse']}; border:none; border-radius:6px; padding:4px 8px; }}\n"
            f"QCalendarWidget QToolButton#qt_calendar_prevmonth:hover, QCalendarWidget QToolButton#qt_calendar_nextmonth:hover {{ background-color:{colors_inputs['button_primary_hover']}; }}\n"
            f"QCalendarWidget QToolButton#qt_calendar_monthbutton {{ background-color:{colors_inputs['border_light']}; color:{colors_inputs['text_primary']}; border:none; border-radius:6px; padding:4px 10px; }}\n"
            f"QCalendarWidget QToolButton#qt_calendar_monthbutton:hover {{ background-color:{colors_inputs['button_primary_hover']}; color:{colors_inputs['text_inverse']}; }}\n"
            f"QCalendarWidget QSpinBox#qt_calendar_yearspinbox {{ background-color:{colors_inputs['input_background']}; color:{colors_inputs['text_primary']}; border:1px solid {colors_inputs['input_border']}; border-radius:6px; padding:2px 6px; }}\n"
            f"QCalendarWidget QTableView {{ background-color:{colors_inputs['card_background']}; alternate-background-color:{colors_inputs['border_light']}; selection-background-color:{colors_inputs['primary']}; selection-color:{colors_inputs['text_inverse']}; gridline-color:{colors_inputs['border']}; }}\n"
            f"QCalendarWidget QTableView::item {{ padding:4px; }}\n"
            f"QCalendarWidget QTableView::item:hover {{ background-color:{colors_inputs['border_light']}; }}\n"
            f"QCalendarWidget QTableView::item:selected {{ background-color:{colors_inputs['primary']}; color:{colors_inputs['text_inverse']}; }}"
        )
        self.ex_permanent = QCheckBox(tr('admin.exams.permanent_checkbox'))
        colors_perm = theme_manager.get_theme_colors()
        self.ex_permanent.setStyleSheet(
            f"QCheckBox {{ color:{colors_perm['text_primary']}; font-size:14px; }}\n"
            f"QCheckBox::indicator {{ width:40px; height:22px; border-radius:11px; }}\n"
            f"QCheckBox::indicator:unchecked {{ background-color:{colors_perm['border_light']}; border:1px solid {colors_perm['border']}; }}\n"
            f"QCheckBox::indicator:checked {{ background-color:{colors_perm['primary']}; border:1px solid {colors_perm['primary']}; }}"
        )
        def on_perm_changed(state):
            checked = state == Qt.CheckState.Checked
            self.ex_end.setEnabled(not checked)
            self.ex_end.setReadOnly(checked)
            if not checked:
                self.ex_end.setFocus()
        self.ex_permanent.stateChanged.connect(on_perm_changed)
        form.addRow(tr('admin.exams.form.title'), self.ex_title)
        form.addRow(tr('admin.exams.form.description'), self.ex_desc)
        form.addRow(tr('admin.exams.form.pass_ratio'), self.ex_pass)
        form.addRow(tr('admin.exams.form.time_limit'), self.ex_time)
        form.addRow(tr('admin.exams.form.end_date'), self.ex_end)
        form.addRow('', self.ex_permanent)
        add_btn = QPushButton(tr('admin.exams.add_btn'))
        add_btn.setIcon(self.icon_manager.get_icon('exam_add'))
        add_btn.clicked.connect(self.add_exam)
        import_btn = QPushButton(tr('admin.import_questions'))
        import_btn.setIcon(self.icon_manager.get_icon('exam_import'))
        import_btn.clicked.connect(self.import_questions)
        export_btn = QPushButton(tr('admin.export_sample'))
        export_btn.setIcon(self.icon_manager.get_icon('exam_export'))
        export_btn.clicked.connect(self.export_sample)
        vb2.addLayout(form)
        hb = QHBoxLayout()
        hb.addWidget(add_btn)
        hb.addWidget(import_btn)
        hb.addWidget(export_btn)
        vb2.addLayout(hb)
        gb2.setLayout(vb2)
        lay.addWidget(gb1, 3)
        lay.addWidget(gb2, 1)
        return page

    def refresh_exams(self):
        tbl = getattr(self, 'exams_table', None)
        if tbl is None:
            return
        tbl.blockSignals(True)
        tbl.setRowCount(0)
        for e in list_exams(include_expired=True):
            r = tbl.rowCount()
            tbl.insertRow(r)
            it_id = QTableWidgetItem(str(e[0]))
            it_id.setData(Qt.ItemDataRole.UserRole, e[6])  # store uuid
            it_id.setFlags(it_id.flags() & ~Qt.ItemFlag.ItemIsEditable)
            tbl.setItem(r, 0, it_id)
            tbl.setItem(r, 1, QTableWidgetItem(e[1] or ''))
            tbl.setItem(r, 2, QTableWidgetItem(f"{int(float(e[3])*100)}%"))
            it_time = QTableWidgetItem(str(e[4]))
            it_time.setFlags(it_time.flags() & ~Qt.ItemFlag.ItemIsEditable)
            tbl.setItem(r, 3, it_time)
            it_end = QTableWidgetItem(e[5] if e[5] else tr('common.permanent'))
            it_end.setFlags(it_end.flags() & ~Qt.ItemFlag.ItemIsEditable)
            tbl.setItem(r, 4, it_end)
            tbl.setItem(r, 5, QTableWidgetItem(e[2] or ''))
            try:
                exam_uuid = e[6] or ''
                stats = get_exam_stats(exam_uuid)
            except Exception:
                stats = {'count': 0, 'total_score': 0}
            it_cnt = QTableWidgetItem(str(int(stats['count']) if stats and 'count' in stats else 0))
            it_cnt.setFlags(it_cnt.flags() & ~Qt.ItemFlag.ItemIsEditable)
            tbl.setItem(r, 6, it_cnt)
            it_total = QTableWidgetItem(str(int(stats['total_score']) if stats and 'total_score' in stats else 0))
            it_total.setFlags(it_total.flags() & ~Qt.ItemFlag.ItemIsEditable)
            tbl.setItem(r, 7, it_total)
            opw = QWidget()
            hb = QHBoxLayout()
            hb.setContentsMargins(0,0,0,0)
            btn_export = QPushButton(tr('admin.export.exam'))
            btn_export.setIcon(self.icon_manager.get_icon('exam_export'))
            btn_clear = QPushButton(tr('common.clear'))
            btn_clear.setIcon(self.icon_manager.get_icon('delete'))
            btn_del = QPushButton(tr('common.delete'))
            btn_del.setIcon(self.icon_manager.get_icon('exam_delete'))
            exam_uuid = e[6] or ''
            exam_id = e[0]
            btn_export.clicked.connect(lambda _, x=(exam_id, exam_uuid, e[1]): self.export_exam_questions(*x))
            btn_clear.clicked.connect(lambda _, x=exam_uuid: self.clear_exam(x))
            btn_del.clicked.connect(lambda _, x=exam_id: self.delete_exam(x))
            hb.addWidget(btn_export)
            hb.addWidget(btn_clear)
            hb.addWidget(btn_del)
            hb.addStretch()
            opw.setLayout(hb)
            tbl.setCellWidget(r, 8, opw)
        try:
            for r in range(tbl.rowCount()):
                for c in range(tbl.columnCount()):
                    it = tbl.item(r, c)
                    if it:
                        it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        except Exception:
            pass
        tbl.blockSignals(False)
    def on_exam_item_changed(self, item):
        row = item.row()
        col = item.column()
        try:
            exam_id = int(self.exams_table.item(row, 0).text())
        except Exception:
            return
        if col == 1:
            title = item.text().strip()
            if not title:
                show_warn(self, tr('common.error'), tr('error.title_required'))
                self.refresh_exams()
                return
            update_exam_title_desc(exam_id, title=title)
            self.refresh_exams()
        elif col == 5:
            desc = item.text().strip()
            update_exam_title_desc(exam_id, description=desc)
            self.refresh_exams()
    def add_exam(self):
        title = self.ex_title.text().strip()
        desc = self.ex_desc.toPlainText().strip()
        pass_ratio = self.ex_pass.value() / 100.0
        tl = self.ex_time.value()
        end = None if self.ex_permanent.isChecked() else self.ex_end.dateTime().toString(Qt.DateFormat.ISODate)
        if not title:
            show_warn(self, tr('common.error'), tr('error.title_required'))
            return
        add_exam(title, desc, pass_ratio, tl, end)
        self.refresh_exams()
        show_info(self, tr('common.success'), tr('info.exam_added'))
    def get_selected_exam_id(self):
        tbl = getattr(self, 'exams_table', None)
        if tbl is None:
            return None
        r = tbl.currentRow()
        if r < 0:
            return None
        it = tbl.item(r, 0)
        return int(it.text()) if it and it.text() else None
    def get_selected_exam_uuid(self):
        tbl = getattr(self, 'exams_table', None)
        if tbl is None:
            return None
        r = tbl.currentRow()
        if r < 0:
            return None
        it = tbl.item(r, 0)
        return it.data(Qt.ItemDataRole.UserRole) if it else None
    def on_exam_double_clicked(self, row, col):
        tbl = getattr(self, 'exams_table', None)
        if tbl is None:
            return
        it = tbl.item(row, 0)
        if it is None:
            return
        try:
            exam_id = int(it.text())
        except Exception:
            return
        exam_uuid = it.data(Qt.ItemDataRole.UserRole)
        if not exam_uuid:
            return
        title_it = tbl.item(row, 1)
        title = title_it.text() if title_it else ''
        self.open_question_page(exam_id, exam_uuid, title)
    def open_question_page(self, exam_id, exam_uuid, title):
        self.question_page.open_exam(exam_id, exam_uuid, title)
        self.stack.setCurrentWidget(self.question_page)
    def back_to_exams(self):
        self.stack.setCurrentWidget(self.exam_list_page)
        self.refresh_exams()
    def import_questions(self):
        exam_id = self.get_selected_exam_id()
        exam_uuid = self.get_selected_exam_uuid()
        if not exam_id or not exam_uuid:
            show_warn(self, tr('common.error'), tr('error.select_exam'))
            return
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents')
        fn, sel = QFileDialog.getOpenFileName(self, tr('admin.import.title'), suggested, 'Excel (*.xlsx)')
        if not fn:
            return
        try:
            wb = load_workbook(fn)
            rand_count = None

            def parse_sheet(ws):
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                header = [str(x).strip() if x else '' for x in header_row]

                def idx(name):
                    try:
                        return header.index(name)
                    except Exception:
                        return -1

                itype = idx('类型')
                icontent = idx('内容')
                icorrect = idx('正确答案')
                iscore = idx('分数')
                start_opts = None
                for i, h in enumerate(header):
                    if h.startswith('选项'):
                        start_opts = i
                        break
                base_cols = [x for x in (itype, icontent, icorrect, iscore) if x >= 0]
                if min(itype, icontent, icorrect) < 0:
                    return []
                if start_opts is None:
                    start_opts = (max(base_cols) + 1) if base_cols else 3

                img_dic = {}
                for image in ws._images:
                    col = image.anchor._from.col
                    row = image.anchor._from.row
                    img_io = BytesIO(image._data())
                    img_dic[row] = {col: img_io}

                data_local = []
                for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                    tval = (str(row[itype]).strip().lower() if row[itype] is not None else '')
                    qtype = None
                    if tval in ('单选', 'single'):
                        qtype = 'single'
                    elif tval in ('多选', 'multiple'):
                        qtype = 'multiple'
                    elif tval in ('判断', 'truefalse', '判断题'):
                        qtype = 'truefalse'
                    elif tval in ('填空', 'fill'):
                        qtype = 'fill'
                    elif tval in ('简答', 'essay'):
                        qtype = 'essay'
                    else:
                        continue
                    text = (str(row[icontent]).strip() if row[icontent] is not None else '')
                    if not text:
                        continue
                    correct_cell = (str(row[icorrect]).strip() if row[icorrect] is not None else '')
                    correct = []
                    if qtype == 'truefalse':
                        lc = correct_cell.lower()
                        if lc in ('true', 'false'):
                            correct = [True] if lc == 'true' else [False]
                        else:
                            continue
                    else:
                        parts = [p.strip().upper() for p in correct_cell.replace('，', ',').replace(';', ',').split(',') if p.strip()]
                        correct = parts[:1] if qtype == 'single' else parts
                    options = []
                    if qtype != 'truefalse':
                        letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
                        cidx = start_opts
                        key_index = 0
                        while cidx < len(row):
                            val = row[cidx]
                            if val is None or str(val).strip() == '':
                                break
                            key = letters[key_index] if key_index < len(letters) else str(key_index + 1)
                            options.append({'key': key, 'text': str(val).strip()})
                            cidx += 1
                            key_index += 1
                        if not options and qtype not in ('truefalse', 'fill', 'essay'):
                            continue
                    sc = 1.0
                    if iscore >= 0 and iscore < len(row):
                        try:
                            v = row[iscore]
                            if v is not None and str(v).strip() != '':
                                sc = float(str(v).strip())
                        except Exception:
                            sc = 1.0
                    pic_group = img_dic.get(idx + 1, None)
                    pic_hash_list = []
                    if isinstance(pic_group, dict):
                        pic_list = [v for k, v in sorted(pic_group.items())]
                        for pic in pic_list:
                            hash_str = save_pic(pic)
                            if hash_str:
                                pic_hash_list.append(hash_str)
                    pic_hash_list_str = json.dumps(pic_hash_list, ensure_ascii=False)
                    item = {'type': qtype, 'text': text, 'score': sc, 'options': options, 'correct': correct, 'pictures': pic_hash_list_str}
                    data_local.append(item)
                return data_local

            data_mand = []
            data_rand = []
            if '配置选项' in wb.sheetnames:
                ws_cfg = wb['配置选项']
                cfg_header = next(ws_cfg.iter_rows(min_row=1, max_row=1, values_only=True))
                cfg = {str(cfg_header[i]).strip(): (ws_cfg.cell(row=2, column=i + 1).value) for i in range(len(cfg_header)) if cfg_header[i] is not None}
                if '随机抽取数量' in cfg:
                    try:
                        rand_count = int(str(cfg['随机抽取数量']).strip())
                    except Exception:
                        rand_count = None
            if '必考题库' in wb.sheetnames:
                data_mand = parse_sheet(wb['必考题库'])
            if '随机题库' in wb.sheetnames:
                data_rand = parse_sheet(wb['随机题库'])
            if not data_mand and not data_rand:
                show_warn(self, tr('common.error'), tr('error.lost_mandatory_or_random'))
                return
            data = {'mandatory': data_mand, 'random': data_rand, 'config': {}}
            if rand_count is not None:
                data['config']['random_pick_count'] = rand_count
            valid = []
            errs = []
            def validate_list(lst, pool_name):
                base_index = len(valid)
                for idx, q in enumerate(lst, start=1):
                    t = (q.get('type') or '').strip().lower()
                    if t not in ('single','multiple','truefalse','fill','essay'):
                        errs.append(f'{pool_name} {tr("common.question")} {idx} {tr("error.invalid_type")}')
                        continue
                    corr = q.get('correct') or []
                    if t in ('single','multiple'):
                        opts = q.get('options') or []
                        keys = {str(o.get('key')).strip().upper() for o in opts if o.get('key')}
                        if not keys:
                            errs.append(f'{pool_name} {tr("common.question")} {idx} {tr("error.missing_options")}')
                            continue
                        corr = [str(x).strip().upper() for x in corr if str(x).strip() != '']
                        if not corr or not set(corr).issubset(keys):
                            errs.append(f'{pool_name} {tr("common.question")} {idx} {tr("error.invalid_correct")}')
                            continue
                        if t == 'single' and len(corr) != 1:
                            errs.append(f'{pool_name} {tr("common.question")} {idx} {tr("error.single_need_one")}')
                            continue
                        q['correct'] = corr
                    elif t == 'fill':
                        # 填空题：验证正确答案不为空
                        if not corr or not any(str(c).strip() for c in corr):
                            errs.append(f'{pool_name} {tr("common.question")} {idx} {tr("error.invalid_correct")}')
                            continue
                        q['correct'] = [str(c).strip() for c in corr if str(c).strip()]
                    elif t == 'essay':
                        # 简答题：无预设正确答案，correct留空
                        q['correct'] = []
                    else:
                        if not corr or len(corr) != 1 or not isinstance(corr[0], bool):
                            errs.append(f'{pool_name} {tr("common.question")} {idx} {tr("error.tf_need_one")}')
                            continue
                    valid.append(q)
            if isinstance(data, dict):
                cfg = data.get('config') or {}
                from models import update_exam_random_pick_count
                if 'random_pick_count' in cfg:
                    try:
                        update_exam_random_pick_count(exam_id, int(cfg.get('random_pick_count') or 0))
                    except Exception:
                        pass
                mand = data.get('mandatory') or []
                rand = data.get('random') or []
                if not mand and not rand:
                    show_warn(self, tr('common.error'), tr('admin.import.error.jsonyaml_missing'))
                    return
                for x in mand:
                    x['pool'] = 'mandatory'
                for x in rand:
                    x['pool'] = 'random'
                validate_list(mand, '必考题库')
                validate_list(rand, '随机题库')
            else:
                show_warn(self, tr('common.error'), tr('admin.import.error.jsonyaml_dict'))
                return
            if not valid:
                detail = '\n'.join(errs[:20]) if errs else tr('admin.import.error.no_valid')
                show_warn(self, tr('common.error'), detail)
                return
            import_questions_from_json(exam_uuid, valid)
            self.refresh_exams()
            cnt_single = sum(1 for d in valid if d.get('type') == 'single')
            cnt_multiple = sum(1 for d in valid if d.get('type') == 'multiple')
            cnt_tf = sum(1 for d in valid if d.get('type') == 'truefalse')
            cnt_fill = sum(1 for d in valid if d.get('type') == 'fill')
            cnt_essay = sum(1 for d in valid if d.get('type') == 'essay')
            cnt_mand = sum(1 for d in valid if (d.get('pool') or 'mandatory') == 'mandatory')
            cnt_rand = sum(1 for d in valid if (d.get('pool') or 'mandatory') == 'random')
            extra = ''
            if errs:
                extra = f'\n{tr("admin.import.extra_prefix")}:\n' + '\n'.join(errs[:10])
            show_info(self, tr('common.success'), tr('admin.import.success', single=cnt_single, multiple=cnt_multiple, truefalse=cnt_tf, fill=cnt_fill, essay=cnt_essay, mandatory=cnt_mand, random=cnt_rand, extra=extra))
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))
    def export_exam_questions(self, exam_id, exam_uuid, title=''):
        """将指定试卷的全部题目导出为与导入模板一致的 Excel 文件"""
        if not exam_id or not exam_uuid:
            show_warn(self, tr('common.error'), tr('error.select_exam'))
            return
        safe_title = (title or 'exam').strip().replace('/', '_').replace('\\', '_')
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents', f'{safe_title}_questions')
        fn, sel = QFileDialog.getSaveFileName(self, tr('admin.export.exam.title'), suggested, 'Excel (*.xlsx)')
        if not fn:
            return
        try:
            ext = os.path.splitext(fn)[1].lower()
            out = fn if ext == '.xlsx' or ext == '' else fn + '.xlsx'
            mand = list_questions_by_pool(exam_uuid, 'mandatory')
            rand = list_questions_by_pool(exam_uuid, 'random')
            wb = Workbook()
            ws_cfg = wb.active
            ws_cfg.title = '配置选项'
            ws_cfg.append(['随机抽取数量'])
            ws_cfg.append([int(get_exam_random_pick_count(exam_uuid) or 0)])
            self._write_question_sheet(wb.create_sheet('必考题库'), mand)
            self._write_question_sheet(wb.create_sheet('随机题库'), rand)
            wb.save(out)
            show_info(self, tr('common.success'), tr('admin.export.exam.done'))
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))

    def _write_question_sheet(self, ws, rows):
        """按导入模板格式写入一个题库工作表（选项列数按实际题目动态扩展）"""
        letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        max_opts = max((len(q.get('options') or []) for q in rows), default=0)
        headers = ['类型', '内容', '正确答案', '分数'] + [f'选项{letters[i]}' for i in range(max_opts)]
        ws.append(headers)
        for item in rows:
            qtype = item.get('type')
            text = item.get('text') or ''
            score = item.get('score')
            correct = item.get('correct') or []
            if qtype == 'truefalse':
                row = ['判断', text, 'true' if correct and correct[0] else 'false', score]
            elif qtype == 'single':
                row = ['单选', text, ','.join(str(c) for c in correct), score]
            elif qtype == 'fill':
                row = ['填空', text, ' / '.join(str(c) for c in correct), score]
            elif qtype == 'essay':
                row = ['简答', text, '', score]
            else:
                row = ['多选', text, ','.join(str(c) for c in correct), score]
            row += [o.get('text') if isinstance(o, dict) else o for o in (item.get('options') or [])]
            ws.append(row)
        self._style_question_sheet(ws, headers)

    def _style_question_sheet(self, ws, headers):
        """为题目工作表套用与导入模板一致的样式"""
        header_fill = PatternFill(start_color='FF409EFF', end_color='FF409EFF', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFFFF', size=13)
        data_font = Font(size=12)
        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center')
        thin = Side(style='thin', color='FFDDDDDD')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ncols = len(headers)
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        ws.row_dimensions[1].height = 26
        for r in range(2, ws.max_row + 1):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                cell.font = data_font
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).alignment = center if c == 4 else left
            ws.row_dimensions[r].height = 22
        widths = [0] * ncols
        for r in ws.iter_rows(values_only=True):
            for idx, val in enumerate(r):
                l = len(str(val)) if val is not None else 0
                widths[idx] = max(widths[idx], l)
        for i, w in enumerate(widths, start=1):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = max(16, min(48, w + 6))
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"

    def clear_exam(self, exam_id):
        reply = ask_yes_no(self, tr('common.hint'), tr('admin.exams.clear_confirm'), default_yes=False)
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            from models import clear_exam_questions
            clear_exam_questions(exam_id)
            self.refresh_exams()
            show_info(self, tr('common.success'), tr('admin.exams.clear_done'))
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))
    def delete_exam(self, exam_id):
        reply = ask_yes_no(self, tr('common.hint'), tr('admin.exams.delete_confirm'), default_yes=False)
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            from models import delete_exam
            delete_exam(exam_id)
            self.refresh_exams()
            show_info(self, tr('common.success'), tr('admin.exams.delete_done'))
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))
    def export_sample(self):
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents/exam')
        fn, sel = QFileDialog.getSaveFileName(self, tr('admin.export.sample.title'), suggested, 'Excel (*.xlsx)')
        if not fn:
            return
        try:
            ext = os.path.splitext(fn)[1].lower()
            mand = [
                {"type":"single","text":"Python中获取列表长度的函数是?","options":[{"key":"A","text":"len(list)"},{"key":"B","text":"size(list)"},{"key":"C","text":"count(list)"},{"key":"D","text":"length(list)"}],"correct":["A"],"score":2},
                {"type":"multiple","text":"以下哪些是Linux常见包管理器?","options":[{"key":"A","text":"apt"},{"key":"B","text":"ls"},{"key":"C","text":"yum"},{"key":"D","text":"pacman"}],"correct":["A","C","D"],"score":3},
                {"type":"truefalse","text":"Python中的list是可变对象","correct":[True],"score":1},
                {"type":"fill","text":"Python中按字节读取文件的函数是?","correct":["read"],"score":2},
                {"type":"essay","text":"请简述Python中__init__方法的作用","correct":[],"score":5},
                {"type":"single","text":"查看当前工作目录的Linux命令是?","options":[{"key":"A","text":"pwd"},{"key":"B","text":"cd"},{"key":"C","text":"ls"},{"key":"D","text":"echo"}],"correct":["A"],"score":2},
                {"type":"multiple","text":"以下哪些工具可用于创建Python虚拟环境?","options":[{"key":"A","text":"venv"},{"key":"B","text":"virtualenv"},{"key":"C","text":"pip"},{"key":"D","text":"conda"}],"correct":["A","B","D"],"score":3}
            ]
            rand = [
                {"type":"truefalse","text":"Linux中/etc目录通常存放系统配置文件","correct":[True],"score":1},
                {"type":"multiple","text":"以下哪些是Python中的可迭代对象?","options":[{"key":"A","text":"list"},{"key":"B","text":"dict"},{"key":"C","text":"int"},{"key":"D","text":"tuple"}],"correct":["A","B","D"],"score":3},
                {"type":"single","text":"Python字典取值且键不存在时不抛异常的方法是?","options":[{"key":"A","text":"d['k']"},{"key":"B","text":"d.get('k')"},{"key":"C","text":"d.k"},{"key":"D","text":"getattr(d,'k')"}],"correct":["B"],"score":2},
                {"type":"single","text":"Linux查看网络端口占用的命令是?","options":[{"key":"A","text":"ss -tuln"},{"key":"B","text":"ps aux"},{"key":"C","text":"top"},{"key":"D","text":"df -h"}],"correct":["A"],"score":2},
                {"type":"multiple","text":"以下哪些属于Python打包/分发相关工具?","options":[{"key":"A","text":"setuptools"},{"key":"B","text":"wheel"},{"key":"C","text":"pip"},{"key":"D","text":"twine"}],"correct":["A","B","D"],"score":3}
            ]
            out = fn if ext == '.xlsx' or ext == '' else fn + '.xlsx'
            wb = Workbook()
            ws_cfg = wb.active
            ws_cfg.title = '配置选项'
            ws_cfg.append(['随机抽取数量'])
            ws_cfg.append([4])

            def write_sheet(ws, rows):
                ws.append(['类型', '内容', '正确答案', '分数', '选项A', '选项B', '选项C', '选项D'])
                for item in rows:
                    if item['type'] == 'truefalse':
                        ws.append(['判断', item['text'], 'true' if item['correct'][0] else 'false', item['score']])
                    elif item['type'] == 'single':
                        ws.append(['单选', item['text'], ','.join(item['correct']), item['score']] + [opt['text'] for opt in item.get('options', [])])
                    elif item['type'] == 'fill':
                        ws.append(['填空', item['text'], ' / '.join(str(c) for c in item['correct']), item['score']])
                    elif item['type'] == 'essay':
                        ws.append(['简答', item['text'], '', item['score']])
                    else:
                        ws.append(['多选', item['text'], ','.join(item['correct']), item['score']] + [opt['text'] for opt in item.get('options', [])])

            ws_m = wb.create_sheet('必考题库')
            write_sheet(ws_m, mand)
            ws_r = wb.create_sheet('随机题库')
            write_sheet(ws_r, rand)
            header_fill = PatternFill(start_color='FF409EFF', end_color='FF409EFF', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFFFF', size=13)
            data_font = Font(size=12)
            center = Alignment(horizontal='center', vertical='center')
            left = Alignment(horizontal='left', vertical='center')
            thin = Side(style='thin', color='FFDDDDDD')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ws in [ws_m, ws_r]:
                headers = ['类型', '内容', '正确答案', '分数', '选项A', '选项B', '选项C', '选项D']
                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=c)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center
                ws.row_dimensions[1].height = 26
                for r in range(2, ws.max_row + 1):
                    for c in range(1, len(headers) + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.border = border
                        cell.font = data_font
                    ws.cell(row=r, column=4).alignment = center
                    for c in (1, 2, 3, 5, 6, 7, 8):
                        ws.cell(row=r, column=c).alignment = left
                    ws.row_dimensions[r].height = 22
                widths = [0] * len(headers)
                for r in ws.iter_rows(values_only=True):
                    for idx, val in enumerate(r):
                        l = len(str(val)) if val is not None else 0
                        widths[idx] = max(widths[idx], l)
                for i, w in enumerate(widths, start=1):
                    letter = get_column_letter(i)
                    ws.column_dimensions[letter].width = max(16, min(48, w + 6))
                ws.freeze_panes = 'A2'
                ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
            wb.save(out)
            show_info(self, tr('common.success'), tr('admin.export.sample.done'))
        except Exception as e:
            show_warn(self, tr('common.error'), str(e))
