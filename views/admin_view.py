import json
import os
import pathlib
import yaml
import re
import re
from openpyxl import load_workbook, Workbook
from PySide6.QtCore import Qt, QThread, Signal, QSize, QRegularExpression, QDateTime
from PySide6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QTextEdit, QFormLayout, QSpinBox, QDateTimeEdit, QFileDialog, QTabWidget, QTableWidget, QTableWidgetItem, QGroupBox, QCheckBox, QComboBox, QMessageBox, QProgressDialog, QListView, QAbstractItemView, QTextBrowser
from PySide6.QtGui import QRegularExpressionValidator
from database import DB_PATH
from models import list_users, create_user, list_exams, add_exam, import_questions_from_json, list_sync_targets, upsert_sync_target, delete_user, update_user_role, update_user_active, delete_sync_target, update_sync_target, get_exam_title, get_exam_stats, update_exam_random_pick_count
from theme_manager import theme_manager
from language import tr
from icon_manager import get_icon
from status_indicators import LoadingIndicator
from sync import rsync_push, rsync_pull

class SyncWorker(QThread):
    progress = Signal(str)
    finished = Signal(str)
    error = Signal(str)
    def __init__(self, targets, operation='push'):
        super().__init__()
        self.targets = targets
        self.operation = operation
    def run(self):
        results = []
        for t in self.targets:
            try:
                ssh_password = t[5] if len(t) > 5 else None
                if self.operation == 'push':
                    code, out, err = rsync_push(t[2], t[3], t[4], ssh_password)
                    if code == 0:
                        result = f'{t[1]} ({t[2]}) 推送成功'
                    else:
                        result = f'{t[1]} ({t[2]}) 推送失败: {err or "未知错误"}'
                else:
                    base_dir = os.path.join(os.path.dirname(DB_PATH), 'pulled')
                    os.makedirs(base_dir, exist_ok=True)
                    ip = t[2]
                    dest_dir = os.path.join(base_dir, ip)
                    os.makedirs(dest_dir, exist_ok=True)
                    code, out, err = rsync_pull(ip, t[3], t[4], dest_dir, ssh_password)
                    if code == 0:
                        result = f'{t[1]} ({ip}) 拉取成功'
                        rp = os.path.join(dest_dir, os.path.basename(t[4]))
                        try:
                            from models import merge_remote_db
                            merge_remote_db(rp)
                            result += ' (成绩已合并)'
                        except Exception as me:
                            result += f' (合并失败: {str(me)})'
                    else:
                        result = f'{t[1]} ({ip}) 拉取失败: {err or "未知错误"}'
                results.append(result)
                self.progress.emit(result)
            except Exception as e:
                error_msg = f'{t[1]} 错误: {str(e)}'
                results.append(error_msg)
                self.error.emit(error_msg)
        self.finished.emit('\n'.join(results))

class AdminView(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        colors = theme_manager.get_theme_colors()
        bkg = 'background' + '-color'
        col = 'co' + 'lor'
        bd = 'bor' + 'der'
        pd = 'pad' + 'ding'
        ff = 'font' + '-family'
        br = 'border' + '-radius'
        ss_admin = (
            f"QWidget {{ {bkg}:{colors['background']}; {ff}:\"PingFang SC\",sans-serif; }}\n"
            f"QTabWidget::pane {{ {bd}:1px solid {colors['border']}; {bkg}:{colors['card_background']}; {br}:8px; }}\n"
            f"QTabBar::tab {{ {bkg}:{colors['border_light']}; {pd}:10px 16px; margin-right:2px; {br}:8px; }}\n"
            f"QTabBar::tab:selected {{ {bkg}:{colors['primary']}; {col}:{colors['text_inverse']}; }}\n"
            f"QGroupBox {{ font-weight:bold; {bd}:1px solid {colors['border']}; {br}:8px; margin-top:8px; padding-top:8px; }}\n"
            f"QPushButton {{ {bkg}:{colors['button_primary']}; {col}:{colors['text_inverse']}; {pd}:6px 12px; border:none; {br}:8px; }}\n"
            f"QPushButton:hover {{ {bkg}:{colors['button_primary_hover']}; }}\n"
            f"QTableWidget {{ {bd}:1px solid {colors['border']}; {br}:8px; {bkg}:{colors['card_background']}; {col}:{colors['text_primary']}; }}\n"
            f"QTableWidget::item:hover {{ {bkg}:{colors['border_light']}; }}\n"
            f"QTableWidget::item:selected {{ {bkg}:{colors['primary']}; {col}:{colors['text_inverse']}; }}\n"
            f"QHeaderView::section {{ {bkg}:{colors['border_light']}; {col}:{colors['text_secondary']}; font-weight:600; {pd}:6px 8px; {bd}:none; {bd}-right:1px solid {colors['border']}; }}\n"
            f"QLineEdit, QTextEdit, QSpinBox, QDateTimeEdit {{ {pd}:6px; {bd}:1px solid {colors['input_border']}; {br}:8px; {bkg}:{colors['input_background']}; {col}:{colors['text_primary']}; }}\n"
            f"QLineEdit:focus, QTextEdit:focus {{ {bd}-color:{colors['primary']}; }}\n"
            f"QComboBox {{ {pd}:6px 8px; {bd}:1px solid {colors['input_border']}; {br}:8px; {bkg}:{colors['input_background']}; {col}:{colors['text_primary']}; }}\n"
            f"QComboBox:focus {{ {bd}-color:{colors['primary']}; }}\n"
            f"QComboBox::drop-down {{ width:24px; border:none; }}\n"
            f"QComboBox QAbstractItemView {{ {bkg}:{colors['card_background']}; {bd}:1px solid {colors['border']}; {pd}:4px; outline:0; }}\n"
        )
        self.setStyleSheet(ss_admin)
        tabs = QTabWidget()
        tabs.addTab(self.users_tab(), tr('admin.users_tab'))
        tabs.addTab(self.exams_tab(), tr('admin.exams_tab'))
        tabs.addTab(self.sync_tab(), tr('admin.sync_tab'))
        tabs.addTab(self.scores_tab(), tr('admin.scores_tab'))
        tabs.setTabIcon(0, get_icon('user'))
        tabs.setTabIcon(1, get_icon('exam'))
        tabs.setTabIcon(2, get_icon('sync'))
        tabs.setTabIcon(3, get_icon('score'))
        layout = QVBoxLayout()
        topbar = QHBoxLayout()
        title = QLabel(tr('admin.dashboard'))
        title.setStyleSheet("font-size:18px; font-weight:bold;")
        topbar.addWidget(title)
        topbar.addStretch()
        logout_btn = QPushButton(tr('common.logout'))
        logout_btn.setIcon(get_icon('confirm'))
        logout_btn.clicked.connect(self.handle_logout)
        topbar.addWidget(logout_btn)
        layout.addLayout(topbar)
        layout.addWidget(tabs)
        self.setLayout(layout)
    def handle_logout(self):
        p = self.parent()
        while p is not None and not hasattr(p, 'logout'):
            p = p.parent()
        if p is not None:
            p.logout()
    def show_sync_progress(self, message):
        if hasattr(self, 'sync_progress_dialog') and getattr(self, 'sync_progress_dialog'):
            try:
                self.sync_progress_dialog.close()
            except Exception:
                pass
        colors = theme_manager.get_theme_colors()
        dlg = QProgressDialog(message, '', 0, 0, self)
        dlg.setWindowTitle(tr('sync.progress.title'))
        dlg.setCancelButton(None)
        dlg.setMinimumDuration(0)
        dlg.setAutoClose(False)
        dlg.setAutoReset(False)
        dlg.setModal(True)
        dlg.setRange(0, 0)
        dlg.setLabelText(message)
        dlg.setFixedSize(420, 140)
        try:
            dlg.setWindowFlags(dlg.windowFlags() & ~Qt.WindowType.WindowContextHelpButtonHint)
        except Exception:
            pass
        dlg.setStyleSheet(
            f"QProgressDialog {{ background-color:{colors['card_background']}; border:1px solid {colors['border']}; border-radius:12px; }}\n"
            f"QLabel {{ color:{colors['text_primary']}; font-size:14px; padding:12px; }}\n"
            f"QProgressBar {{ background-color:{colors['border_light']}; border:1px solid {colors['border']}; border-radius:10px; height:16px; margin:8px 12px; }}\n"
            f"QProgressBar::chunk {{ background-color:{colors['primary']}; border-radius:10px; }}"
        )
        dlg.show()
        self.sync_progress_dialog = dlg
    def make_tag(self, text, bg, fg):
        lab = QLabel(text)
        lab.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lab.setStyleSheet(f"QLabel {{ background-color:{bg}; color:{fg}; border-radius:10px; padding:2px 8px; font-size:12px; }}")
        return lab
    def users_tab(self):
        w = QWidget()
        lay = QVBoxLayout()
        gb = QGroupBox(tr('admin.users_group'))
        vb = QVBoxLayout()
        self.users_table = QTableWidget(0, 7)
        self.users_table.setHorizontalHeaderLabels([tr('admin.users.headers.id'), tr('admin.users.headers.username'), tr('admin.users.headers.full_name'), tr('admin.users.headers.role'), tr('admin.users.headers.status'), tr('admin.users.headers.created_at'), tr('admin.users.headers.actions')])
        self.users_table.setColumnWidth(0, 50)
        self.users_table.setColumnWidth(1, 100)
        self.users_table.setColumnWidth(2, 100)
        self.users_table.setColumnWidth(3, 80)
        self.users_table.setColumnWidth(4, 80)
        self.users_table.setColumnWidth(5, 300)
        self.users_table.horizontalHeader().setStretchLastSection(True)
        self.users_table.setAlternatingRowColors(True)
        self.refresh_users()
        vb.addWidget(self.users_table)
        gb.setLayout(vb)
        lay.addWidget(gb)
        self.users_table.itemChanged.connect(self.on_user_item_changed)
        gb2 = QGroupBox(tr('admin.new_user_group'))
        form = QFormLayout()
        self.new_user = QLineEdit()
        self.new_user.setPlaceholderText(tr('admin.users.username_ph'))
        self.new_user.setInputMethodHints(Qt.InputMethodHint.ImhNoPredictiveText | Qt.InputMethodHint.ImhNoAutoUppercase | Qt.InputMethodHint.ImhPreferLowercase)
        self.new_user.setValidator(QRegularExpressionValidator(QRegularExpression(r"^[A-Za-z0-9_@.\-]+$")))
        self.new_pwd = QLineEdit()
        self.new_pwd.setEchoMode(QLineEdit.EchoMode.Password)
        self.new_pwd.setPlaceholderText(tr('admin.users.password_ph'))
        self.new_pwd.setInputMethodHints(Qt.InputMethodHint.ImhHiddenText | Qt.InputMethodHint.ImhNoPredictiveText | Qt.InputMethodHint.ImhSensitiveData)
        self.new_pwd.setValidator(QRegularExpressionValidator(QRegularExpression(r"^[\x20-\x7E]+$")))
        self.new_fullname = QLineEdit()
        self.new_fullname.setPlaceholderText(tr('admin.users.full_name_ph'))
        try:
            self.new_fullname.setValidator(None)
        except Exception:
            pass
        self.new_role = QComboBox()
        colors = theme_manager.get_theme_colors()
        self.new_role.setStyleSheet(
            f"QComboBox {{ padding:6px 10px; border:1px solid {colors['input_border']}; border-radius:12px; background-color:{colors['input_background']}; color:{colors['text_primary']}; min-height:32px; }}\n"
            f"QComboBox:focus {{ border-color:{colors['primary']}; }}\n"
            f"QComboBox::drop-down {{ width:28px; border:none; }}\n"
            f"QComboBox QAbstractItemView {{ background-color:{colors['card_background']}; border:1px solid {colors['border']}; padding:6px; outline:0; }}\n"
            f"QComboBox QAbstractItemView::item {{ padding:6px 10px; height:28px; color:{colors['text_primary']}; }}\n"
            f"QComboBox QAbstractItemView::item:hover {{ background-color:{colors['border_light']}; }}\n"
            f"QComboBox QAbstractItemView::item:selected {{ background-color:{colors['primary']}; color:{colors['text_inverse']}; }}"
        )
        role_view = QListView()
        role_view.setStyleSheet(
            f"QListView {{ background-color:{colors['card_background']}; border:1px solid {colors['border']}; padding:6px; }}\n"
            f"QListView::item {{ padding:6px 10px; height:28px; color:{colors['text_primary']}; }}\n"
            f"QListView::item:hover {{ background-color:{colors['border_light']}; }}\n"
            f"QListView::item:selected {{ background-color:{colors['primary']}; color:{colors['text_inverse']}; }}"
        )
        self.new_role.setView(role_view)
        self.new_role.setIconSize(QSize(16, 16))
        self.new_role.addItem(get_icon('user'), 'user')
        self.new_role.addItem(get_icon('user_admin'), 'admin')
        add_btn = QPushButton(tr('admin.users.add_button'))
        add_btn.clicked.connect(self.add_user)
        form.addRow(tr('admin.users.headers.username'), self.new_user)
        form.addRow(tr('admin.users.password_ph'), self.new_pwd)
        form.addRow(tr('admin.users.headers.full_name'), self.new_fullname)
        form.addRow(tr('admin.users.headers.role'), self.new_role)
        form.addRow(add_btn)
        gb2.setLayout(form)
        lay.addWidget(gb2)
        hb_users_excel = QHBoxLayout()
        btn_export_users_tpl = QPushButton(tr('admin.users.export_tpl'))
        btn_export_users_tpl.setIcon(get_icon('exam_export'))
        btn_export_users_tpl.clicked.connect(self.export_users_template)
        btn_import_users_excel = QPushButton(tr('admin.users.import_excel'))
        btn_import_users_excel.setIcon(get_icon('exam_import'))
        btn_import_users_excel.clicked.connect(self.import_users_from_excel)
        hb_users_excel.addWidget(btn_export_users_tpl)
        hb_users_excel.addWidget(btn_import_users_excel)
        lay.addLayout(hb_users_excel)
        lay.addStretch()
        w.setLayout(lay)
        return w
    def refresh_users(self):
        self.users_table = getattr(self, 'users_table', QTableWidget(0, 7))
        self.users_table.blockSignals(True)
        self.users_table.setRowCount(0)
        for u in list_users():
            row = self.users_table.rowCount()
            self.users_table.insertRow(row)
            self.users_table.setItem(row, 0, QTableWidgetItem(str(u[0])))
            self.users_table.setItem(row, 1, QTableWidgetItem(u[1] or ''))
            self.users_table.setItem(row, 2, QTableWidgetItem((u[2] or '') if u[2] else ''))
            role_text = tr('admin.role.admin') if u[3] == 'admin' else tr('admin.role.user')
            role_bg = '#e1f3d8' if u[3] == 'admin' else '#d9ecff'
            role_fg = '#67c23a' if u[3] == 'admin' else '#409eff'
            self.users_table.setCellWidget(row, 3, self.make_tag(role_text, role_bg, role_fg))
            status_text = tr('admin.status.active') if u[4] == 1 else tr('admin.status.inactive')
            status_bg = '#e1f3d8' if u[4] == 1 else '#fde2e2'
            status_fg = '#67c23a' if u[4] == 1 else '#f56c6c'
            self.users_table.setCellWidget(row, 4, self.make_tag(status_text, status_bg, status_fg))
            self.users_table.setItem(row, 5, QTableWidgetItem(u[5]))
            it_id = self.users_table.item(row, 0)
            it_ct = self.users_table.item(row, 5)
            if it_id:
                it_id.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if it_ct:
                it_ct.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            action_widget = QWidget()
            action_layout = QHBoxLayout()
            action_layout.setContentsMargins(4, 4, 4, 4)
            delete_btn = QPushButton(tr('admin.user.delete'))
            delete_btn.setIcon(get_icon('delete'))
            delete_btn.setStyleSheet("QPushButton { background-color:#f56c6c; color:#fff; padding:4px 8px; font-size:12px; border-radius:6px; }")
            delete_btn.clicked.connect(lambda checked, uid=u[0]: self.delete_user(uid))
            action_layout.addWidget(delete_btn)
            role_btn = QPushButton(tr('admin.user.set_admin') if u[3] == 'user' else tr('admin.user.set_user'))
            role_btn.setIcon(get_icon('user_edit'))
            role_btn.setStyleSheet("QPushButton { background-color:#67c23a; color:#fff; padding:4px 8px; font-size:12px; border-radius:6px; }")
            role_btn.clicked.connect(lambda checked, uid=u[0], current_role=u[3]: self.toggle_user_role(uid, current_role))
            action_layout.addWidget(role_btn)
            active_btn = QPushButton(tr('admin.user.disable') if u[4] == 1 else tr('admin.user.enable'))
            active_btn.setIcon(get_icon('user_active' if u[4] == 1 else 'user_inactive'))
            active_btn.setStyleSheet("QPushButton { background-color:#e6a23c; color:#fff; padding:4px 8px; font-size:12px; border-radius:6px; }")
            active_btn.clicked.connect(lambda checked, uid=u[0], current_active=u[4]: self.toggle_user_active(uid, current_active))
            action_layout.addWidget(active_btn)
            action_widget.setLayout(action_layout)
            self.users_table.setCellWidget(row, 6, action_widget)
        try:
            for r in range(self.users_table.rowCount()):
                for c in range(self.users_table.columnCount()):
                    it = self.users_table.item(r, c)
                    if it:
                        it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        except Exception:
            pass
        self.users_table.blockSignals(False)
    def add_user(self):
        name = self.new_user.text().strip()
        pwd = self.new_pwd.text()
        role = self.new_role.currentText()
        if not name or not pwd:
            QMessageBox.warning(self, tr('common.error'), tr('error.input_username_password'))
            return
        try:
            full_name = self.new_fullname.text().strip() or None
            if not re.fullmatch(r"[A-Za-z0-9_@.\-]+", name):
                QMessageBox.warning(self, tr('common.error'), tr('error.username_format'))
                return
            if not re.fullmatch(r"[\x20-\x7E]+", pwd):
                QMessageBox.warning(self, tr('common.error'), tr('error.password_format'))
                return
            create_user(name, pwd, role, 1, full_name)
            self.refresh_users()
            self.new_user.clear()
            self.new_pwd.clear()
            self.new_fullname.clear()
            QMessageBox.information(self, tr('common.success'), tr('info.user_created'))
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))
    def delete_user(self, user_id):
        reply = QMessageBox.question(self, tr('common.hint'), tr('confirm.delete_user'), QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            try:
                delete_user(user_id)
                self.refresh_users()
                QMessageBox.information(self, tr('common.success'), tr('info.user_deleted'))
            except Exception as e:
                QMessageBox.warning(self, '错误', str(e))
    def toggle_user_role(self, user_id, current_role):
        new_role = 'admin' if current_role == 'user' else 'user'
        try:
            update_user_role(user_id, new_role)
            self.refresh_users()
            QMessageBox.information(self, tr('common.success'), tr('info.user_role_updated', role=new_role))
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))
    def toggle_user_active(self, user_id, current_active):
        new_active = 0 if current_active == 1 else 1
        try:
            update_user_active(user_id, new_active)
            self.refresh_users()
            status = tr('admin.user.enable') if new_active == 1 else tr('admin.user.disable')
            QMessageBox.information(self, tr('common.success'), tr('info.user_status_updated', status=status))
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))
    def exams_tab(self):
        w = QWidget()
        lay = QVBoxLayout()
        gb1 = QGroupBox(tr('admin.exams_group'))
        vb1 = QVBoxLayout()
        self.exams_table = QTableWidget(0, 9)
        self.exams_table.setHorizontalHeaderLabels([tr('admin.exams.headers.id'), tr('admin.exams.headers.title'), tr('admin.exams.headers.pass_ratio'), tr('admin.exams.headers.time_limit'), tr('admin.exams.headers.deadline'), tr('admin.exams.headers.description'), tr('admin.exams.headers.q_count'), tr('admin.exams.headers.total'), tr('admin.exams.headers.actions')])
        self.exams_table.setColumnWidth(0, 50)
        self.exams_table.setColumnWidth(1, 120)
        self.exams_table.setColumnWidth(2, 120)
        self.exams_table.setColumnWidth(3, 120)
        self.exams_table.setColumnWidth(4, 120)
        self.exams_table.setColumnWidth(5, 480)
        self.exams_table.setColumnWidth(6, 80)
        self.exams_table.setColumnWidth(7, 80)
        self.exams_table.setColumnWidth(8, 150)
        self.exams_table.horizontalHeader().setStretchLastSection(True)
        self.exams_table.setAlternatingRowColors(True)
        self.exams_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.exams_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.refresh_exams()
        vb1.addWidget(self.exams_table)
        self.exams_table.itemChanged.connect(self.on_exam_item_changed)
        gb1.setLayout(vb1)
        gb2 = QGroupBox(tr('admin.new_exam_group'))
        vb2 = QVBoxLayout()
        form = QFormLayout()
        self.ex_title = QLineEdit()
        self.ex_title.setPlaceholderText(tr('admin.exams.form.title'))
        self.ex_desc = QTextEdit()
        self.ex_desc.setPlaceholderText(tr('admin.exams.form.description'))
        self.ex_desc.setMaximumHeight(80)
        self.ex_pass = QSpinBox()
        self.ex_pass.setRange(0, 100)
        self.ex_pass.setValue(60)
        colors_inputs = theme_manager.get_theme_colors()
        spin_style = (
            f"QSpinBox {{ padding:6px 10px; border:1px solid {colors_inputs['input_border']}; border-radius:12px; background-color:{colors_inputs['input_background']}; color:{colors_inputs['text_primary']}; }}\n"
            f"QSpinBox:focus {{ border-color:{colors_inputs['primary']}; }}\n"
            f"QSpinBox::up-button, QSpinBox::down-button {{ width:20px; border:none; background-color:{colors_inputs['border_light']}; border-radius:8px; }}\n"
            f"QSpinBox::up-button:hover, QSpinBox::down-button:hover {{ background-color:{colors_inputs['button_primary_hover']}; }}"
        )
        self.ex_pass.setStyleSheet(spin_style)
        self.ex_time = QSpinBox()
        self.ex_time.setRange(1, 600)
        self.ex_time.setValue(60)
        self.ex_time.setStyleSheet(spin_style)
        self.ex_random_count = QSpinBox()
        self.ex_random_count.setRange(0, 1000)
        self.ex_random_count.setValue(4)
        self.ex_random_count.setStyleSheet(spin_style)
        self.ex_end = QDateTimeEdit()
        self.ex_end.setDateTime(QDateTime.currentDateTime())
        self.ex_end.setDisplayFormat('yyyy-MM-dd HH:mm')
        self.ex_end.setCalendarPopup(True)
        dt_style = (
            f"QDateTimeEdit {{ padding:6px 10px; border:1px solid {colors_inputs['input_border']}; border-radius:12px; background-color:{colors_inputs['input_background']}; color:{colors_inputs['text_primary']}; }}\n"
            f"QDateTimeEdit:focus {{ border-color:{colors_inputs['primary']}; }}\n"
            f"QDateTimeEdit::up-button, QDateTimeEdit::down-button {{ width:22px; border:none; background-color:{colors_inputs['border_light']}; border-radius:10px; margin:2px; }}\n"
            f"QDateTimeEdit::up-button:hover, QDateTimeEdit::down-button:hover {{ background-color:{colors_inputs['button_primary_hover']}; }}\n"
            f"QDateTimeEdit::up-button:pressed, QDateTimeEdit::down-button:pressed {{ background-color:{colors_inputs['primary']}; }}\n"
            f"QDateTimeEdit::up-arrow, QDateTimeEdit::down-arrow {{ width: 0; height: 0; }}"
        )
        self.ex_end.setStyleSheet(dt_style)
        cal = self.ex_end.calendarWidget()
        cal.setStyleSheet(
            f"QCalendarWidget {{ background-color:{colors_inputs['card_background']}; border:1px solid {colors_inputs['border']}; border-radius:8px; }}\n"
            f"QCalendarWidget QWidget#qt_calendar_navigationbar {{ background-color:{colors_inputs['card_background']}; border:none; padding:6px; }}\n"
            f"QCalendarWidget QToolButton#qt_calendar_prevmonth, QCalendarWidget QToolButton#qt_calendar_nextmonth {{ background-color:{colors_inputs['button_primary']}; color:{colors_inputs['text_inverse']}; border:none; border-radius:6px; padding:4px 8px; }}\n"
            f"QCalendarWidget QToolButton#qt_calendar_prevmonth:hover, QCalendarWidget QToolButton#qt_calendar_nextmonth:hover {{ background-color:{colors_inputs['button_primary_hover']}; }}\n"
            f"QCalendarWidget QToolButton#qt_calendar_monthbutton {{ background-color:{colors_inputs['border_light']}; color:{colors_inputs['text_primary']}; border:none; border-radius:6px; padding:4px 10px; }}\n"
            f"QCalendarWidget QToolButton#qt_calendar_monthbutton:hover {{ background-color:{colors_inputs['button_primary_hover']}; color:{colors_inputs['text_inverse']}; }}\n"
            f"QCalendarWidget QSpinBox#qt_calendar_yearspinbox {{ background-color:{colors_inputs['input_background']}; color:{colors_inputs['text_primary']}; border:1px solid {colors_inputs['input_border']}; border-radius:6px; padding:2px 6px; }}\n"
            f"QCalendarWidget QTableView {{ background-color:{colors_inputs['card_background']}; alternate-background-color:{colors_inputs['border_light']}; selection-background-color:{colors_inputs['primary']}; selection-color:{colors_inputs['text_inverse']}; gridline-color:{colors_inputs['border']}; }}\n"
            f"QCalendarWidget QTableView::item {{ padding:4px; }}\n"
            f"QCalendarWidget QTableView::item:hover {{ background-color:{colors_inputs['border_light']}; }}\n"
            f"QCalendarWidget QTableView::item:selected {{ background-color:{colors_inputs['primary']}; color:{colors_inputs['text_inverse']}; }}"
        )
        self.ex_permanent = QCheckBox(tr('admin.exams.permanent_checkbox'))
        colors_perm = theme_manager.get_theme_colors()
        self.ex_permanent.setStyleSheet(
            f"QCheckBox {{ color:{colors_perm['text_primary']}; font-size:14px; }}\n"
            f"QCheckBox::indicator {{ width:40px; height:22px; border-radius:11px; }}\n"
            f"QCheckBox::indicator:unchecked {{ background-color:{colors_perm['border_light']}; border:1px solid {colors_perm['border']}; }}\n"
            f"QCheckBox::indicator:checked {{ background-color:{colors_perm['primary']}; border:1px solid {colors_perm['primary']}; }}"
        )
        def on_perm_changed(state):
            checked = state == Qt.CheckState.Checked
            self.ex_end.setEnabled(not checked)
            self.ex_end.setReadOnly(checked)
            if not checked:
                self.ex_end.setFocus()
        self.ex_permanent.stateChanged.connect(on_perm_changed)
        form.addRow(tr('admin.exams.form.title'), self.ex_title)
        form.addRow(tr('admin.exams.form.description'), self.ex_desc)
        form.addRow(tr('admin.exams.form.pass_ratio'), self.ex_pass)
        form.addRow(tr('admin.exams.form.time_limit'), self.ex_time)
        form.addRow(tr('admin.exams.form.end_date'), self.ex_end)
        form.addRow(tr('admin.exams.form.random_pick'), self.ex_random_count)
        form.addRow('', self.ex_permanent)
        add_btn = QPushButton(tr('admin.exams.add_btn'))
        add_btn.setIcon(get_icon('exam_add'))
        add_btn.clicked.connect(self.add_exam)
        import_btn = QPushButton(tr('admin.import_questions'))
        import_btn.setIcon(get_icon('exam_import'))
        import_btn.clicked.connect(self.import_questions)
        export_btn = QPushButton(tr('admin.export_sample'))
        export_btn.setIcon(get_icon('exam_export'))
        export_btn.clicked.connect(self.export_sample)
        vb2.addLayout(form)
        hb = QHBoxLayout()
        hb.addWidget(add_btn)
        hb.addWidget(import_btn)
        hb.addWidget(export_btn)
        vb2.addLayout(hb)
        gb2.setLayout(vb2)
        lay.addWidget(gb1, 3)
        lay.addWidget(gb2, 1)
        w.setLayout(lay)
        return w
    def refresh_exams(self):
        tbl = getattr(self, 'exams_table', None)
        if tbl is None:
            return
        tbl.blockSignals(True)
        tbl.setRowCount(0)
        for e in list_exams(include_expired=True):
            r = tbl.rowCount()
            tbl.insertRow(r)
            it_id = QTableWidgetItem(str(e[0]))
            it_id.setFlags(it_id.flags() & ~Qt.ItemFlag.ItemIsEditable)
            tbl.setItem(r, 0, it_id)
            tbl.setItem(r, 1, QTableWidgetItem(e[1] or ''))
            tbl.setItem(r, 2, QTableWidgetItem(f"{int(float(e[3])*100)}%"))
            it_time = QTableWidgetItem(str(e[4]))
            it_time.setFlags(it_time.flags() & ~Qt.ItemFlag.ItemIsEditable)
            tbl.setItem(r, 3, it_time)
            it_end = QTableWidgetItem(e[5] if e[5] else tr('common.permanent'))
            it_end.setFlags(it_end.flags() & ~Qt.ItemFlag.ItemIsEditable)
            tbl.setItem(r, 4, it_end)
            tbl.setItem(r, 5, QTableWidgetItem(e[2] or ''))
            try:
                stats = get_exam_stats(int(e[0]))
            except Exception:
                stats = {'count': 0, 'total_score': 0}
            it_cnt = QTableWidgetItem(str(int(stats['count']) if stats and 'count' in stats else 0))
            it_cnt.setFlags(it_cnt.flags() & ~Qt.ItemFlag.ItemIsEditable)
            tbl.setItem(r, 6, it_cnt)
            it_total = QTableWidgetItem(str(int(stats['total_score']) if stats and 'total_score' in stats else 0))
            it_total.setFlags(it_total.flags() & ~Qt.ItemFlag.ItemIsEditable)
            tbl.setItem(r, 7, it_total)
            opw = QWidget()
            hb = QHBoxLayout()
            hb.setContentsMargins(0,0,0,0)
            btn_clear = QPushButton(tr('common.clear'))
            btn_clear.setIcon(get_icon('delete'))
            btn_del = QPushButton(tr('common.delete'))
            btn_del.setIcon(get_icon('exam_delete'))
            exam_id = e[0]
            btn_clear.clicked.connect(lambda _, x=exam_id: self.clear_exam(x))
            btn_del.clicked.connect(lambda _, x=exam_id: self.delete_exam(x))
            hb.addWidget(btn_clear)
            hb.addWidget(btn_del)
            hb.addStretch()
            opw.setLayout(hb)
            tbl.setCellWidget(r, 8, opw)
        try:
            for r in range(tbl.rowCount()):
                for c in range(tbl.columnCount()):
                    it = tbl.item(r, c)
                    if it:
                        it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        except Exception:
            pass
        tbl.blockSignals(False)
    def add_exam(self):
        title = self.ex_title.text().strip()
        desc = self.ex_desc.toPlainText().strip()
        pass_ratio = self.ex_pass.value() / 100.0
        tl = self.ex_time.value()
        end = None if self.ex_permanent.isChecked() else self.ex_end.dateTime().toString(Qt.DateFormat.ISODate)
        if not title:
            QMessageBox.warning(self, tr('common.error'), tr('error.title_required'))
            return
        add_exam(title, desc, pass_ratio, tl, end, self.ex_random_count.value())
        self.refresh_exams()
        QMessageBox.information(self, tr('common.success'), tr('info.exam_added'))
    def get_selected_exam_id(self):
        tbl = getattr(self, 'exams_table', None)
        if tbl is None:
            return None
        r = tbl.currentRow()
        if r < 0:
            return None
        it = tbl.item(r, 0)
        return int(it.text()) if it and it.text() else None
    def import_questions(self):
        exam_id = self.get_selected_exam_id()
        if not exam_id:
            QMessageBox.warning(self, tr('common.error'), tr('error.select_exam'))
            return
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents')
        fn, sel = QFileDialog.getOpenFileName(self, tr('admin.import.title'), suggested, 'Excel (*.xlsx);;JSON (*.json);;YAML (*.yaml *.yml)')
        if not fn:
            return
        try:
            ext = os.path.splitext(fn)[1].lower().strip()
            is_zip = False
            try:
                with open(fn, 'rb') as f:
                    head = f.read(4)
                    is_zip = head.startswith(b'PK')
            except Exception:
                is_zip = False
            if (sel and sel.startswith('Excel')) or ext in ('.xlsx', '.xlsm', '.xltx', '.xltm') or is_zip:
                wb = load_workbook(fn)
                rand_count = None
                def parse_sheet(ws):
                    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                    header = [str(x).strip() if x else '' for x in header_row]
                    def idx(name):
                        try:
                            return header.index(name)
                        except Exception:
                            return -1
                    itype = idx('类型'); icontent = idx('内容'); icorrect = idx('正确答案'); iscore = idx('分数')
                    start_opts = None
                    for i, h in enumerate(header):
                        if h.startswith('选项'):
                            start_opts = i
                            break
                    base_cols = [x for x in (itype, icontent, icorrect, iscore) if x >= 0]
                    if min(itype, icontent, icorrect) < 0:
                        return []
                    if start_opts is None:
                        start_opts = (max(base_cols) + 1) if base_cols else 3
                    data_local = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        tval = (str(row[itype]).strip().lower() if row[itype] is not None else '')
                        qtype = None
                        if tval in ('单选', 'single'):
                            qtype = 'single'
                        elif tval in ('多选', 'multiple'):
                            qtype = 'multiple'
                        elif tval in ('判断', 'truefalse', '判断题'):
                            qtype = 'truefalse'
                        else:
                            continue
                        text = (str(row[icontent]).strip() if row[icontent] is not None else '')
                        if not text:
                            continue
                        correct_cell = (str(row[icorrect]).strip() if row[icorrect] is not None else '')
                        correct = []
                        if qtype == 'truefalse':
                            lc = correct_cell.lower()
                            if lc in ('true', 'false'):
                                correct = [True] if lc == 'true' else [False]
                            else:
                                continue
                        else:
                            parts = [p.strip().upper() for p in correct_cell.replace('，', ',').replace(';', ',').split(',') if p.strip()]
                            correct = parts[:1] if qtype == 'single' else parts
                        options = []
                        if qtype != 'truefalse':
                            letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
                            cidx = start_opts
                            key_index = 0
                            while cidx < len(row):
                                val = row[cidx]
                                if val is None or str(val).strip() == '':
                                    break
                                key = letters[key_index] if key_index < len(letters) else str(key_index+1)
                                options.append({'key': key, 'text': str(val).strip()})
                                cidx += 1
                                key_index += 1
                            if not options and qtype != 'truefalse':
                                continue
                        sc = 1.0
                        if iscore >= 0 and iscore < len(row):
                            try:
                                v = row[iscore]
                                if v is not None and str(v).strip() != '':
                                    sc = float(str(v).strip())
                            except Exception:
                                sc = 1.0
                        item = {'type': qtype, 'text': text, 'score': sc, 'options': options, 'correct': correct}
                        data_local.append(item)
                    return data_local
                data_mand = []
                data_rand = []
                if '配置选项' in wb.sheetnames:
                    ws_cfg = wb['配置选项']
                    cfg_header = next(ws_cfg.iter_rows(min_row=1, max_row=1, values_only=True))
                    cfg = {str(cfg_header[i]).strip(): (ws_cfg.cell(row=2, column=i+1).value) for i in range(len(cfg_header)) if cfg_header[i] is not None}
                    if '随机抽取数量' in cfg:
                        try:
                            rand_count = int(str(cfg['随机抽取数量']).strip())
                        except Exception:
                            rand_count = None
                if '必考题库' in wb.sheetnames:
                    data_mand = parse_sheet(wb['必考题库'])
                if '随机题库' in wb.sheetnames:
                    data_rand = parse_sheet(wb['随机题库'])
                if not data_mand and not data_rand:
                    QMessageBox.warning(self, tr('common.error'), '缺少新格式工作表：请提供“必考题库”或“随机题库”（至少之一），可选“配置选项”设置随机抽取数量')
                    return
                data = []
                for x in data_mand:
                    x['pool'] = 'mandatory'
                    data.append(x)
                for x in data_rand:
                    x['pool'] = 'random'
                    data.append(x)
                if rand_count is not None:
                    update_exam_random_pick_count(exam_id, rand_count)
            else:
                def read_text_with_fallback(path):
                    with open(path, 'rb') as f:
                        raw = f.read()
                    try:
                        return raw.decode('utf-8')
                    except UnicodeDecodeError:
                        try:
                            return raw.decode('gb18030')
                        except UnicodeDecodeError:
                            return None
                text = read_text_with_fallback(fn)
                if text is None:
                    QMessageBox.warning(self, tr('common.error'), tr('admin.import.error.file_decode'))
                    return
                data = None
                if (sel and sel.startswith('JSON')) or ext == '.json':
                    data = json.loads(text)
                elif (sel and sel.startswith('YAML')) or ext in ('.yaml', '.yml'):
                    data = yaml.safe_load(text)
                else:
                    try:
                        data = json.loads(text)
                    except Exception:
                        try:
                            data_yaml = yaml.safe_load(text)
                            data = data_yaml
                        except Exception:
                            QMessageBox.warning(self, tr('common.error'), tr('admin.import.error.not_supported'))
                            return
            if data is None:
                QMessageBox.warning(self, tr('common.error'), tr('admin.import.error.no_data'))
                return
            valid = []
            errs = []
            def validate_list(lst, pool_name):
                base_index = len(valid)
                for idx, q in enumerate(lst, start=1):
                    t = (q.get('type') or '').strip().lower()
                    if t not in ('single','multiple','truefalse'):
                        errs.append(f'{pool_name} 第{idx}题: 类型无效')
                        continue
                    corr = q.get('correct') or []
                    if t in ('single','multiple'):
                        opts = q.get('options') or []
                        keys = {str(o.get('key')).strip().upper() for o in opts if o.get('key')}
                        if not keys:
                            errs.append(f'{pool_name} 第{idx}题: 缺少选项')
                            continue
                        corr = [str(x).strip().upper() for x in corr if str(x).strip() != '']
                        if not corr or not set(corr).issubset(keys):
                            errs.append(f'{pool_name} 第{idx}题: 正确答案不在选项中')
                            continue
                        if t == 'single' and len(corr) != 1:
                            errs.append(f'{pool_name} 第{idx}题: 单选需1个答案')
                            continue
                        q['correct'] = corr
                    else:
                        if not corr or len(corr) != 1 or not isinstance(corr[0], bool):
                            errs.append(f'{pool_name} 第{idx}题: 判断题答案需为true/false')
                            continue
                    valid.append(q)
            if isinstance(data, dict):
                cfg = data.get('config') or {}
                if 'random_pick_count' in cfg:
                    try:
                        update_exam_random_pick_count(exam_id, int(cfg.get('random_pick_count') or 0))
                    except Exception:
                        pass
                mand = data.get('mandatory') or []
                rand = data.get('random') or []
                if not mand and not rand:
                    QMessageBox.warning(self, tr('common.error'), tr('admin.import.error.jsonyaml_missing'))
                    return
                for x in mand:
                    x['pool'] = 'mandatory'
                for x in rand:
                    x['pool'] = 'random'
                validate_list(mand, '必考题库')
                validate_list(rand, '随机题库')
            else:
                QMessageBox.warning(self, tr('common.error'), tr('admin.import.error.jsonyaml_dict'))
                return
            if not valid:
                detail = '\n'.join(errs[:20]) if errs else tr('admin.import.error.no_valid')
                QMessageBox.warning(self, tr('common.error'), detail)
                return
            import_questions_from_json(exam_id, valid)
            self.refresh_exams()
            cnt_single = sum(1 for d in valid if d.get('type') == 'single')
            cnt_multiple = sum(1 for d in valid if d.get('type') == 'multiple')
            cnt_tf = sum(1 for d in valid if d.get('type') == 'truefalse')
            cnt_mand = sum(1 for d in valid if (d.get('pool') or 'mandatory') == 'mandatory')
            cnt_rand = sum(1 for d in valid if (d.get('pool') or 'mandatory') == 'random')
            extra = ''
            if errs:
                extra = '\n部分题目未导入:\n' + '\n'.join(errs[:10])
            QMessageBox.information(self, tr('common.success'), tr('admin.import.success', single=cnt_single, multiple=cnt_multiple, truefalse=cnt_tf, mandatory=cnt_mand, random=cnt_rand, extra=(tr('admin.import.extra_prefix') + '\n'.join(errs[:10]) if errs else '')))
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))
    def clear_exam(self, exam_id):
        reply = QMessageBox.question(self, tr('common.hint'), tr('admin.exams.clear_confirm'), QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            from models import clear_exam_questions
            clear_exam_questions(exam_id)
            self.refresh_exams()
            QMessageBox.information(self, tr('common.success'), tr('admin.exams.clear_done'))
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))
    def delete_exam(self, exam_id):
        reply = QMessageBox.question(self, tr('common.hint'), tr('admin.exams.delete_confirm'), QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            from models import delete_exam
            delete_exam(exam_id)
            self.refresh_exams()
            QMessageBox.information(self, tr('common.success'), tr('admin.exams.delete_done'))
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))
    def export_sample(self):
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents/exam')
        fn, sel = QFileDialog.getSaveFileName(self, tr('admin.export.sample.title'), suggested, 'Excel (*.xlsx);;JSON (*.json);;YAML (*.yaml)')
        if not fn:
            return
        try:
            ext = os.path.splitext(fn)[1].lower()
            mand = [
                {"type":"single","text":"Python中获取列表长度的函数是?","options":[{"key":"A","text":"len(list)"},{"key":"B","text":"size(list)"},{"key":"C","text":"count(list)"},{"key":"D","text":"length(list)"}],"correct":["A"],"score":2},
                {"type":"multiple","text":"以下哪些是Linux常见包管理器?","options":[{"key":"A","text":"apt"},{"key":"B","text":"ls"},{"key":"C","text":"yum"},{"key":"D","text":"pacman"}],"correct":["A","C","D"],"score":3},
                {"type":"truefalse","text":"Python中的list是可变对象","correct":[True],"score":1},
                {"type":"single","text":"查看当前工作目录的Linux命令是?","options":[{"key":"A","text":"pwd"},{"key":"B","text":"cd"},{"key":"C","text":"ls"},{"key":"D","text":"echo"}],"correct":["A"],"score":2},
                {"type":"multiple","text":"以下哪些工具可用于创建Python虚拟环境?","options":[{"key":"A","text":"venv"},{"key":"B","text":"virtualenv"},{"key":"C","text":"pip"},{"key":"D","text":"conda"}],"correct":["A","B","D"],"score":3}
            ]
            rand = [
                {"type":"truefalse","text":"Linux中/etc目录通常存放系统配置文件","correct":[True],"score":1},
                {"type":"multiple","text":"以下哪些是Python中的可迭代对象?","options":[{"key":"A","text":"list"},{"key":"B","text":"dict"},{"key":"C","text":"int"},{"key":"D","text":"tuple"}],"correct":["A","B","D"],"score":3},
                {"type":"single","text":"Python字典取值且键不存在时不抛异常的方法是?","options":[{"key":"A","text":"d['k']"},{"key":"B","text":"d.get('k')"},{"key":"C","text":"d.k"},{"key":"D","text":"getattr(d,'k')"}],"correct":["B"],"score":2},
                {"type":"single","text":"Linux查看网络端口占用的命令是?","options":[{"key":"A","text":"ss -tuln"},{"key":"B","text":"ps aux"},{"key":"C","text":"top"},{"key":"D","text":"df -h"}],"correct":["A"],"score":2},
                {"type":"multiple","text":"以下哪些属于Python打包/分发相关工具?","options":[{"key":"A","text":"setuptools"},{"key":"B","text":"wheel"},{"key":"C","text":"pip"},{"key":"D","text":"twine"}],"correct":["A","B","D"],"score":3}
            ]
            if (sel and sel.startswith('Excel')) or ext == '.xlsx' or ext == '':
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter
                out = fn if ext == '.xlsx' else fn + '.xlsx'
                wb = Workbook()
                ws_cfg = wb.active
                ws_cfg.title = '配置选项'
                ws_cfg.append(['随机抽取数量'])
                ws_cfg.append([4])
                def write_sheet(ws, rows):
                    ws.append(['类型', '内容', '正确答案', '分数', '选项A', '选项B', '选项C', '选项D'])
                    for item in rows:
                        if item['type'] == 'truefalse':
                            ws.append(['判断', item['text'], 'true' if item['correct'][0] else 'false', item['score']])
                        elif item['type'] == 'single':
                            ws.append(['单选', item['text'], ','.join(item['correct']), item['score']] + [opt['text'] for opt in item.get('options', [])])
                        else:
                            ws.append(['多选', item['text'], ','.join(item['correct']), item['score']] + [opt['text'] for opt in item.get('options', [])])
                ws_m = wb.create_sheet('必考题库')
                write_sheet(ws_m, mand)
                ws_r = wb.create_sheet('随机题库')
                write_sheet(ws_r, rand)
                header_fill = PatternFill(start_color='FF409EFF', end_color='FF409EFF', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFFFF')
                center = Alignment(horizontal='center', vertical='center')
                left = Alignment(horizontal='left', vertical='center')
                thin = Side(style='thin', color='FFDDDDDD')
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                for ws in [ws_m, ws_r]:
                    headers = ['类型', '内容', '正确答案', '分数', '选项A', '选项B', '选项C', '选项D']
                    for c in range(1, len(headers)+1):
                        cell = ws.cell(row=1, column=c)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center
                    for r in range(2, ws.max_row+1):
                        for c in range(1, len(headers)+1):
                            ws.cell(row=r, column=c).border = border
                        ws.cell(row=r, column=4).alignment = center
                        for c in (1,2,3,5,6,7,8):
                            ws.cell(row=r, column=c).alignment = left
                    widths = [0] * len(headers)
                    for r in ws.iter_rows(values_only=True):
                        for idx, val in enumerate(r):
                            l = len(str(val)) if val is not None else 0
                            widths[idx] = max(widths[idx], l)
                    for i, w in enumerate(widths, start=1):
                        letter = get_column_letter(i)
                        ws.column_dimensions[letter].width = max(12, min(36, w + 2))
                    ws.freeze_panes = 'A2'
                    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
                wb.save(out)
            elif (sel and sel.startswith('JSON')) or ext == '.json':
                out = fn if ext == '.json' else fn + '.json'
                payload = {"config": {"random_pick_count": 4}, "mandatory": mand, "random": rand}
                with open(out, 'w', encoding='utf-8') as f:
                    json.dump(payload, f, ensure_ascii=False, indent=2)
            elif (sel and sel.startswith('YAML')) or ext in ('.yaml', '.yml'):
                out = fn if ext in ('.yaml', '.yml') else fn + '.yaml'
                lines = []
                lines.append('config:')
                lines.append('  random_pick_count: 4')
                def write_yaml_block(name, rows):
                    lines.append(f'{name}:')
                    for item in rows:
                        lines.append('  -')
                        lines.append(f'    type: {item["type"]}')
                        lines.append(f'    text: "{item["text"]}"')
                        lines.append(f'    score: {item["score"]}')
                        if item.get('options'):
                            lines.append('    options:')
                            for opt in item['options']:
                                lines.append('      - key: ' + opt['key'])
                                lines.append(f'        text: "{opt["text"]}"')
                        if item.get('correct') is not None:
                            lines.append('    correct:')
                            for v in item['correct']:
                                if isinstance(v, bool):
                                    lines.append('      - ' + ('true' if v else 'false'))
                                else:
                                    lines.append('      - ' + str(v))
                write_yaml_block('mandatory', mand)
                write_yaml_block('random', rand)
                with open(out, 'w', encoding='utf-8') as f:
                    f.write('\n'.join(lines) + '\n')
            QMessageBox.information(self, tr('common.success'), tr('admin.export.sample.done'))
        except ImportError:
            QMessageBox.warning(self, '错误', '需要安装openpyxl: pip install openpyxl')
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))
    def sync_tab(self):
        w = QWidget()
        lay = QVBoxLayout()
        gb1 = QGroupBox(tr('admin.targets.group'))
        vb1 = QVBoxLayout()
        self.targets_table = QTableWidget(0, 6)
        self.targets_table.setHorizontalHeaderLabels([tr('admin.targets.headers.name'), tr('admin.targets.headers.ip'), tr('admin.targets.headers.username'), tr('admin.targets.headers.remote_path'), tr('admin.targets.headers.ssh_password'), tr('admin.users.headers.actions')])
        self.targets_table.setColumnWidth(0, 150)
        self.targets_table.setColumnWidth(1, 150)
        self.targets_table.setColumnWidth(2, 150)
        self.targets_table.setColumnWidth(3, 300)
        self.targets_table.setColumnWidth(4, 150)
        self.targets_table.horizontalHeader().setStretchLastSection(True)
        self.refresh_targets()
        vb1.addWidget(self.targets_table)
        gb1.setLayout(vb1)
        lay.addWidget(gb1)
        gb2 = QGroupBox(tr('admin.targets.add_group'))
        form = QFormLayout()
        self.t_name = QLineEdit()
        self.t_name.setPlaceholderText(tr('admin.targets.name_ph'))
        try:
            self.t_name.setValidator(None)
        except Exception:
            pass
        self.t_ip = QLineEdit()
        self.t_ip.setPlaceholderText(tr('admin.targets.ip_ph'))
        self.t_user = QLineEdit()
        self.t_user.setPlaceholderText(tr('admin.targets.username_ph'))
        self.t_path = QLineEdit()
        self.t_path.setText(tr('admin.targets.remote_path_ph'))
        self.t_password = QLineEdit()
        self.t_password.setPlaceholderText(tr('admin.targets.ssh_password_ph'))
        self.t_password.setEchoMode(QLineEdit.EchoMode.Password)
        add_btn = QPushButton(tr('admin.targets.add_btn'))
        add_btn.clicked.connect(self.add_target)
        form.addRow('名称', self.t_name)
        form.addRow('IP', self.t_ip)
        form.addRow('用户名', self.t_user)
        form.addRow('远程路径', self.t_path)
        form.addRow('SSH密码', self.t_password)
        form.addRow(add_btn)
        gb2.setLayout(form)
        lay.addWidget(gb2)
        hb_tpl = QHBoxLayout()
        btn_export_targets_tpl = QPushButton(tr('admin.export.targets_tpl.title'))
        btn_export_targets_tpl.setIcon(get_icon('exam_export'))
        btn_export_targets_tpl.clicked.connect(self.export_targets_template)
        btn_import_targets_excel = QPushButton(tr('admin.import.targets.title'))
        btn_import_targets_excel.setIcon(get_icon('exam_import'))
        btn_import_targets_excel.clicked.connect(self.import_targets_from_excel)
        hb_tpl.addWidget(btn_export_targets_tpl)
        hb_tpl.addWidget(btn_import_targets_excel)
        lay.addLayout(hb_tpl)
        hb = QHBoxLayout()
        self.push_btn = QPushButton(tr('sync.push_btn'))
        self.push_btn.setIcon(get_icon('push'))
        self.push_btn.clicked.connect(self.push_all)
        self.pull_btn = QPushButton(tr('sync.pull_btn'))
        self.pull_btn.setIcon(get_icon('pull'))
        self.pull_btn.clicked.connect(self.pull_all)
        hb.addWidget(self.push_btn)
        hb.addWidget(self.pull_btn)
        lay.addLayout(hb)
        self.sync_spinner = LoadingIndicator(self)
        self.sync_spinner.hide()
        lay.addWidget(self.sync_spinner)
        colors_log = theme_manager.get_theme_colors()
        self.sync_log = QTextBrowser()
        self.sync_log.setReadOnly(True)
        self.sync_log.setMinimumHeight(140)
        self.sync_log.setStyleSheet(
            f"QTextBrowser {{ background-color:{colors_log['card_background']}; border:1px solid {colors_log['border']}; border-radius:8px; padding:8px; color:{colors_log['text_primary']}; }}"
        )
        lay.addWidget(self.sync_log)
        lay.addStretch()
        w.setLayout(lay)
        return w
    def export_users_template(self):
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents/users_template')
        fn, sel = QFileDialog.getSaveFileName(self, '导出用户Excel模板', suggested, 'Excel (*.xlsx)')
        if not fn:
            return
        headers = ['用户名', '密码', '姓名', '角色', '状态']
        try:
            from openpyxl import Workbook
            ext = os.path.splitext(fn)[1].lower()
            out = fn if ext == '.xlsx' else fn + '.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = 'Users'
            ws.append(headers)
            ws.append(['', '', '', 'user/admin', '1或0'])
            wb.save(out)
            QMessageBox.information(self, '成功', '用户模板已导出')
        except ImportError:
            QMessageBox.warning(self, '错误', '需要安装openpyxl: pip install openpyxl')
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))
    def import_users_from_excel(self):
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents')
        fn, sel = QFileDialog.getOpenFileName(self, '选择用户Excel', suggested, 'Excel (*.xlsx)')
        if not fn:
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(fn)
            ws = wb['Users'] if 'Users' in wb.sheetnames else wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            header = [str(x).strip() if x else '' for x in header_row]
            def idx(name):
                try:
                    return header.index(name)
                except Exception:
                    return -1
            iu = idx('用户名'); ip = idx('密码'); iname = idx('姓名'); ir = idx('角色'); ia = idx('状态')
            if min(iu, ip, ir, ia) < 0:
                QMessageBox.warning(self, '错误', '缺少必要列: 用户名/密码/角色/状态')
                return
            ok = 0; fail = 0
            format_errs = []
            for idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    username = (str(r[iu]).strip() if iu >= 0 and iu < len(r) and r[iu] is not None else '')
                    password = (str(r[ip]).strip() if ip >= 0 and ip < len(r) and r[ip] is not None else '')
                    full_name = (str(r[iname]).strip() if iname >= 0 and iname < len(r) and r[iname] is not None else '') or None
                    role = (str(r[ir]).strip().lower() if ir >= 0 and ir < len(r) and r[ir] is not None else 'user')
                    active_str = (str(r[ia]).strip() if ia >= 0 and ia < len(r) and r[ia] is not None else '1')
                    active = 1 if active_str in ('1', '是', '启用', 'true', 'True') else 0
                    if not username or not password:
                        fail += 1
                        format_errs.append(f'第{idx}行：用户名或密码为空')
                        continue
                    if not re.fullmatch(r"[A-Za-z0-9_@.\-]+", username):
                        fail += 1
                        format_errs.append(f'第{idx}行：用户名格式错误')
                        continue
                    if not re.fullmatch(r"[\x20-\x7E]+", password):
                        fail += 1
                        format_errs.append(f'第{idx}行：密码格式错误')
                        continue
                    if role not in ('user', 'admin'):
                        role = 'user'
                    create_user(username, password, role, active, full_name)
                    ok += 1
                except Exception:
                    fail += 1
            self.refresh_users()
            QMessageBox.information(self, '完成', f'导入成功:{ok} 失败:{fail}')
            if format_errs:
                detail = '\n'.join(format_errs[:20])
                QMessageBox.warning(self, '格式错误', detail)
        except ImportError:
            QMessageBox.warning(self, '错误', '需要安装openpyxl: pip install openpyxl')
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))
    def export_targets_template(self):
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents/targets_template')
        fn, sel = QFileDialog.getSaveFileName(self, '导出设备Excel模板', suggested, 'Excel (*.xlsx)')
        if not fn:
            return
        headers = ['名称', 'IP', '用户名', '远程路径', 'SSH密码']
        try:
            from openpyxl import Workbook
            ext = os.path.splitext(fn)[1].lower()
            out = fn if ext == '.xlsx' else fn + '.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = 'Targets'
            ws.append(headers)
            ws.append(['设备A', '192.168.1.10', 'user', '~/.exam_system/exam.db', ''])
            wb.save(out)
            QMessageBox.information(self, '成功', '设备模板已导出')
        except ImportError:
            QMessageBox.warning(self, '错误', '需要安装openpyxl: pip install openpyxl')
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))
    def import_targets_from_excel(self):
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents')
        fn, sel = QFileDialog.getOpenFileName(self, '选择设备Excel', suggested, 'Excel (*.xlsx)')
        if not fn:
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(fn)
            ws = wb['Targets'] if 'Targets' in wb.sheetnames else wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            header = [str(x).strip() if x else '' for x in header_row]
            def idx(name):
                try:
                    return header.index(name)
                except Exception:
                    return -1
            iname = idx('名称'); iip = idx('IP'); iuser = idx('用户名'); ipath = idx('远程路径'); ipwd = idx('SSH密码')
            if min(iname, iip, iuser, ipath) < 0:
                QMessageBox.warning(self, '错误', '缺少必要列: 名称/IP/用户名/远程路径')
                return
            ok = 0; fail = 0
            for r in ws.iter_rows(min_row=2, values_only=True):
                try:
                    name = (str(r[iname]).strip() if iname >= 0 and iname < len(r) and r[iname] is not None else '')
                    ip = (str(r[iip]).strip() if iip >= 0 and iip < len(r) and r[iip] is not None else '')
                    user = (str(r[iuser]).strip() if iuser >= 0 and iuser < len(r) and r[iuser] is not None else '')
                    path = (str(r[ipath]).strip() if ipath >= 0 and ipath < len(r) and r[ipath] is not None else '')
                    password = (str(r[ipwd]).strip() if ipwd >= 0 and ipwd < len(r) and r[ipwd] is not None else '') or None
                    if not name or not ip or not user or not path:
                        fail += 1
                        continue
                    upsert_sync_target(name, ip, user, path, password)
                    ok += 1
                except Exception:
                    fail += 1
            self.refresh_targets()
            QMessageBox.information(self, '完成', f'导入成功:{ok} 失败:{fail}')
        except ImportError:
            QMessageBox.warning(self, '错误', '需要安装openpyxl: pip install openpyxl')
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))
    def refresh_targets(self):
        targets = list_sync_targets()
        self.targets_table.blockSignals(True)
        self.targets_table.setRowCount(0)
        for t in targets:
            r = self.targets_table.rowCount()
            self.targets_table.insertRow(r)
            it_name = QTableWidgetItem(t[1])
            it_name.setData(Qt.ItemDataRole.UserRole, t[0])
            self.targets_table.setItem(r, 0, it_name)
            self.targets_table.setItem(r, 1, QTableWidgetItem(t[2]))
            self.targets_table.setItem(r, 2, QTableWidgetItem(t[3]))
            self.targets_table.setItem(r, 3, QTableWidgetItem(t[4]))
            password_text = '******' if t[5] else ''
            self.targets_table.setItem(r, 4, QTableWidgetItem(password_text))
            action_widget = QWidget()
            action_layout = QHBoxLayout()
            action_layout.setContentsMargins(4, 4, 4, 4)
            edit_btn = QPushButton('编辑')
            edit_btn.setStyleSheet("QPushButton { background-color:#409eff; color:#fff; padding:4px 8px; font-size:12px; border-radius:6px; }")
            edit_btn.clicked.connect(lambda checked, tid=t[0]: self.edit_target(tid))
            action_layout.addWidget(edit_btn)
            delete_btn = QPushButton('删除')
            delete_btn.setStyleSheet("QPushButton { background-color:#f56c6c; color:#fff; padding:4px 8px; font-size:12px; }")
            delete_btn.clicked.connect(lambda checked, tid=t[0]: self.delete_target(tid))
            action_layout.addWidget(delete_btn)
            action_widget.setLayout(action_layout)
            self.targets_table.setCellWidget(r, 5, action_widget)
        try:
            for r in range(self.targets_table.rowCount()):
                for c in range(self.targets_table.columnCount()):
                    it = self.targets_table.item(r, c)
                    if it:
                        it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        except Exception:
            pass
        self.targets_table.blockSignals(False)
        try:
            self.targets_table.itemChanged.disconnect()
        except Exception:
            pass
        self.targets_table.itemChanged.connect(self.on_target_item_changed)
    def on_user_item_changed(self, item):
        row = item.row()
        col = item.column()
        try:
            user_id = int(self.users_table.item(row, 0).text())
        except Exception:
            return
        from models import update_user_basic
        if col == 1:
            username = item.text().strip()
            if not username:
                QMessageBox.warning(self, '错误', '用户名不能为空')
                self.refresh_users()
                return
            update_user_basic(user_id, username=username)
            self.refresh_users()
        elif col == 2:
            full_name = item.text().strip() or None
            update_user_basic(user_id, full_name=full_name)
            self.refresh_users()
    def on_exam_item_changed(self, item):
        row = item.row()
        col = item.column()
        try:
            exam_id = int(self.exams_table.item(row, 0).text())
        except Exception:
            return
        from models import update_exam_title_desc
        if col == 1:
            title = item.text().strip()
            if not title:
                QMessageBox.warning(self, '错误', '标题不能为空')
                self.refresh_exams()
                return
            update_exam_title_desc(exam_id, title=title)
            self.refresh_exams()
        elif col == 5:
            desc = item.text().strip()
            update_exam_title_desc(exam_id, description=desc)
            self.refresh_exams()
    def on_target_item_changed(self, item):
        row = item.row()
        col = item.column()
        if col not in (0,1,2,3):
            return
        id_item = self.targets_table.item(row, 0)
        target_id = id_item.data(Qt.ItemDataRole.UserRole) if id_item else None
        if target_id is None:
            return
        name = self.targets_table.item(row, 0).text() if self.targets_table.item(row, 0) else ''
        ip = self.targets_table.item(row, 1).text() if self.targets_table.item(row, 1) else ''
        username = self.targets_table.item(row, 2).text() if self.targets_table.item(row, 2) else ''
        remote_path = self.targets_table.item(row, 3).text() if self.targets_table.item(row, 3) else ''
        if not name or not ip or not username or not remote_path:
            QMessageBox.warning(self, '错误', '设备信息不能为空')
            self.refresh_targets()
            return
        try:
            update_sync_target(int(target_id), name, ip, username, remote_path, None)
            self.refresh_targets()
            QMessageBox.information(self, '成功', '设备信息已更新')
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))
    def add_target(self):
        name = self.t_name.text().strip()
        ip = self.t_ip.text().strip()
        user = self.t_user.text().strip()
        path = self.t_path.text().strip()
        password = self.t_password.text().strip()
        if not name or not ip or not user or not path:
            QMessageBox.warning(self, '错误', '请完整填写设备信息')
            return
        upsert_sync_target(name, ip, user, path, password if password else None)
        self.refresh_targets()
        self.t_name.clear()
        self.t_ip.clear()
        self.t_user.clear()
        self.t_path.clear()
        self.t_password.clear()
        QMessageBox.information(self, '成功', '设备已添加')
    def edit_target(self, target_id):
        targets = list_sync_targets()
        target = None
        for t in targets:
            if t[0] == target_id:
                target = t
                break
        if not target:
            return
        from PySide6.QtWidgets import QDialog, QDialogButtonBox
        dialog = QDialog(self)
        dialog.setWindowTitle('编辑设备')
        dialog.setFixedSize(400, 300)
        layout = QFormLayout()
        name_edit = QLineEdit(target[1])
        ip_edit = QLineEdit(target[2])
        user_edit = QLineEdit(target[3])
        path_edit = QLineEdit(target[4])
        password_edit = QLineEdit()
        password_edit.setPlaceholderText('留空保持原密码')
        password_edit.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addRow('名称', name_edit)
        layout.addRow('IP', ip_edit)
        layout.addRow('用户名', user_edit)
        layout.addRow('远程路径', path_edit)
        layout.addRow('SSH密码', password_edit)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        layout.addWidget(buttons)
        dialog.setLayout(layout)
        def save_changes():
            new_name = name_edit.text().strip()
            new_ip = ip_edit.text().strip()
            new_user = user_edit.text().strip()
            new_path = path_edit.text().strip()
            new_password = password_edit.text().strip()
            if not new_name or not new_ip or not new_user or not new_path:
                QMessageBox.warning(dialog, '错误', '请完整填写设备信息')
                return
            try:
                update_sync_target(target_id, new_name, new_ip, new_user, new_path, new_password if new_password else None)
                self.refresh_targets()
                dialog.accept()
                QMessageBox.information(self, '成功', '设备已更新')
            except Exception as e:
                QMessageBox.warning(dialog, '错误', str(e))
        buttons.accepted.connect(save_changes)
        buttons.rejected.connect(dialog.reject)
        dialog.exec()
    def delete_target(self, target_id):
        reply = QMessageBox.question(self, '确认', '确定要删除该设备吗？', QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            try:
                delete_sync_target(target_id)
                self.refresh_targets()
                QMessageBox.information(self, '成功', '设备已删除')
            except Exception as e:
                QMessageBox.warning(self, '错误', str(e))
    def push_all(self):
        targets = list_sync_targets()
        if not targets:
            QMessageBox.information(self, '提示', '没有配置任何设备')
            return
        self.set_sync_buttons_enabled(False)
        if hasattr(self, 'push_btn'):
            self.push_btn.setEnabled(False)
        if hasattr(self, 'pull_btn'):
            self.pull_btn.setEnabled(False)
        if hasattr(self, 'sync_spinner'):
            self.sync_spinner.show()
        if hasattr(self, 'sync_log'):
            self.sync_log.clear()
        self.sync_worker = SyncWorker(targets, 'push')
        self.sync_worker.finished.connect(self.on_sync_finished)
        self.sync_worker.error.connect(self.on_sync_error)
        self.sync_worker.progress.connect(self.append_sync_log)
        self.sync_worker.progress.connect(self.update_progress_message)
        self.sync_worker.start()
        self.show_sync_progress('正在同步题库到设备，请稍候...')
    def pull_all(self):
        targets = list_sync_targets()
        if not targets:
            QMessageBox.information(self, '提示', '没有配置任何设备')
            return
        self.set_sync_buttons_enabled(False)
        if hasattr(self, 'push_btn'):
            self.push_btn.setEnabled(False)
        if hasattr(self, 'pull_btn'):
            self.pull_btn.setEnabled(False)
        if hasattr(self, 'sync_spinner'):
            self.sync_spinner.show()
        if hasattr(self, 'sync_log'):
            self.sync_log.clear()
        self.sync_worker = SyncWorker(targets, 'pull')
        self.sync_worker.finished.connect(self.on_sync_finished)
        self.sync_worker.error.connect(self.on_sync_error)
        self.sync_worker.progress.connect(self.append_sync_log)
        self.sync_worker.progress.connect(self.update_progress_message)
        self.sync_worker.start()
        self.show_sync_progress('正在拉取成绩，请稍候...')

    def update_progress_message(self, msg):
        try:
            if hasattr(self, 'sync_progress_dialog') and getattr(self, 'sync_progress_dialog'):
                self.sync_progress_dialog.setLabelText(msg)
        except Exception:
            pass

    def append_sync_log(self, message):
        colors = theme_manager.get_theme_colors()
        fg_ok = '#67c23a'
        bg_ok = '#e1f3d8'
        fg_err = '#f56c6c'
        bg_err = '#fde2e2'
        fg_info = colors['text_secondary']
        bg_info = colors['border_light']
        status = '成功' if ('成功' in message and '失败' not in message) else ('失败' if '失败' in message else '信息')
        if status == '成功':
            badge_fg, badge_bg = fg_ok, bg_ok
        elif status == '失败':
            badge_fg, badge_bg = fg_err, bg_err
        else:
            badge_fg, badge_bg = fg_info, bg_info
        html = f"<div style=\"margin:4px 0;\">"
        html += f"<span style=\"background:{badge_bg}; color:{badge_fg}; border-radius:10px; padding:2px 8px; font-size:12px;\">{status}</span>"
        html += f"<span style=\"margin-left:8px;\">{message}</span>"
        html += "</div>"
        self.sync_log.append(html)
    def set_sync_buttons_enabled(self, enabled):
        for i in range(self.targets_table.rowCount()):
            widget = self.targets_table.cellWidget(i, 5)
            if widget:
                for child in widget.findChildren(QPushButton):
                    child.setEnabled(enabled)
        if hasattr(self, 'push_btn'):
            self.push_btn.setEnabled(enabled)
        if hasattr(self, 'pull_btn'):
            self.pull_btn.setEnabled(enabled)
    def on_sync_finished(self, results):
        self.set_sync_buttons_enabled(True)
        if hasattr(self, 'sync_worker'):
            self.sync_worker.deleteLater()
        if hasattr(self, 'sync_spinner'):
            self.sync_spinner.hide()
        if hasattr(self, 'sync_progress_dialog') and getattr(self, 'sync_progress_dialog'):
            try:
                self.sync_progress_dialog.close()
            except Exception:
                pass
            self.sync_progress_dialog = None
        if '拉取' in results:
            self.refresh_scores()
        QMessageBox.information(self, tr('sync.finished.title'), tr('sync.operation_done', results=results))
    def on_sync_error(self, error):
        self.set_sync_buttons_enabled(True)
        if hasattr(self, 'sync_worker'):
            self.sync_worker.deleteLater()
        if hasattr(self, 'sync_spinner'):
            self.sync_spinner.hide()
        if hasattr(self, 'sync_progress_dialog') and getattr(self, 'sync_progress_dialog'):
            try:
                self.sync_progress_dialog.close()
            except Exception:
                pass
            self.sync_progress_dialog = None
        QMessageBox.warning(self, tr('sync.error.title'), tr('sync.error.message', error=error))
    def scores_tab(self):
        w = QWidget()
        lay = QVBoxLayout()
        gb = QGroupBox(tr('scores.group'))
        vb = QVBoxLayout()
        self.scores_table = QTableWidget(0, 8)
        self.scores_table.setHorizontalHeaderLabels([tr('scores.headers.uuid'), tr('scores.headers.username'), tr('scores.headers.full_name'), tr('scores.headers.user_id'), tr('scores.headers.exam_title'), tr('scores.headers.started'), tr('scores.headers.submitted'), tr('scores.headers.score_total_pass')])
        self.scores_table.horizontalHeader().setStretchLastSection(True)
        self.scores_table.setColumnWidth(0, 280)
        self.scores_table.setColumnWidth(1, 75)
        self.scores_table.setColumnWidth(2, 75)
        self.scores_table.setColumnWidth(3, 50)
        self.scores_table.setColumnWidth(4, 250)
        self.scores_table.setColumnWidth(5, 200)
        self.scores_table.setColumnWidth(6, 200)
        self.scores_table.setAlternatingRowColors(True)
        self.refresh_scores()
        vb.addWidget(self.scores_table)
        hb = QHBoxLayout()
        btn_export_scores = QPushButton(tr('scores.export_excel'))
        btn_export_scores.setIcon(get_icon('exam_export'))
        btn_export_scores.clicked.connect(self.export_scores_to_excel)
        hb.addWidget(btn_export_scores)
        vb.addLayout(hb)
        gb.setLayout(vb)
        lay.addWidget(gb)
        w.setLayout(lay)
        return w
    def refresh_scores(self):
        self.scores_table = getattr(self, 'scores_table', QTableWidget(0, 8))
        self.scores_table.setRowCount(0)
        from models import list_attempts_with_user
        for a in list_attempts_with_user():
            r = self.scores_table.rowCount()
            self.scores_table.insertRow(r)
            self.scores_table.setItem(r, 0, QTableWidgetItem(a[0]))
            self.scores_table.setItem(r, 1, QTableWidgetItem(a[1] or ''))
            self.scores_table.setItem(r, 2, QTableWidgetItem(a[2] or ''))
            self.scores_table.setItem(r, 3, QTableWidgetItem(str(a[3])))
            exam_title = get_exam_title(int(a[4])) if a[4] is not None else ''
            self.scores_table.setItem(r, 4, QTableWidgetItem(exam_title or ''))
            self.scores_table.setItem(r, 5, QTableWidgetItem(a[5] or ''))
            self.scores_table.setItem(r, 6, QTableWidgetItem(a[6] or tr('scores.not_submitted')))
            is_valid = (len(a) > 9 and a[9] == 1)
            passed_text = '数据异常' if not is_valid else ('通过' if a[8] == 1 else '未通过')
            badge_bg = '#fff3cd' if not is_valid else ('#e1f3d8' if a[8] == 1 else '#fde2e2')
            badge_fg = '#8a6d3b' if not is_valid else ('#67c23a' if a[8] == 1 else '#f56c6c')
            stats = get_exam_stats(int(a[4])) if a[4] is not None else {'total_score': 0}
            total = int(stats['total_score']) if stats and stats.get('total_score') is not None else 0
            self.scores_table.setCellWidget(r, 7, self.make_tag(f'{a[7]} / {total} / {passed_text}', badge_bg, badge_fg))
            for c in (3,4):
                it = self.scores_table.item(r, c)
                if it:
                    it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        try:
            for r in range(self.scores_table.rowCount()):
                for c in range(self.scores_table.columnCount()):
                    it = self.scores_table.item(r, c)
                    if it:
                        it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        except Exception:
            pass
    def export_scores_to_excel(self):
        suggested = os.path.join(str(pathlib.Path.home()), 'Documents/scores')
        fn, sel = QFileDialog.getSaveFileName(self, '导出成绩Excel', suggested, 'Excel (*.xlsx)')
        if not fn:
            return
        try:
            from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
            from openpyxl.utils import get_column_letter
            ext = os.path.splitext(fn)[1].lower()
            out = fn if ext == '.xlsx' else fn + '.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = 'Scores'
            headers = ['尝试UUID', '用户名', '姓名', '用户ID', '试题标题', '开始', '提交', '分数', '通过']
            ws.append(headers)
            from models import list_attempts_with_user
            green_fill = PatternFill(start_color='FF67C23A', end_color='FF67C23A', fill_type='solid')
            red_fill = PatternFill(start_color='FFF56C6C', end_color='FFF56C6C', fill_type='solid')
            header_fill = PatternFill(start_color='FF409EFF', end_color='FF409EFF', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFFFF')
            center = Alignment(horizontal='center', vertical='center')
            left = Alignment(horizontal='left', vertical='center')
            thin = Side(style='thin', color='FFDDDDDD')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for a in list_attempts_with_user():
                exam_title = get_exam_title(int(a[4])) if a[4] is not None else ''
                is_valid = (len(a) > 9 and a[9] == 1)
                text_pass = '数据异常' if not is_valid else ('通过' if a[8] == 1 else '未通过')
                ws.append([a[0], a[1] or '', a[2] or '', int(a[3]), exam_title or '', a[5] or '', a[6] or '', a[7], text_pass])
                cell = ws.cell(row=ws.max_row, column=9)
                if not is_valid:
                    from openpyxl.styles import PatternFill
                    cell.fill = PatternFill(start_color='FFFFF3CD', end_color='FFFFF3CD', fill_type='solid')
                else:
                    cell.fill = green_fill if a[8] == 1 else red_fill
            for c in range(1, len(headers)+1):
                cell = ws.cell(row=1, column=c)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
            for r in range(2, ws.max_row+1):
                for c in range(1, len(headers)+1):
                    ws.cell(row=r, column=c).border = border
                ws.cell(row=r, column=4).alignment = center
                ws.cell(row=r, column=6).alignment = center
                ws.cell(row=r, column=7).alignment = center
                ws.cell(row=r, column=8).alignment = center
                ws.cell(row=r, column=9).alignment = center
                ws.cell(row=r, column=1).alignment = left
                ws.cell(row=r, column=2).alignment = left
                ws.cell(row=r, column=3).alignment = left
                ws.cell(row=r, column=5).alignment = left
            widths = [0] * len(headers)
            for r in ws.iter_rows(values_only=True):
                for idx, val in enumerate(r):
                    l = len(str(val)) if val is not None else 0
                    widths[idx] = max(widths[idx], l)
            for i, w in enumerate(widths, start=1):
                letter = get_column_letter(i)
                ws.column_dimensions[letter].width = max(12, min(36, w + 2))
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
            wb.save(out)
            QMessageBox.information(self, tr('common.success'), tr('export.scores.done'))
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))
